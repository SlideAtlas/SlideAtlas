"""
기관 포털 P2(구독 플랜) pytest — 읽기 전용.

CLAUDE.md §9(멀티테넌시)·§0(단일 진실)·§16(구독 모델)·§8(슬라이드 게이트).
내부 QA 체크리스트(이 세션 CEO 확정):
  (a) 스코프 격리 — g.institution_id 강제, inst_id 를 body/쿼리로 받지 않음(IDOR 불가)
  (b) /viewer 게이트 우회 없음 — 슬라이드 목록은 메타데이터 카탈로그(타일·토큰 발급 없음)
  (c) 내보내기 수식주입 방어 — XLSX/CSV 모두 _xlsx_safe
  (+) export 포함 모든 slides 경로 과목 격리 — 비구독 subject_code 는 빈 목록 아닌 403

DB 는 mock(로컬 RDS 접속 불가).
"""
import io
import os
from datetime import date
from unittest.mock import MagicMock, patch

os.environ.setdefault("JWT_SECRET_KEY", "test-secret-key-for-pytest")
os.environ.setdefault("GMAIL_USER", "test@gmail.com")
os.environ.setdefault("GMAIL_APP_PW", "test-app-pw")
os.environ.setdefault("ADMIN_SECRET_KEY", "test-admin-secret-for-pytest")

import pytest
import server_render as sr
from server_render import app
from auth.decorators import ADMIN_ROSTER_SUBJECT


def _norm(sql):
    return " ".join(str(sql).split()).lower()


@pytest.fixture
def client():
    app.config["TESTING"] = True
    with app.test_client() as c:
        yield c


def _fake_auth(uid="5", inst="CNU", role="admin", subject="HST"):
    def f():
        from flask import g
        g.user_id = uid
        g.institution_id = inst
        g.role = role
        g.subject_code = subject
        return None
    return f


def _mock_db():
    mock_conn = MagicMock()
    mock_cur = MagicMock()
    mock_conn.cursor.return_value.__enter__.return_value = mock_cur
    return mock_conn, mock_cur


# ─────────────────────────────────────────────
# 인증·관리자 게이트 (세 엔드포인트 공통)
# ─────────────────────────────────────────────
@pytest.mark.parametrize("path", [
    "/portal/api/plans",
    "/portal/api/plans/slides?subject_code=HST",
    "/portal/api/plans/slides/export?subject_code=HST&format=csv",
])
def test_plans_requires_auth(client, path):
    """비로그인 → 401(login_required)."""
    assert client.get(path).status_code == 401


@pytest.mark.parametrize("path", [
    "/portal/api/plans",
    "/portal/api/plans/slides?subject_code=HST",
    "/portal/api/plans/slides/export?subject_code=HST&format=xlsx",
])
def test_plans_forbidden_non_admin(client, path):
    """로그인했으나 __ADMIN__ roster 행 없음 → 403 FORBIDDEN(role 단독 우회 불가)."""
    with patch("auth.decorators._authenticate", _fake_auth(role="viewer")), \
         patch("server_render._is_institution_admin", return_value=False):
        resp = client.get(path)
    assert resp.status_code == 403
    assert resp.get_json()["error"] == "FORBIDDEN"


# ─────────────────────────────────────────────
# (a) 스코프 격리 + §0 단일 진실(subscriptions·active_seat_count)
# ─────────────────────────────────────────────
def test_plans_scope_and_single_source(client):
    """구독 카드는 자기 기관 subscriptions 만, 좌석은 active_seat_count(§0)로 산출."""
    mock_conn, mock_cur = _mock_db()
    # 1) subscriptions SELECT → fetchall, 2) active_seat_count → fetchone
    mock_cur.fetchall.return_value = [
        (1, "HST", "campus", 300, "2026-fall", 2,
         date(2026, 8, 1), date(2027, 2, 28), 2500000, "학기 선불", "active", "조직학"),
    ]
    mock_cur.fetchone.side_effect = [(12,)]   # active_seat_count
    with patch("auth.decorators._authenticate", _fake_auth(inst="CNU")), \
         patch("server_render._is_institution_admin", return_value=True), \
         patch("server_render.get_db_conn", return_value=mock_conn), \
         patch("server_render.release_db_conn"):
        resp = client.get("/portal/api/plans")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["institution_id"] == "CNU"
    p = data["plans"][0]
    assert p["subject_code"] == "HST" and p["subject_name"] == "조직학"
    assert p["used_seats"] == 12 and p["max_seats"] == 300
    assert p["seat_rate"] == 4            # round(12/300*100)
    assert p["fee"] == 2500000
    # 단일 진실: subscriptions 에서만, institutions 옛 구독 컬럼 미참조
    sqls = " ".join(_norm(c.args[0]) for c in mock_cur.execute.call_args_list)
    assert "from subscriptions" in sqls
    assert "max_users" not in sqls
    assert "subscription_plan" not in sqls   # institutions deprecated 컬럼
    # scope: 모든 쿼리 파라미터에 자기 기관만
    used = []
    for c in mock_cur.execute.call_args_list:
        if len(c.args) > 1 and c.args[1]:
            used.extend(c.args[1])
    assert "CNU" in used


def test_plans_no_inst_id_from_query(client):
    """body/쿼리의 inst_id 를 신뢰하지 않는다 — 타 기관 id 를 줘도 g.institution_id(CNU)만 쓴다(IDOR)."""
    mock_conn, mock_cur = _mock_db()
    mock_cur.fetchall.return_value = []
    with patch("auth.decorators._authenticate", _fake_auth(inst="CNU")), \
         patch("server_render._is_institution_admin", return_value=True), \
         patch("server_render.get_db_conn", return_value=mock_conn), \
         patch("server_render.release_db_conn"):
        resp = client.get("/portal/api/plans?inst_id=SNU&institution_id=SNU")
    assert resp.status_code == 200
    used = []
    for c in mock_cur.execute.call_args_list:
        if len(c.args) > 1 and c.args[1]:
            used.extend(c.args[1])
    assert "SNU" not in used and "CNU" in used


# ─────────────────────────────────────────────
# (b) 슬라이드 목록 = 배포 메타데이터 / 과목 격리(비구독 403)
# ─────────────────────────────────────────────
def test_plan_slides_subscribed_returns_deployed_metadata(client):
    """구독 과목 → 배포 슬라이드 메타데이터(타일·토큰 없음). deploy_status='deployed' 만."""
    mock_conn, mock_cur = _mock_db()
    mock_cur.fetchall.side_effect = [
        [("HST", "조직학")],                                  # _subscribed_subjects
        [("SA-HST-001", "소장", "HST", "H&E")],               # _portal_subject_slides
    ]
    with patch("auth.decorators._authenticate", _fake_auth(inst="CNU")), \
         patch("server_render._is_institution_admin", return_value=True), \
         patch("server_render.get_db_conn", return_value=mock_conn), \
         patch("server_render.release_db_conn"):
        resp = client.get("/portal/api/plans/slides?subject_code=HST")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["slides"][0]["id"] == "SA-HST-001"
    assert data["slides"][0]["title_ko"] == "소장"
    # 슬라이드 쿼리는 deployed 만 — 미배포 노출 차단(§8 라이선스 격리)
    slide_sql = [_norm(c.args[0]) for c in mock_cur.execute.call_args_list
                 if "from slides" in _norm(c.args[0])][0]
    assert "deploy_status = 'deployed'" in slide_sql
    # (b) 게이트 우회 없음: 포털 경로는 타일 토큰을 발급하지 않는다
    sqls = " ".join(_norm(c.args[0]) for c in mock_cur.execute.call_args_list)
    assert "tile" not in sqls


def test_plan_slides_non_subscribed_403(client):
    """비구독 과목 → 빈 목록 아닌 403 SUBJECT_NOT_SUBSCRIBED(CEO 확정 #2)."""
    mock_conn, mock_cur = _mock_db()
    mock_cur.fetchall.return_value = [("HST", "조직학")]   # _subscribed_subjects(PATH 없음)
    with patch("auth.decorators._authenticate", _fake_auth(inst="CNU")), \
         patch("server_render._is_institution_admin", return_value=True), \
         patch("server_render.get_db_conn", return_value=mock_conn), \
         patch("server_render.release_db_conn"):
        resp = client.get("/portal/api/plans/slides?subject_code=PATH")
    assert resp.status_code == 403
    assert resp.get_json()["error"] == "SUBJECT_NOT_SUBSCRIBED"
    # 슬라이드 쿼리는 아예 실행되지 않음
    assert not any("from slides" in _norm(c.args[0]) for c in mock_cur.execute.call_args_list)


# ─────────────────────────────────────────────
# (+) export 과목 격리(비구독 403) — list 와 동일 allowlist (CEO 보완 #1)
# ─────────────────────────────────────────────
@pytest.mark.parametrize("fmt", ["xlsx", "csv"])
def test_export_non_subscribed_403(client, fmt):
    mock_conn, mock_cur = _mock_db()
    mock_cur.fetchall.return_value = [("HST", "조직학")]
    with patch("auth.decorators._authenticate", _fake_auth(inst="CNU")), \
         patch("server_render._is_institution_admin", return_value=True), \
         patch("server_render.get_db_conn", return_value=mock_conn), \
         patch("server_render.release_db_conn"):
        resp = client.get(f"/portal/api/plans/slides/export?subject_code=PATH&format={fmt}")
    assert resp.status_code == 403
    assert resp.get_json()["error"] == "SUBJECT_NOT_SUBSCRIBED"


def test_export_bad_format_400(client):
    with patch("auth.decorators._authenticate", _fake_auth(inst="CNU")), \
         patch("server_render._is_institution_admin", return_value=True):
        resp = client.get("/portal/api/plans/slides/export?subject_code=HST&format=exe")
    assert resp.status_code == 400
    assert resp.get_json()["error"] == "BAD_FORMAT"


# ─────────────────────────────────────────────
# (c) 내보내기 수식주입 방어 — XLSX/CSV 모두 _xlsx_safe
# ─────────────────────────────────────────────
def _export(client, fmt, slides):
    mock_conn, mock_cur = _mock_db()
    mock_cur.fetchall.side_effect = [
        [("HST", "조직학")],   # _subscribed_subjects
        slides,                # _portal_subject_slides
    ]
    with patch("auth.decorators._authenticate", _fake_auth(inst="CNU")), \
         patch("server_render._is_institution_admin", return_value=True), \
         patch("server_render.get_db_conn", return_value=mock_conn), \
         patch("server_render.release_db_conn"):
        return client.get(f"/portal/api/plans/slides/export?subject_code=HST&format={fmt}")


def test_export_xlsx_formula_injection_defused(client):
    """제목이 '='로 시작하면 셀 값 앞에 ' 가 붙어 수식이 무력화된다(§18 D9)."""
    pytest.importorskip("openpyxl")   # prod requirements 에 포함, 로컬 미설치 시 skip
    resp = _export(client, "xlsx", [("SA-HST-001", "=cmd|'/c calc'!A1", "HST", "H&E")])
    assert resp.status_code == 200
    assert "spreadsheetml.sheet" in resp.headers["Content-Type"]
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(resp.data))
    ws = wb.active
    vals = [c.value for row in ws.iter_rows() for c in row if c.value]
    assert any(isinstance(v, str) and v.startswith("'=cmd") for v in vals)


def test_export_csv_formula_injection_defused(client):
    """CSV 도 동일하게 무력화 + BOM(한글) 포함."""
    resp = _export(client, "csv", [("SA-HST-001", "=HYPERLINK(0)", "HST", "H&E")])
    assert resp.status_code == 200
    body = resp.data.decode("utf-8")
    assert body.startswith("﻿")        # BOM
    assert "'=HYPERLINK(0)" in body          # 프리픽스로 무력화
    assert "SA-HST-001" in body


def test_export_xlsx_ok_basic(client):
    """정상 데이터 xlsx 다운로드(첨부 파일명·시트)."""
    pytest.importorskip("openpyxl")
    resp = _export(client, "xlsx", [("SA-HST-001", "소장", "HST", "H&E")])
    assert resp.status_code == 200
    assert "attachment" in resp.headers.get("Content-Disposition", "")
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(resp.data))
    vals = [c.value for row in wb.active.iter_rows() for c in row if c.value]
    assert "소장" in vals and "SA-HST-001" in vals
