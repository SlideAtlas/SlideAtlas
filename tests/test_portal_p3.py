"""
기관 포털 P3(이용 리포트) pytest — 읽기 전용.

CLAUDE.md §9(멀티테넌시)·§0(단일 진실)·§15-7·§18 D9(과목축 분리→기관 롤업).
내부 QA 체크리스트(이 세션 CEO 확정):
  (a) 스코프 격리 — g.institution_id 강제, inst_id 쿼리 미참조(IDOR 불가), 학교 드롭다운 없음
  (b) 과목 격리 — 비구독 subject_code 는 빈 결과 아닌 403
  (c) 내보내기 수식주입 방어 — XLSX _xlsx_safe
  (d) 집계 과목별 산출→기관 롤업 — active_users·max_seats 과목 스코프(축 미혼합)
  (+) 빈 데이터 graceful — 0/빈 차트·0 나눗셈 가드, ANY(빈배열) 회피

DB 는 mock(로컬 RDS 접속 불가).
"""
import io
import os
from unittest.mock import MagicMock, patch

os.environ.setdefault("JWT_SECRET_KEY", "test-secret-key-for-pytest")
os.environ.setdefault("GMAIL_USER", "test@gmail.com")
os.environ.setdefault("GMAIL_APP_PW", "test-app-pw")
os.environ.setdefault("ADMIN_SECRET_KEY", "test-admin-secret-for-pytest")

import pytest
import server_render as sr
from server_render import app


def _norm(sql):
    return " ".join(str(sql).split()).lower()


@pytest.fixture
def client():
    app.config["TESTING"] = True
    with app.test_client() as c:
        yield c


def _fake_auth(uid="5", inst="CNU", role="admin", subject="HST"):
    def f():
        from flask import g
        g.user_id = uid
        g.institution_id = inst
        g.role = role
        g.subject_code = subject
        return None
    return f


def _mock_db():
    mock_conn = MagicMock()
    mock_cur = MagicMock()
    mock_conn.cursor.return_value.__enter__.return_value = mock_cur
    return mock_conn, mock_cur


def _setup_report(mock_cur, *, subjects=(("HST", "조직학"),),
                  by_position=(("학생", 10), ("교수", 2)),
                  window_rows=(("HST", 150),),   # 접근창 열린 active 구독 (subject_code, max_seats)
                  members=(8, 3, 1), total_views=400,
                  monthly=(("2026-05", 120), ("2026-06", 280)),
                  top=(("SA-HST-001", "소장", "H&E", 90),),
                  ai_q=42, ai_monthly=(("2026-05", 20), ("2026-06", 22))):
    """_subscribed_subjects + _portal_report_data 호출 순서대로 mock 적재(v3.15 2R 반영).

    v3.15 변경: 좌석은 window 구독 rows(subject_code, max_seats)를 fetchall 로 받아 Python 합산(max_seats
    fetchone 제거). 구성원 쿼리는 window_codes 비어있지 않을 때만 fetchone 1회.
    """
    mock_cur.fetchall.side_effect = [
        list(subjects),       # _subscribed_subjects (route)
        list(by_position),    # 1) by_position
        list(window_rows),    # 2-A) 접근창 active 구독 (분자=분모 정합)
        list(monthly),        # monthly_views
        list(top),            # top_slides
        list(ai_monthly),     # ai_monthly
    ]
    fone = []
    if window_rows:           # 2-B) members fetchone 은 window_codes 비어있지 않을 때만 호출됨
        fone.append(members)
    fone += [(total_views,), (ai_q,)]
    mock_cur.fetchone.side_effect = fone


def _run(client, mock_conn, path, auth=None):
    auth = auth or _fake_auth(inst="CNU")
    with patch("auth.decorators._authenticate", auth), \
         patch("server_render._is_institution_admin", return_value=True), \
         patch("server_render.get_db_conn", return_value=mock_conn), \
         patch("server_render.release_db_conn"):
        return client.get(path)


# ─────────────────────────────────────────────
# 인증·관리자 게이트
# ─────────────────────────────────────────────
@pytest.mark.parametrize("path", [
    "/portal/api/report?period=3m&subject_code=all",
    "/portal/api/report/export?period=3m&subject_code=all&format=xlsx",
])
def test_report_requires_auth(client, path):
    assert client.get(path).status_code == 401


@pytest.mark.parametrize("path", [
    "/portal/api/report",
    "/portal/api/report/export?format=xlsx",
])
def test_report_forbidden_non_admin(client, path):
    with patch("auth.decorators._authenticate", _fake_auth(role="viewer")), \
         patch("server_render._is_institution_admin", return_value=False):
        resp = client.get(path)
    assert resp.status_code == 403
    assert resp.get_json()["error"] == "FORBIDDEN"


# ─────────────────────────────────────────────
# (a) 스코프 격리 + 기본 표시
# ─────────────────────────────────────────────
def test_report_basic_shape(client):
    mock_conn, mock_cur = _mock_db()
    _setup_report(mock_cur)
    resp = _run(client, mock_conn, "/portal/api/report?period=3m&subject_code=HST")
    assert resp.status_code == 200
    d = resp.get_json()
    assert d["success"] and d["institution_id"] == "CNU"
    assert d["registered_users"] == 12 and d["by_position"] == {"학생": 10, "교수": 2}
    assert d["members"] == {"active": 8, "unverified": 3, "inactive": 1}
    assert d["active_users"] == 8 and d["max_seats"] == 150
    assert d["util_pct"] == 5            # round(8/150*100)
    assert d["total_views"] == 400
    assert d["per_user_views"] == 50.0   # round(400/8,1)
    assert d["ai_questions"] == 42
    assert len(d["monthly_views"]) == 2 and len(d["top_slides"]) == 1
    assert d["top_slides"][0]["pct"] == 100


def test_report_scope_uses_g_institution_not_query(client):
    """inst_id 를 쿼리로 줘도 무시 — g.institution_id(CNU)만 파라미터화(IDOR, 학교 드롭다운 없음)."""
    mock_conn, mock_cur = _mock_db()
    _setup_report(mock_cur)
    resp = _run(client, mock_conn,
                "/portal/api/report?subject_code=HST&inst_id=SNU&institution_id=SNU")
    assert resp.status_code == 200
    used = []
    for c in mock_cur.execute.call_args_list:
        if len(c.args) > 1 and c.args[1]:
            used.extend([a for a in c.args[1] if isinstance(a, str)])
    assert "CNU" in used and "SNU" not in used


# ─────────────────────────────────────────────
# (b) 과목 격리 — 비구독 403 (report + export)
# ─────────────────────────────────────────────
def test_report_non_subscribed_subject_403(client):
    mock_conn, mock_cur = _mock_db()
    mock_cur.fetchall.return_value = [("HST", "조직학")]   # _subscribed_subjects(PATH 없음)
    resp = _run(client, mock_conn, "/portal/api/report?subject_code=PATH")
    assert resp.status_code == 403
    assert resp.get_json()["error"] == "SUBJECT_NOT_SUBSCRIBED"


def test_export_non_subscribed_subject_403(client):
    mock_conn, mock_cur = _mock_db()
    mock_cur.fetchall.return_value = [("HST", "조직학")]
    resp = _run(client, mock_conn,
                "/portal/api/report/export?subject_code=PATH&format=xlsx")
    assert resp.status_code == 403
    assert resp.get_json()["error"] == "SUBJECT_NOT_SUBSCRIBED"


def test_export_bad_format_400(client):
    """PDF 등은 클라이언트 print — 서버 export 는 xlsx 만."""
    with patch("auth.decorators._authenticate", _fake_auth(inst="CNU")), \
         patch("server_render._is_institution_admin", return_value=True):
        resp = client.get("/portal/api/report/export?subject_code=all&format=pdf")
    assert resp.status_code == 400
    assert resp.get_json()["error"] == "BAD_FORMAT"


# ─────────────────────────────────────────────
# (d) 집계 과목축 분리 — active_users·max_seats 가 과목 스코프
# ─────────────────────────────────────────────
def test_aggregation_subject_scoped(client):
    """특정 과목 선택 시 users·subscriptions 집계가 subject_code 로 스코프된다(§18 D9 축 미혼합)."""
    mock_conn, mock_cur = _mock_db()
    _setup_report(mock_cur)
    _run(client, mock_conn, "/portal/api/report?subject_code=HST")
    # users 집계(활성/지위)에 subject_code 필터
    def _flat(params):
        out = []
        for x in (params or []):
            out.extend(x) if isinstance(x, list) else out.append(x)
        return out
    user_sqls = [(_norm(c.args[0]), _flat(c.args[1]))
                 for c in mock_cur.execute.call_args_list if "from users u" in _norm(c.args[0])]
    assert user_sqls, "users 집계 쿼리가 있어야 함"
    for sql, params in user_sqls:
        assert "u.subject_code" in sql
        assert "HST" in params
    # 좌석: subscriptions 접근창 쿼리(subject_code ANY + access_open_date 윈도우), today 파라미터 포함
    seat_sql = [(_norm(c.args[0]), c.args[1]) for c in mock_cur.execute.call_args_list
                if "from subscriptions" in _norm(c.args[0]) and "access_open_date" in _norm(c.args[0])][0]
    assert "subject_code = any(%s)" in seat_sql[0]
    assert ["HST"] in seat_sql[1]
    # 단일 진실: institutions 옛 구독 컬럼 미참조
    allsql = " ".join(_norm(c.args[0]) for c in mock_cur.execute.call_args_list)
    assert "max_users" not in allsql and "subscription_plan" not in allsql


# ─────────────────────────────────────────────
# (+) 빈 데이터 graceful — 0 나눗셈/빈 차트
# ─────────────────────────────────────────────
def test_empty_data_graceful_zeros(client):
    """구독 과목은 있으나 로그/사용자 0 → 0 나눗셈 없이 0·빈 배열."""
    mock_conn, mock_cur = _mock_db()
    # window_rows=() → 접근창 열린 구독 없음 → members 쿼리 skip, max_seats=0, active_users=0
    _setup_report(mock_cur, by_position=(), window_rows=(),
                  total_views=0, monthly=(), top=(), ai_q=0, ai_monthly=())
    resp = _run(client, mock_conn, "/portal/api/report?subject_code=HST")
    assert resp.status_code == 200
    d = resp.get_json()
    assert d["util_pct"] == 0 and d["per_user_views"] == 0   # 0 나눗셈 가드
    assert d["registered_users"] == 0
    assert d["monthly_views"] == [] and d["top_slides"] == [] and d["ai_monthly"] == []


def test_all_subject_no_subscription_empty_report(client):
    """구독이 하나도 없으면 'all' 은 ANY(빈배열) 쿼리 없이 _empty_report 로 0 반환."""
    mock_conn, mock_cur = _mock_db()
    mock_cur.fetchall.return_value = []   # _subscribed_subjects = {} (구독 0)
    resp = _run(client, mock_conn, "/portal/api/report?subject_code=all")
    assert resp.status_code == 200
    d = resp.get_json()
    assert d["registered_users"] == 0 and d["max_seats"] == 0
    assert d["util_pct"] == 0 and d["subjects"] == []
    # 집계 본쿼리(users/subscriptions/access_logs)는 실행되지 않음(빈배열 회피)
    assert not any("from users u" in _norm(c.args[0]) for c in mock_cur.execute.call_args_list)


def test_chat_logs_failure_graceful(client):
    """chat_logs 부재/오류 시 ai_questions=0·ai_monthly=[] 로 graceful, 나머지 집계 보존."""
    mock_conn, mock_cur = _mock_db()
    # ai_questions 의 fetchone(4번째)에서 예외 → except 가 0/[] 로 처리
    mock_cur.fetchall.side_effect = [
        [("HST", "조직학")], [("학생", 5)], [("HST", 150)],   # subjects, by_position, window_rows
        [("2026-06", 10)], [("SA-HST-001", "소장", "H&E", 7)],  # monthly, top (ai_monthly 미도달)
    ]
    # fetchone: members, total_views, ai_questions(예외) → except 가 0/[] 처리
    mock_cur.fetchone.side_effect = [(4, 1, 0), (10,), Exception("no chat_logs table")]
    resp = _run(client, mock_conn, "/portal/api/report?subject_code=HST")
    assert resp.status_code == 200
    d = resp.get_json()
    assert d["ai_questions"] == 0 and d["ai_monthly"] == []
    assert d["total_views"] == 10 and d["active_users"] == 4   # 나머지 보존
    mock_cur.connection.rollback.assert_called()   # 트랜잭션 정리(다른 집계 보호)


# ─────────────────────────────────────────────
# (c) 내보내기 — xlsx 정상 + 수식주입 방어
# ─────────────────────────────────────────────
def _export(client, slides_top, fmt="xlsx", subject="HST"):
    mock_conn, mock_cur = _mock_db()
    mock_cur.fetchall.side_effect = [
        [("HST", "조직학")],   # _subscribed_subjects
        [("학생", 5)],         # by_position
        [("HST", 150)],        # 2-A) window 구독 (subject_code, max_seats)
        [("2026-06", 10)],     # monthly
        slides_top,            # top_slides
        [("2026-06", 3)],      # ai_monthly
    ]
    mock_cur.fetchone.side_effect = [
        ("충남대학교 의과대학",),   # institutions name_ko
        (4, 1, 0),                  # members (window_codes 있음)
        (10,),                      # total_views
        (3,),                       # ai_questions
    ]
    return _run(client, mock_conn,
                f"/portal/api/report/export?subject_code={subject}&period=3m&format={fmt}")


def test_export_xlsx_ok(client):
    pytest.importorskip("openpyxl")
    resp = _export(client, [("SA-HST-001", "소장", "H&E", 7)])
    assert resp.status_code == 200
    assert "spreadsheetml.sheet" in resp.headers["Content-Type"]
    assert "attachment" in resp.headers.get("Content-Disposition", "")
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(resp.data))
    assert {"요약", "월별 조회수", "인기 슬라이드", "AI 월별 호출"}.issubset(set(wb.sheetnames))


def test_export_xlsx_formula_injection_defused(client):
    """인기 슬라이드 제목이 '='로 시작하면 셀 앞에 ' 가 붙어 무력화(§18 D9)."""
    pytest.importorskip("openpyxl")
    resp = _export(client, [("SA-HST-001", "=cmd|'/c calc'!A1", "H&E", 7)])
    assert resp.status_code == 200
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(resp.data))
    vals = [c.value for ws in wb.worksheets for row in ws.iter_rows()
            for c in row if isinstance(c.value, str)]
    assert any(v.startswith("'=cmd") for v in vals)
