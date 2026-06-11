"""
기관 포털 — 명단 일괄 업로드 '빈 양식 다운로드'(GET /portal/api/roster/template) pytest.

CLAUDE.md §9(스코프 격리)·§3(업로드 파서 정합)·§8/§18 D9(_xlsx_safe).
핵심 단언:
  (a) 인증·관리자 게이트(401 / FORBIDDEN) — 상태변경 아님(GET)이나 _portal_guard 동일.
  (b) ★ 양식 헤더가 업로드 파서 기대 헤더와 '글자 단위' 일치 — 생성한 xlsx 를 그대로
      _parse_xlsx_roster 에 넣으면 헤더가 자동 스킵되고 예시 행만 남는다(자기 양식을 거부하지 않음).
  (c) 과목 안내 = g.institution_id 구독 과목만(타 기관 과목 노출 0, scope 격리) + 구독0 graceful.
  (d) 모든 셀 _xlsx_safe(수식주입 방어).

DB 는 mock — 과목 allowlist 는 _subscribed_subjects 를 직접 패치해 격리한다.
"""
import io
import os
from unittest.mock import patch

os.environ.setdefault("JWT_SECRET_KEY", "test-secret-key-for-pytest")
os.environ.setdefault("GMAIL_USER", "test@gmail.com")
os.environ.setdefault("GMAIL_APP_PW", "test-app-pw")
os.environ.setdefault("ADMIN_SECRET_KEY", "test-admin-secret-for-pytest")

import pytest
import server_render as sr
from server_render import app, _PORTAL_ROSTER_HEADER, _PORTAL_POSITIONS

TEMPLATE = "/portal/api/roster/template"


@pytest.fixture
def client():
    app.config["TESTING"] = True
    with app.test_client() as c:
        yield c


def _fake_auth(uid="5", inst="CNU", role="admin", subject=None):
    def f():
        from flask import g
        g.user_id = uid
        g.institution_id = inst
        g.role = role
        g.subject_code = subject
        return None
    return f


def _admin_ctx(subjects, inst="CNU"):
    """인증+관리자 게이트 통과 + _subscribed_subjects 를 주어진 과목 dict 로 고정한 contextmanager 묶음."""
    from contextlib import ExitStack
    st = ExitStack()
    st.enter_context(patch("auth.decorators._authenticate", _fake_auth(inst=inst)))
    st.enter_context(patch("server_render._is_institution_admin", return_value=True))
    # _subscribed_subjects(cur, inst_id) → 고정 dict. inst_id 인자도 캡처해 scope 단언에 사용.
    seen = {}

    def fake_subs(cur, institution_id):
        seen["inst"] = institution_id
        return dict(subjects)
    st.enter_context(patch("server_render._subscribed_subjects", side_effect=fake_subs))
    # DB 커넥션은 실제로 안 쓰지만 with conn.cursor() 컨텍스트만 통과시키면 됨.
    from unittest.mock import MagicMock
    conn = MagicMock()
    conn.cursor.return_value.__enter__.return_value = MagicMock()
    st.enter_context(patch("server_render.get_db_conn", return_value=conn))
    st.enter_context(patch("server_render.release_db_conn"))
    return st, seen


def _load_xlsx(data):
    import openpyxl
    return openpyxl.load_workbook(io.BytesIO(data), data_only=True)


# ─────────────────────────────────────────────
# (a) 인증·관리자 게이트
# ─────────────────────────────────────────────
def test_template_requires_auth(client):
    assert client.get(TEMPLATE).status_code == 401


def test_template_forbidden_non_admin(client):
    with patch("auth.decorators._authenticate", _fake_auth(role="viewer")), \
         patch("server_render._is_institution_admin", return_value=False):
        resp = client.get(TEMPLATE)
    assert resp.status_code == 403
    assert resp.get_json()["error"] == "FORBIDDEN"


def test_template_bad_format(client):
    with patch("auth.decorators._authenticate", _fake_auth()), \
         patch("server_render._is_institution_admin", return_value=True):
        resp = client.get(TEMPLATE + "?format=exe")
    assert resp.status_code == 400
    assert resp.get_json()["error"] == "BAD_FORMAT"


# ─────────────────────────────────────────────
# (b) ★ 양식 헤더 ↔ 업로드 파서 글자단위 일치 (round-trip)
# ─────────────────────────────────────────────
def test_template_header_matches_upload_parser_exactly(client):
    """생성한 양식을 그대로 _parse_xlsx_roster 에 넣으면 헤더는 스킵되고 예시 행만 남는다."""
    st, _seen = _admin_ctx({"HST": "조직학"})
    with st:
        resp = client.get(TEMPLATE + "?format=xlsx")
    assert resp.status_code == 200

    # 1) 활성 시트 첫 행이 파서 기대 헤더와 정확히 일치
    wb = _load_xlsx(resp.data)
    first_row = [c.value for c in next(wb.active.iter_rows(max_row=1))]
    assert tuple(first_row[:4]) == _PORTAL_ROSTER_HEADER
    # 2) 그 헤더는 파서의 헤더 검출을 통과(=재업로드 시 자동 스킵)
    assert sr._looks_like_header(list(_PORTAL_ROSTER_HEADER)) is True
    # 3) 양식 전체를 파서에 넣으면 헤더 스킵 후 예시 행 1개만(자기 양식을 거부하지 않음)
    rows = sr._parse_xlsx_roster(resp.data, 2000)
    assert rows == [("홍길동", _PORTAL_POSITIONS[-1], "HST", "gildong@univ.ac.kr")]


def test_template_csv_header_and_example(client):
    st, _seen = _admin_ctx({"HST": "조직학"})
    with st:
        resp = client.get(TEMPLATE + "?format=csv")
    assert resp.status_code == 200
    rows = sr._parse_csv_roster(resp.data, 2000)
    # CSV 도 헤더 스킵 후 예시 1행 — xlsx 와 동일 컬럼 정합
    assert rows == [("홍길동", _PORTAL_POSITIONS[-1], "HST", "gildong@univ.ac.kr")]


# ─────────────────────────────────────────────
# (c) 과목 안내 = g.institution_id 구독 과목만 (scope 격리) + graceful
# ─────────────────────────────────────────────
def test_template_subjects_scoped_to_own_institution(client):
    """과목 안내는 _subscribed_subjects(g.institution_id) 결과만 — 타 기관 과목 노출 0."""
    st, seen = _admin_ctx({"HST": "조직학", "PATH": "병리학"}, inst="CNU")
    with st:
        resp = client.get(TEMPLATE + "?format=xlsx&inst_id=SNU&institution_id=SNU")
    assert resp.status_code == 200
    # scope = g.institution_id(CNU) 강제 — 쿼리스트링 inst_id(SNU) 무시
    assert seen["inst"] == "CNU"
    wb = _load_xlsx(resp.data)
    guide = "\n".join(str(c.value) for row in wb["안내"].iter_rows() for c in row if c.value)
    assert "HST(조직학)" in guide and "PATH(병리학)" in guide
    assert "SNU" not in guide          # 타 기관 식별자 노출 없음
    # 예시 과목 = 구독 과목 중 첫 코드(HST)
    assert wb.active["C2"].value == "HST"


def test_template_empty_subscription_graceful(client):
    """구독 0 → 200 + 안내에 '구독 중인 과목 없음' 문구, 예시 과목 빈칸(크래시 없음)."""
    st, _seen = _admin_ctx({})
    with st:
        resp = client.get(TEMPLATE + "?format=xlsx")
    assert resp.status_code == 200
    wb = _load_xlsx(resp.data)
    guide = "\n".join(str(c.value) for row in wb["안내"].iter_rows() for c in row if c.value)
    assert "구독 중인 과목이 없습니다" in guide
    assert (wb.active["C2"].value or "") == ""   # 예시 과목 빈칸


def test_template_positions_from_parser_allowlist(client):
    """지위 안내가 파서 allowlist(_PORTAL_POSITIONS)와 동일 — 행정직원은 제외(파서가 거절하므로)."""
    st, _seen = _admin_ctx({"HST": "조직학"})
    with st:
        resp = client.get(TEMPLATE + "?format=xlsx")
    wb = _load_xlsx(resp.data)
    guide = "\n".join(str(c.value) for row in wb["안내"].iter_rows() for c in row if c.value)
    for p in _PORTAL_POSITIONS:
        assert p in guide
    assert "행정직원" in guide        # 단, '여기 넣지 말라'는 안내로만 등장
    # 행정직원이 지위 allowlist 줄에는 포함되지 않음(파서 거절 값이므로)
    pos_line = [ln for ln in guide.splitlines() if ln.startswith("• 지위:")][0]
    assert "행정직원" not in pos_line


# ─────────────────────────────────────────────
# (d) _xlsx_safe 수식주입 방어
# ─────────────────────────────────────────────
def test_template_formula_injection_neutralized(client):
    """셀 값이 수식문자로 '시작'하면 작은따옴표로 무력화된다(_xlsx_safe).

    예시 과목 셀(C2)은 구독 코드를 그대로 선두에 싣는 셀 — 코드가 '='로 시작하면 프리픽스돼야 한다.
    (구성 문자열 안내 줄은 '   - ' 등으로 시작해 선두가 수식문자가 아니므로 Excel 상 무해 = 가드 no-op.)
    """
    st, _seen = _admin_ctx({"=cmd|'/c calc'!A1": "조직학"})   # 코드가 '='로 시작(수식 위험)
    with st:
        resp = client.get(TEMPLATE + "?format=xlsx")
    wb = _load_xlsx(resp.data)
    assert str(wb.active["C2"].value).startswith("'")    # C2(예시 과목) 무력화 프리픽스
    # CSV 도 동일 가드
    st2, _ = _admin_ctx({"=cmd|'/c calc'!A1": "조직학"})
    with st2:
        resp2 = client.get(TEMPLATE + "?format=csv")
    csv_rows = sr._parse_csv_roster(resp2.data, 2000)
    assert csv_rows[0][2].startswith("'")                 # 예시 행 과목 셀 무력화
