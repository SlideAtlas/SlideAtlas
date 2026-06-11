from dotenv import load_dotenv
load_dotenv()

import os
import sys
import re
import threading
import json
from functools import wraps

from flask import Flask, send_file, Response, request, jsonify, session, redirect, url_for, render_template, g
from werkzeug.security import check_password_hash
from PIL import Image
import io

# -- DB Connection Pool (RDS PostgreSQL) ---------------------------------------
try:
    import psycopg2
    from psycopg2 import pool as pg_pool
    _PSYCOPG2_AVAILABLE = True
except ImportError:
    _PSYCOPG2_AVAILABLE = False

_db_pool = None

def get_db_pool():
    global _db_pool
    if _db_pool is None:
        if not _PSYCOPG2_AVAILABLE:
            raise RuntimeError("psycopg2 미설치 -- pip install psycopg2-binary")
        _db_pool = pg_pool.ThreadedConnectionPool(
            minconn=1,
            maxconn=5,
            host=os.environ["DB_HOST"],
            dbname=os.environ["DB_NAME"],
            user=os.environ["DB_USER"],
            password=os.environ["DB_PASSWORD"],
            port=int(os.environ.get("DB_PORT", "5432")),
        )
    return _db_pool

def get_db_conn():
    return get_db_pool().getconn()

def release_db_conn(conn):
    get_db_pool().putconn(conn)

CATEGORY_MAP = {
    "histology":    "HST",
    "pathology":    "PATH",
    "parasitology": "PARA",
    "anatomy":      "ANAT",
    "embryology":   "EMBRY",
}

app = Flask(__name__)
# [#6] fail-closed: 고정 폴백 금지. 키 누락 시 기동 실패시켜 알려진 secret으로 Flask 세션
#   쿠키(admin 세션)가 위조되는 사고를 막는다(§18 D3). 환경변수/.env 에만 설정.
_admin_secret = os.environ.get('ADMIN_SECRET_KEY')
if not _admin_secret:
    raise RuntimeError("ADMIN_SECRET_KEY 미설정 — 환경변수/.env 확인 (§18 D3). 고정 폴백 금지(어드민 세션 위조 방지).")
app.secret_key = _admin_secret

# -- JWT 인증 Blueprint 등록 ---------------------------------------------------
from auth.auth import auth_bp, active_window_subscription, active_seat_count
app.register_blueprint(auth_bp)
from auth.decorators import (
    login_required, page_login_required, tile_token_required,
    generate_tile_token, verify_tile_token,
    ADMIN_ROSTER_SUBJECT, _today_kst,
)

# -- 어드민 템플릿 컨텍스트 프로세서 -------------------------------------------
# admin_role, admin_name, admin_csrf 를 모든 admin/ 템플릿에 자동 주입.
# g.admin_* 는 admin_required / super_admin_required 통과 후 설정된다.
@app.context_processor
def _admin_context():
    return {
        'admin_role': getattr(g, 'admin_role', None),
        'admin_name': getattr(g, 'admin_name', session.get('admin_name', '')),
        'admin_csrf': session.get('admin_csrf_token', ''),
    }

# -- JSON 데이터 경로 (롤백용 주석 보존) ----------------------------------------
SLIDES_JSON = os.path.join(os.path.dirname(__file__), 'slides.json')
INSTITUTIONS_JSON = os.path.join(os.path.dirname(__file__), 'institutions.json')

# [ROLLBACK] JSON 파일 직접 읽기 -- RDS 전환 전 코드
# def load_slides():
#     if os.path.exists(SLIDES_JSON):
#         with open(SLIDES_JSON, 'r', encoding='utf-8') as f:
#             return json.load(f)
#     return {"slides": []}

def load_slides():
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, institution_id AS institution,
                       subject_code AS category, subject_code,
                       title_ko, title_en, description,
                       s3_key, s3_minimap_key, s3_thumbnail_key,
                       mpp, width, height,
                       stain, organ AS system,
                       species, original_format AS format,
                       conversion_status, deploy_status,
                       (deploy_status = 'deployed') AS active,
                       reject_reason, knowledge_base,
                       'ec2' AS tileserver
                FROM slides
                ORDER BY created_at
            """)
            cols = [d[0] for d in cur.description]
            rows = [dict(zip(cols, row)) for row in cur.fetchall()]
        return {"slides": rows}
    finally:
        release_db_conn(conn)

# [ROLLBACK] JSON 파일 직접 쓰기 -- RDS 전환 전 코드
# def save_slides(data):
#     with open(SLIDES_JSON, 'w', encoding='utf-8') as f:
#         json.dump(data, f, ensure_ascii=False, indent=2)

def save_slides(data):
    # admin_save_slide / admin_delete_slide 가 직접 SQL 처리하므로 no-op.
    pass

# [ROLLBACK] JSON 파일 직접 읽기 -- RDS 전환 전 코드
# def load_institutions():
#     if os.path.exists(INSTITUTIONS_JSON):
#         with open(INSTITUTIONS_JSON, 'r', encoding='utf-8') as f:
#             return json.load(f)
#     return {"institutions": [], "subjects": []}

def load_institutions():
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT id AS code, name_ko, name_en FROM institutions ORDER BY id")
            cols = [d[0] for d in cur.description]
            institutions = [dict(zip(cols, row)) for row in cur.fetchall()]

            cur.execute("SELECT code, name_ko, name_en FROM subject_codes ORDER BY code")
            cols = [d[0] for d in cur.description]
            subjects = [dict(zip(cols, row)) for row in cur.fetchall()]

        return {"institutions": institutions, "subjects": subjects}
    finally:
        release_db_conn(conn)

# 어드민 계정 잠금 정책 (학생 users 정책과 동일, Gemini#1).
ADMIN_LOCK_THRESHOLD = 10    # 24h 윈도우 내 실패 누적 임계값
ADMIN_LOCK_WINDOW_HRS = 24   # 카운팅 윈도우 / 자동 해제 시간


def _get_admin_user():
    """session에서 admin_user_id를 읽어 DB status + session_token을 확인. 유효하면 dict, 아니면 None.

    [Codex#2] 어드민 세션도 매 요청 DB 대조한다. admin_login이 발급한 session_token을 Flask 세션과
      DB(admin_users.session_token) 양쪽에 저장해 두고, 매 요청 compare_digest로 일치를 확인한다.
      → 다른 곳에서 재로그인(토큰 회전)·세션 무효화 시 구 세션 쿠키는 즉시 무효(None 반환).
    """
    admin_user_id = session.get('admin_user_id')
    if not admin_user_id:
        return None
    sess_token = session.get('admin_session_token')
    if not sess_token:
        return None
    try:
        conn = get_db_conn()
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT id, role, name, status, session_token, locked_at"
                    " FROM admin_users WHERE id = %s",
                    (admin_user_id,),
                )
                row = cur.fetchone()
        finally:
            release_db_conn(conn)
        if row is None or row[3] != 'active':
            return None
        # [Codex#1/Gemini#3] 잠금 중(24h 내)이면 기존 세션도 차단한다(매 요청 locked_at 검사).
        locked_at = row[5]
        if locked_at is not None:
            from datetime import datetime, timezone
            if locked_at.tzinfo is None:
                locked_at = locked_at.replace(tzinfo=timezone.utc)
            if (datetime.now(timezone.utc) - locked_at).total_seconds() <= ADMIN_LOCK_WINDOW_HRS * 3600:
                return None
        import secrets as _sec
        db_token = row[4]
        if not db_token or not _sec.compare_digest(str(db_token), str(sess_token)):
            return None
        return {'id': row[0], 'role': row[1], 'name': row[2] or ''}
    except Exception:
        return None


def _admin_check_and_increment_failed(cur, admin_id):
    """어드민 로그인 실패 카운터 증가 + 임계값 도달 시 잠금. 잠겼으면 True (학생 정책 미러)."""
    from datetime import datetime, timezone, timedelta
    cur.execute(
        "SELECT failed_attempts, failed_window_start FROM admin_users WHERE id = %s FOR UPDATE",
        (admin_id,),
    )
    r = cur.fetchone()
    if r is None:
        return False
    failed, window_start = r
    now = datetime.now(timezone.utc)
    if window_start is not None and window_start.tzinfo is None:
        window_start = window_start.replace(tzinfo=timezone.utc)
    if window_start is None or (now - window_start).total_seconds() > ADMIN_LOCK_WINDOW_HRS * 3600:
        new_count = 1
        cur.execute(
            "UPDATE admin_users SET failed_attempts = 1, failed_window_start = %s WHERE id = %s",
            (now, admin_id),
        )
    else:
        new_count = (failed or 0) + 1
        cur.execute(
            "UPDATE admin_users SET failed_attempts = %s WHERE id = %s",
            (new_count, admin_id),
        )
    if new_count >= ADMIN_LOCK_THRESHOLD:
        # [Codex#1/Gemini#3] 잠금 시 session_token도 회전(NULL)해 기존 어드민 세션을 즉시 무효화한다.
        #   (이렇게 안 하면 잠가도 이미 로그인된 세션은 계속 살아 있음.)
        cur.execute(
            "UPDATE admin_users SET locked_at = %s, session_token = NULL WHERE id = %s",
            (now, admin_id),
        )
        return True
    return False


def _admin_is_locked(cur, admin_id, locked_at):
    """locked_at 기준 24h 경과 시 자동 해제. 현재 잠금 여부 반환(True=잠김)."""
    from datetime import datetime, timezone
    if locked_at is None:
        return False
    if locked_at.tzinfo is None:
        locked_at = locked_at.replace(tzinfo=timezone.utc)
    now = datetime.now(timezone.utc)
    if (now - locked_at).total_seconds() > ADMIN_LOCK_WINDOW_HRS * 3600:
        cur.execute(
            "UPDATE admin_users SET locked_at = NULL, failed_attempts = 0, "
            "failed_window_start = NULL WHERE id = %s",
            (admin_id,),
        )
        return False
    return True


def _admin_json_err(message, status):
    resp = jsonify({'ok': False, 'error': message})
    resp.status_code = status
    resp.headers['Cache-Control'] = 'no-store'
    return resp


def admin_required(f):
    """어드민 로그인 + DB status='active' 확인. staff·super_admin 공용."""
    @wraps(f)
    def decorated(*args, **kwargs):
        admin = _get_admin_user()
        if not admin:
            if request.path.startswith('/admin/api/'):
                return _admin_json_err('로그인이 필요합니다', 401)
            return redirect('/admin/login')
        g.admin_user_id = admin['id']
        g.admin_role = admin['role']
        g.admin_name = admin['name']
        return f(*args, **kwargs)
    return decorated


def super_admin_required(f):
    """role='super_admin' 전용. staff가 URL 직접 호출해도 403."""
    @wraps(f)
    def decorated(*args, **kwargs):
        admin = _get_admin_user()
        if not admin:
            if request.path.startswith('/admin/api/'):
                return _admin_json_err('로그인이 필요합니다', 401)
            return redirect('/admin/login')
        if admin['role'] != 'super_admin':
            if request.path.startswith('/admin/api/'):
                return _admin_json_err('권한이 없습니다', 403)
            # HTML 페이지: 대시보드로 리다이렉트 (staff는 운영 페이지 접근 불가)
            return redirect('/admin')
        g.admin_user_id = admin['id']
        g.admin_role = admin['role']
        g.admin_name = admin['name']
        return f(*args, **kwargs)
    return decorated


def admin_csrf_required(f):
    """Admin 세션 기반 CSRF 검증 (JWT CSRF와 별개). POST/PUT/DELETE에만 적용."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if request.method in ('POST', 'PUT', 'DELETE', 'PATCH'):
            import secrets as _sec
            client_token = request.headers.get('X-CSRF-Token')
            stored_token = session.get('admin_csrf_token')
            if not client_token or not stored_token or not _sec.compare_digest(client_token, stored_token):
                return jsonify({'ok': False, 'error': 'CSRF 검증 실패'}), 403
        return f(*args, **kwargs)
    return decorated


def get_slide_institution(slide_id):
    """DB에서 slide의 (institution_id, subject_code, deploy_status, conversion_status) 반환. 없으면 None.

    institution_id는 콘텐츠 소유자 표시('SA')일 뿐 접근 격리 기준이 아니다(§6-1·CEO 결정).
    접근 격리는 subject_code(과목 구독)로 한다. 함수명은 호출부 호환을 위해 유지.
    conversion_status는 특별계정의 QC 미완료 노출 차단(Gemini#7)에 쓰인다.
    """
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT institution_id, subject_code, deploy_status, conversion_status"
                " FROM slides WHERE id = %s",
                (slide_id,),
            )
            return cur.fetchone()
    finally:
        release_db_conn(conn)


# ── 이용 로그 (이용 리포트 데이터 소스, best-effort) ─────────────────────────
#   §8 v3.2: 고빈도 타일 스트리밍 경로엔 절대 INSERT하지 않는다. 뷰어 '진입' 1회,
#   AI 챗 '호출' 1회만 기록한다(타일 1장마다 DB쓰기 금지 — RDS 고갈/DoS 방지).
#   로깅 실패가 본 기능(뷰어/챗) 응답을 막지 않도록 모든 예외를 삼킨다(best-effort).
#   access_logs.subject_code/institution_id는 멱등 마이그레이션으로 보장(db/p05_logging_schema.sql).
def _log_slide_view(user_id, slide_id, institution_id, subject_code):
    """뷰어 진입 1회 열람 로그. subject_code는 콘텐츠 축(슬라이드 과목)으로 기록(§15-7 과목별 집계)."""
    if not _PSYCOPG2_AVAILABLE:
        return
    try:
        uid = int(user_id)
    except (TypeError, ValueError):
        return
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """INSERT INTO access_logs (user_id, slide_id, institution_id, subject_code)
                   VALUES (%s, %s, %s, %s)""",
                (uid, slide_id, institution_id, subject_code),
            )
        conn.commit()
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        print(f"[access_log] 기록 실패(slide={slide_id}): {e}")
    finally:
        release_db_conn(conn)


def _log_chat(user_id, institution_id, slide_id, tab, subject_code):
    """AI 튜터 호출 1회 로그. tab='qa'|'quiz'. subject_code는 슬라이드 과목 축."""
    if not _PSYCOPG2_AVAILABLE:
        return
    try:
        uid = int(user_id)
    except (TypeError, ValueError):
        uid = None
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """INSERT INTO chat_logs (user_id, institution_id, slide_id, tab, subject_code)
                   VALUES (%s, %s, %s, %s, %s)""",
                (uid, institution_id, slide_id, tab, subject_code),
            )
        conn.commit()
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        print(f"[chat_log] 기록 실패(slide={slide_id}): {e}")
    finally:
        release_db_conn(conn)


def _institution_subject_access(institution_id, subject_code):
    """기관이 해당 과목 콘텐츠 접근권을 보유하는가? (단일 게이트 (3))

    institution_subject_access.granted=TRUE 이거나, (institution_id, subject_code)에
    매칭되는 active 구독의 접근창(access_open_date <= today <= subscription_end)이 열려 있으면 True.
    좌석 플랜(subscriptions)과 콘텐츠 접근권(institution_subject_access)은 직교(§16).
    """
    from auth.decorators import _today_kst
    if not institution_id or not subject_code:
        return False
    today = _today_kst()
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT
                       EXISTS(SELECT 1 FROM institution_subject_access
                               WHERE institution_id = %s AND subject_code = %s AND granted = TRUE)
                       OR
                       EXISTS(SELECT 1 FROM subscriptions
                               WHERE institution_id = %s AND subject_code = %s
                                 AND status = 'active'
                                 AND access_open_date <= %s AND subscription_end >= %s)""",
                (institution_id, subject_code, institution_id, subject_code, today, today),
            )
            return bool(cur.fetchone()[0])
    finally:
        release_db_conn(conn)


def _tile_err(error, message, status=403):
    from flask import g as _g
    resp = jsonify({"success": False, "error": error, "message": message})
    resp.status_code = status
    resp.headers["Cache-Control"] = "no-store, no-cache"
    return resp


def _slide_access_allowed(slide_id):
    """슬라이드 접근 단일 게이트. (허용여부, 에러응답) 튜플.

    ★ 격리 기준은 '슬라이드 기관 == 사용자 기관'(구 유튜브형 화석)이 아니라 '과목 구독'이다.
       콘텐츠는 institution_id='SA'로 채번되며, 'SA'는 소유자 표시일 뿐 공용/기본제공이 아니다(§6-1).

    일반 사용자는 다음 전부(AND) 충족 시에만 접근:
      (1) deploy_status == 'deployed'
      (2) g.subject_code == slide.subject_code   (등록 과목 == 슬라이드 과목)
      (3) 사용자 기관이 그 과목 접근권 보유 (_institution_subject_access)
    is_special: deploy_status=='rejected' 차단 + conversion_status 미완료 차단,
                institution·subject 축은 우회(§15-8).
    """
    from flask import g
    info = get_slide_institution(slide_id)
    if info is None:
        return False, _tile_err("SLIDE_NOT_FOUND", "슬라이드를 찾을 수 없습니다", 404)
    _inst_id, slide_subject, deploy_status, conversion_status = info
    if getattr(g, 'is_special', False):
        # [M1] 특별계정: rejected(반려)만 차단, 나머지(qc_pending/deployed) 허용 (CEO 결정, §15-8).
        # 반려 원본은 품질 문제/재공급 대상이므로 특별계정에도 노출 금지(§15-3 라이선스 격리).
        if deploy_status == 'rejected':
            return False, _tile_err("FORBIDDEN", "반려된 슬라이드입니다", 403)
        # [Gemini#7] 변환 미완료(converting/qc_check/pending/failed) 슬라이드는 타일이 깨졌거나
        #   QC 전이므로 특별계정에도 노출 금지. ready / ready_no_mpp만 열람 허용(§4-5).
        if conversion_status not in ('ready', 'ready_no_mpp'):
            return False, _tile_err("FORBIDDEN", "변환이 완료되지 않은 슬라이드입니다", 403)
        return True, None
    # (1) 미배포 슬라이드 학생 접근 불가 (§8 라이선스 격리)
    if deploy_status != 'deployed':
        return False, _tile_err("FORBIDDEN", "접근 권한이 없습니다", 403)
    # (2) 과목 격리: 사용자 등록 과목과 슬라이드 과목 일치 (§0-4·§8 과목 축)
    if slide_subject != getattr(g, 'subject_code', None):
        return False, _tile_err("FORBIDDEN", "접근 권한이 없습니다", 403)
    # (3) 사용자 기관이 그 과목을 구독/접근권 보유
    if not _institution_subject_access(getattr(g, 'institution_id', None), slide_subject):
        return False, _tile_err("FORBIDDEN", "접근 권한이 없습니다", 403)
    return True, None


def _visible_slides(slides):
    """목록/뷰어용 가시 슬라이드 필터 — _slide_access_allowed와 동일 기준의 단일 진실.

    일반 사용자: deploy_status='deployed' AND slide.subject==g.subject_code AND 기관이 그 과목 접근권 보유.
       (2)가 과목 일치를 강제하므로 (3)은 g.subject_code 접근권 1회 확인으로 충분(과목당 N쿼리 회피).
    is_special: rejected 제외 + 변환 미완료 제외(§15-8 검수 목적, ready/ready_no_mpp만).
    """
    from flask import g
    if getattr(g, 'is_special', False):
        return [
            s for s in slides
            if s.get('deploy_status') != 'rejected'
            and s.get('conversion_status') in ('ready', 'ready_no_mpp')
        ]
    subject = getattr(g, 'subject_code', None)
    inst = getattr(g, 'institution_id', None)
    if not subject or not _institution_subject_access(inst, subject):
        return []
    return [
        s for s in slides
        if s.get('deploy_status') == 'deployed' and s.get('subject_code') == subject
    ]


def _verify_tile_request(slide_id):
    """타일 접근 토큰(?t=) 검증. 통과 시 None, 실패 시 에러 응답."""
    from flask import g
    t = request.args.get("t")
    if not t or not verify_tile_token(
        t, str(getattr(g, 'user_id', '')),
        getattr(g, 'institution_id', ''), slide_id
    ):
        # TILE_TOKEN_INVALID: 로그인 세션과 무관한 타일 전용 에러.
        # 프론트 인터셉터가 SESSION_REVOKED·SUBSCRIPTION_EXPIRED와 오판하지 않도록
        # 코드 문자열을 명확히 구분한다.
        return _tile_err("TILE_TOKEN_INVALID", "타일 접근 토큰이 만료되었습니다. 뷰어를 새로고침하세요.", 401)
    return None

# ── 온디맨드 슬라이드 캐시 ──
SLIDE_CACHE = {}
SLIDE_LOCKS = {}
SLIDE_STATUS = {}
TILE_SIZE = 256
OVERLAP = 1
SLIDES_DIR = "/tmp/slides"

EC2_TILESERVER = "http://43.200.171.90:8000"

def get_s3_client():
    import boto3
    return boto3.client(
        's3',
        region_name=os.environ.get('AWS_REGION', 'ap-northeast-2'),
        aws_access_key_id=os.environ.get('AWS_ACCESS_KEY_ID'),
        aws_secret_access_key=os.environ.get('AWS_SECRET_ACCESS_KEY')
    )

def init_slide(slide_id):
    if slide_id not in SLIDE_LOCKS:
        SLIDE_LOCKS[slide_id] = threading.Lock()
    with SLIDE_LOCKS[slide_id]:
        if slide_id in SLIDE_CACHE:
            return True
        SLIDE_STATUS[slide_id] = {"done": False, "error": None}
        try:
            import openslide
            from openslide import deepzoom

            data = load_slides()
            slide_info = next((s for s in data.get('slides', []) if s['id'] == slide_id), None)
            if not slide_info:
                SLIDE_STATUS[slide_id]["error"] = "슬라이드 정보를 찾을 수 없습니다."
                return False

            if slide_info.get('tileserver') == 'ec2':
                ec2_w = slide_info.get("width", 83663)
                ec2_h = slide_info.get("height", 60416)
                SLIDE_CACHE[slide_id] = {"ec2": True, "W": ec2_w, "H": ec2_h, "levels": 18}
                SLIDE_STATUS[slide_id]["done"] = True
                print(f"✅ [{slide_id}] EC2 타일서버 모드 {ec2_w}x{ec2_h}")
                return True

            bucket = os.environ.get('AWS_S3_BUCKET', 'slideatlas-slides')
            s3_key = slide_info.get('s3_key', '')
            fmt = slide_info.get('format', 'dcm').lower()

            os.makedirs(SLIDES_DIR, exist_ok=True)

            if fmt == 'svs':
                local_path = os.path.join(SLIDES_DIR, f"{slide_id}.svs")
                if not os.path.exists(local_path):
                    print(f"SVS 다운로드 중: {s3_key}")
                    s3 = get_s3_client()
                    s3.download_file(bucket, s3_key, local_path)
                slide_obj = openslide.OpenSlide(local_path)
            else:
                dcm_files = slide_info.get('dcm_files', [])
                dcm_entry = slide_info.get('dcm_entry', dcm_files[0] if dcm_files else '')
                slide_dir = os.path.join(SLIDES_DIR, slide_id)
                os.makedirs(slide_dir, exist_ok=True)
                s3 = get_s3_client()
                for fname in dcm_files:
                    local_f = os.path.join(slide_dir, fname)
                    if not os.path.exists(local_f):
                        print(f"DCM 다운로드: {fname}")
                        s3.download_file(bucket, fname, local_f)
                local_path = os.path.join(slide_dir, dcm_entry)
                slide_obj = openslide.OpenSlide(local_path)

            W, H = slide_obj.dimensions
            dz = deepzoom.DeepZoomGenerator(slide_obj, tile_size=TILE_SIZE, overlap=OVERLAP, limit_bounds=True)
            SLIDE_CACHE[slide_id] = {"slide": slide_obj, "dz": dz, "W": W, "H": H, "levels": dz.level_count}
            SLIDE_STATUS[slide_id]["done"] = True
            print(f"✅ [{slide_id}] 로드 완료! {W}x{H}, {dz.level_count}레벨")
            return True
        except Exception as e:
            SLIDE_STATUS[slide_id]["error"] = str(e)
            print(f"❌ [{slide_id}] 오류: {e}")
            return False

# ── EC2 타일서버 프록시 ──
@app.route('/ec2tile/<path:subpath>')
@tile_token_required
def ec2_proxy(subpath):
    # [Gemini#1] 고빈도 타일 경로 — DB 권한 조회 없이 HMAC tile_token만 검증(발급 시점에 _slide_access_allowed 통과).
    # slide_id 추출: "dzi/SA-HST-001.dzi" → "SA-HST-001"
    parts = subpath.split('/')
    raw = parts[1] if len(parts) > 1 else parts[0]
    slide_id = raw.replace('.dzi', '').split('_files')[0]
    # 타일 토큰 검증 (TTL 5분, HMAC). 신원은 tile_token_required가 JWT 복호화로 g에 적재.
    err = _verify_tile_request(slide_id)
    if err:
        return err

    import urllib.request
    url = f"{EC2_TILESERVER}/{subpath}"
    try:
        with urllib.request.urlopen(url, timeout=120) as resp:
            data = resp.read()
            ct = resp.headers.get('Content-Type', 'image/jpeg')
            r = Response(data, mimetype=ct)
            r.headers['Cache-Control'] = 'no-store, no-cache'
            return r
    except Exception as e:
        print(f"EC2 proxy error {url}: {e}")
        return Response(str(e), status=502)

# ── 공지사항 로드 (announcements 테이블 미존재 시 빈 목록 반환) ──
def _load_notices():
    try:
        conn = get_db_conn()
        try:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT title, updated_at
                    FROM announcements
                    WHERE is_published = TRUE AND is_archived = FALSE
                    ORDER BY display_order ASC, updated_at DESC
                    LIMIT 5
                """)
                rows = cur.fetchall()
                return [
                    {
                        'title': r[0],
                        'date': r[1].strftime('%Y.%m.%d') if r[1] else '',
                    }
                    for r in rows
                ]
        finally:
            release_db_conn(conn)
    except Exception:
        return []


# ── 랜딩페이지 ──
@app.route('/')
def landing():
    is_logged = bool(request.cookies.get('access_token'))
    notices = _load_notices()
    return render_template('landing.html', is_logged=is_logged, notices=notices)


@app.route('/login')
def login_page():
    is_logged = bool(request.cookies.get('access_token'))
    # 이미 로그인 상태이면 홈으로 리다이렉트
    if is_logged:
        return redirect('/')
    next_url = request.args.get('next', '/slides')
    # 오픈 리다이렉트 방어: '/'로 시작하는 내부 경로만 허용
    if not next_url.startswith('/') or next_url.startswith('//'):
        next_url = '/slides'
    return render_template('login.html', is_logged=is_logged, next=next_url)


@app.route('/api/institutions', methods=['GET'])
def api_public_institutions():
    """가입 폼 기관 드롭다운용 공개 목록(인증 불요). (§6-4 v3.4, 결정#3 a안)

    id·name_ko만 반환한다 — 구독·좌석·도메인 등 내부 운영 필드는 절대 포함하지 않는다.
    작은 공개 목록이라 rate limit·no-store는 두지 않는다(민감정보 없음).

    노출 기준(§18 D18 재설계, v3.8): '구독(subscriptions) 행이 존재하는 기관'만 반환한다.
    슈퍼관리자가 구독 플랜을 입력한 기관(=고객 학교, 무료 베타 포함 status 무관)만 자동 노출되고,
    콘텐츠 소유자(SA)·공급사·미판매 파트너는 구독 행이 없어 자동 제외된다. 플래그 수동관리 불필요.
    (구 is_subscribable 컬럼 의존 제거 — 컬럼은 죽은 컬럼으로 두고 v1.5에 정리, §18 D18.)
    """
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT DISTINCT i.id, i.name_ko FROM institutions i "
                "JOIN subscriptions s ON s.institution_id = i.id "
                "ORDER BY i.name_ko"
            )
            rows = cur.fetchall()
        institutions = [{"id": r[0], "name_ko": r[1]} for r in rows]
        return jsonify({"success": True, "institutions": institutions})
    finally:
        release_db_conn(conn)


@app.route('/viewer')
@page_login_required
def viewer_default():
    data = load_slides()
    slides = _visible_slides(data.get('slides', []))
    if slides:
        return redirect(f'/viewer/{slides[0]["id"]}')
    return redirect('/')

@app.route('/viewer/<slide_id>')
@page_login_required
def viewer(slide_id):
    data = load_slides()
    slide_info = next((s for s in data.get('slides', []) if s['id'] == slide_id), None)
    if not slide_info:
        return redirect('/slides')
    # 단일 게이트: 과목 구독 기준 접근 판정 — 타일 토큰 발급(아래) 전 반드시 통과 (§8 라이선스 격리).
    allowed, _aerr = _slide_access_allowed(slide_id)
    if not allowed:
        return redirect('/slides')

    # 진입 1회 열람 로그(이용 리포트용, best-effort). 과목 축은 슬라이드 콘텐츠 과목으로 기록.
    #   타일 경로가 아닌 '진입' 시점에만 기록한다(§8 v3.2 — 고빈도 경로 DB쓰기 금지).
    _log_slide_view(
        getattr(g, 'user_id', None), slide_id,
        getattr(g, 'institution_id', None), slide_info.get('subject_code'),
    )

    use_ec2 = slide_info.get('tileserver') == 'ec2'

    if slide_id not in SLIDE_CACHE:
        t = threading.Thread(target=init_slide, args=(slide_id,))
        t.daemon = True
        t.start()

    status = SLIDE_STATUS.get(slide_id, {"done": False, "error": None})

    if status.get("error"):
        return f'''<!DOCTYPE html>
<html><head><meta charset="UTF-8"><title>SlideAtlas</title>
<style>body{{background:#0d1219;color:white;font-family:sans-serif;display:flex;align-items:center;justify-content:center;height:100vh;margin:0;}}</style>
</head><body><div style="text-align:center">
<h2 style="color:#e76f51">오류 발생</h2>
<p style="color:rgba(255,255,255,0.5)">{status["error"]}</p>
<a href="/slides" style="color:#2A9D8F;margin-top:16px;display:block;">← 목록으로</a>
</div></body></html>'''

    if not status.get("done"):
        return f'''<!DOCTYPE html>
<html><head><meta charset="UTF-8"><title>SlideAtlas — 로딩 중</title>
<meta http-equiv="refresh" content="3">
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ background:#0d1219; color:white; font-family:"Segoe UI",sans-serif;
  display:flex; align-items:center; justify-content:center; height:100vh; }}
.progress-bar {{ width:160px; height:2px; background:rgba(91,184,212,0.2);
  border-radius:2px; overflow:hidden; margin:0 auto 20px; }}
.progress-bar::after {{ content:''; display:block; height:100%;
  width:40%; background:#5BB8D4; border-radius:2px;
  animation:progress 1.2s ease-in-out infinite; }}
@keyframes progress {{ 0% {{ transform:translateX(-100%); }} 100% {{ transform:translateX(350%); }} }}
p {{ color:rgba(255,255,255,0.45); font-size:14px; }}
small {{ color:rgba(255,255,255,0.25); font-size:12px; margin-top:8px; display:block; }}
</style>
</head>
<body>
<div style="text-align:center">
  <img src="/static/slideatlas_logo.png" alt="SlideAtlas" style="height:160px;width:auto;margin-bottom:32px;">
  <p>{slide_info.get("title_ko", slide_id)} 로딩 중...</p>
  <small>처음 접속 시 잠시 소요됩니다. 페이지가 자동으로 새로고침됩니다.</small>
</div>
</body></html>'''

    cache = SLIDE_CACHE[slide_id]
    W = cache.get("W", 83663)
    H = cache.get("H", 13119)
    DZ_LEVELS = cache.get("levels", 18)
    title_ko = slide_info.get("title_ko", slide_id)
    title_en = slide_info.get("title_en", "")
    system = slide_info.get("system", "")
    stain = slide_info.get("stain", "H&E")
    mpp = slide_info.get("mpp") or 0.25

    # 타일 접근 토큰 발급 (TTL 5분, §8 Presigned URL)
    from flask import g as _g
    tile_token = generate_tile_token(str(getattr(_g, 'user_id', '')), getattr(_g, 'institution_id', ''), slide_id)

    if use_ec2:
        tile_source_url = f"/ec2tile/dzi/{slide_id}.dzi?t={tile_token}"
        thumbnail_url = f"/ec2tile/thumbnail/{slide_id}?t={tile_token}"
        W = slide_info.get("width", W)
        H = slide_info.get("height", H)
    else:
        tile_source_url = f"/dzi/{slide_id}.dzi?t={tile_token}"
        thumbnail_url = f"/thumbnail/{slide_id}?t={tile_token}"

    # ── 뷰어 HTML ──
    return render_template('viewer.html',
        slide_id=slide_id,
        slide_info=slide_info,
        tile_token=tile_token,
        tile_source_url=tile_source_url,
        thumbnail_url=thumbnail_url,
        W=W, H=H, mpp=mpp,
        title_ko=title_ko,
        title_en=title_en,
        system=system,
        stain=stain,
    )

# ─────────────────────────────────────────────
# 이하 /slides, /api/chat, /dzi, /thumbnail, /admin 등은 원본과 동일
# ─────────────────────────────────────────────

@app.route('/slides')
@page_login_required
def slides():
    # 학생 홈(/home)으로 통합 — 기존 북마크·next 링크 보존을 위해 라우트는 유지하되 진입은 redirect만.
    #   (slides.html 템플릿은 당장 삭제하지 않고 남겨둔다 — §21 1단계.)
    return redirect('/home')


@app.route('/home')
@page_login_required
def home():
    """로그인 후 학생 홈 — 수업 탭 / 전체 탭 2탭 셸(§21 1단계).

    - subject_code 없는 admin-only(콘텐츠 비소비자)는 슬라이드 0개 화면 대신 /portal로 보낸다.
    - 그 외(viewer·겸직)는 home.html 렌더. '전체 탭' 슬라이드 목록은 기존 단일 게이트의
      _visible_slides(load_slides())를 그대로 재사용한다 — 접근 정책·필터 로직을 새로 만들지 않는다(§8 불변).
    - '수업 탭'은 1단계에선 빈 상태 placeholder(데이터 연동은 3단계).
    """
    # admin-only(순수 관리자, 좌석 0·콘텐츠 비소비, §6-4)는 홈 대신 포털로.
    if g.role == 'admin' and g.subject_code is None:
        return redirect('/portal')

    data = load_slides()
    # 단일 게이트와 동일 기준의 가시 슬라이드 필터(§6-1·§8) — 새 필터 로직 추가 금지.
    all_slides = _visible_slides(data.get('slides', []))
    total = len(all_slides)
    stain_class = {'H&E': 'he', 'PAS': 'pas', 'Masson Trichrome': 'masson', 'Silver': 'silver'}
    # 계통(organ) 드롭다운 옵션 — 전체 탭 클라이언트 필터용(빈 값 제외, 정렬).
    organs = sorted({s.get('system') for s in all_slides if s.get('system')})

    # 표시명(roster.name)·과목명(subject_codes.name_ko) — 헤더 표시용.
    display_name, subject_name = '', ''
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT COALESCE(r.name, ''), COALESCE(sc.name_ko, u.subject_code, '')
                     FROM users u
                     LEFT JOIN institution_rosters r
                       ON lower(r.email) = lower(u.email)
                      AND r.institution_id = u.institution_id
                      AND r.subject_code = u.subject_code
                     LEFT JOIN subject_codes sc ON sc.code = u.subject_code
                    WHERE u.id = %s""",
                (g.user_id,),
            )
            row = cur.fetchone()
            if row:
                display_name, subject_name = row[0] or '', row[1] or ''
    finally:
        release_db_conn(conn)

    # 헤더 "관리자 포털" 링크 노출 여부 — 포털 게이트와 동일 기준(현재 __ADMIN__ roster 행 존재, §9).
    is_admin = _is_institution_admin(g.user_id, g.institution_id)
    return render_template('home.html',
        slides=all_slides,
        total=total,
        organs=organs,
        stain_class=stain_class,
        display_name=display_name,
        subject_name=subject_name,
        is_admin=is_admin,
    )

def _is_institution_admin(user_id, institution_id):
    """로그인 사용자가 자기 기관 관리자 명단에 '현재' 등록돼 있는지 확인(§9).

    [Codex#1/Gemini#3] 포털 접근의 유일한 권위는 관리자 roster 행의 존재다.
      (institution_id, subject_code='__ADMIN__', email, role='admin') 행이 지금 존재해야만 True.
      users.role 단독으로는 통과시키지 않는다 → 관리자 권한 회수(roster 행 DELETE) 시 즉시 차단.
    같은 이메일이 과목 행('HST' 등)과 관리자 행('__ADMIN__')으로 공존할 수 있으므로 센티넬 과목코드까지
    함께 대조한다(과목 학생 행을 관리자 권한으로 오인하지 않게).

    [perf 단계1] 같은 요청에서 _authenticate 가 이미 동일 사실(__ADMIN__ 행 존재)을 조회해
      g._admin_roster_cache 에 적재했으면 재조회 없이 재사용한다(중복 admin 조회 1회 제거). 캐시는
      '인증된 본인 user + 본인 기관(institution_id == g.institution_id)' 일 때만 적용 — 다른 기관 인자로
      호출되면(향후 D15 다기관) 캐시 불일치로 폴백 조회한다. _authenticate 의 _has_admin_roster(=
      r.institution_id=u.institution_id) 와 본 함수(=r.institution_id=%s)는 institution_id==본인기관일 때
      동일 술어이므로 판정 의미가 100% 일치한다(반환 bool 불변).
    """
    cache = getattr(g, '_admin_roster_cache', None)
    if cache is not None:
        c_uid, c_inst, c_val = cache
        if c_uid == str(user_id) and c_inst == institution_id:
            return c_val
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT 1 FROM institution_rosters r
                   JOIN users u ON lower(u.email) = lower(r.email)
                  WHERE u.id = %s AND r.institution_id = %s
                    AND r.subject_code = %s AND r.role = 'admin'
                  LIMIT 1""",
                (user_id, institution_id, ADMIN_ROSTER_SUBJECT),
            )
            return cur.fetchone() is not None
    finally:
        release_db_conn(conn)


@app.route('/portal')
@page_login_required
def portal():
    # 멀티테넌시 1급 규칙(§9): scope는 로그인 사용자의 기관으로 강제. 다른 기관 데이터 접근 불가.
    #   [Codex#1/Gemini#3] 포털 접근 게이트는 '현재 관리자 roster 행 존재' 단일 기준으로만 판정한다.
    #   users.role=='admin' 단독 우회는 제거 — 권한 회수(roster DELETE) 시 즉시 포털 차단되어야 하므로.
    inst_id = g.institution_id
    if not _is_institution_admin(g.user_id, inst_id):
        return redirect('/')

    inst_name = inst_id
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT name_ko FROM institutions WHERE id = %s", (inst_id,))
            row = cur.fetchone()
            if row and row[0]:
                inst_name = row[0]
    finally:
        release_db_conn(conn)

    # "슬라이드 보기" 링크 노출 여부. 겸직(subject_code 보유)만 콘텐츠 접근권이 있으므로 노출하고,
    #   순수 admin-only(subject_code NULL=좌석0·콘텐츠 비소비, §6-4)는 슬라이드 0개라 숨긴다.
    has_slides = g.subject_code is not None
    # 최소 라우트(§18 D15): scope 격리·게이트만 우선 구현. 3탭(명단관리·구독플랜·이용리포트)
    #   본화면은 D15 별도 작업에서 institution_portal.html 목업(§17) 기준으로 구현.
    return render_template('portal.html', institution_id=inst_id, institution_name=inst_name,
                           has_slides=has_slides)


# ═════════════════════════════════════════════════════════════════════════════
# 기관 포털 P1 — 명단 관리 API (§9 멀티테넌시 / §21 / D17 sync 해결)
#   포털 관리자(_is_institution_admin 통과)가 자기 기관 '과목 이용자 명단'을 관리한다.
#   · scope 는 g.institution_id 로 강제 — 타 기관 명단 접근·수정 불가(§9).
#   · 상태변경 API 는 login_required(→ 더블서밋 CSRF 자동) + _portal_guard(_is_institution_admin 재확인).
#   · 과목 입력 allowlist = '구독 행 보유 과목'(CEO 확정). 미래학기 구독=접근창 닫힘은 분기 C 안전망.
#   · __ADMIN__(기관 관리자) 행은 읽기전용 표시 — 추가·삭제는 슈퍼관리자 기관수정 화면 관할(CEO 확정).
#   · sync 판정식은 register()와 동일한 공통 헬퍼(active_window_subscription/active_seat_count)만 사용(§0).
#   · role 은 어떤 sync 경로에서도 UPDATE 하지 않는다 — 겸직 admin 보존(§3).
# ═════════════════════════════════════════════════════════════════════════════
# [v3.9 Codex/Gemini High#2 저장형 XSS 방어] 따옴표·괄호·세미콜론·제어문자를 허용하던 느슨한
#   `[^@\s]+` 를 운영 안전 allowlist 로 좁힌다. 개별추가·업로드 공통. (템플릿 측은 inline onclick 제거.)
_PORTAL_EMAIL_RE = re.compile(r'^[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}$')
_PORTAL_POSITIONS = ('학생', '조교', '교수')   # 과목 명단 지위(행정직원=admin-only, 과목행 없음 §21-1)
_PORTAL_UPLOAD_MAX_ROWS = 2000
_PORTAL_UPLOAD_MAX_BYTES = 10 * 1024 * 1024      # 업로드 스트림 바이트 상한(Content-Length 헤더 비신뢰, 실계수)
_PORTAL_XLSX_MAX_UNCOMPRESSED = 50 * 1024 * 1024  # xlsx 압축해제 총합 상한(압축폭탄 1차 방어)
# [v3.9 Codex 2차 Low#2] entry 수는 '보조 backstop'으로만 — 시트·이미지·로고·스타일 포함 정상 업무 xlsx는
#   entry 100을 넘기 쉬워 정상 명단 오탐을 유발했다. 핵심 방어는 압축해제 총량+실측 업로드 크기+행/셀 상한이며,
#   entry 수는 극단적 비정상(수천 entry)만 거른다. xlsx 실무 편의 우선.
_PORTAL_XLSX_MAX_ENTRIES = 1000
_PORTAL_OUTCOME_MSG = {
    'synced': '동기화(과목 전환)',
    'no_change': '변경 없음',
    'seat_full': '좌석 부족 보류',
    'pending_window': '구독 시작일에 자동 반영',
    'multi_subject_hold': '다과목 미지원 보류(D12)',
    'added_no_user': '명단 추가(가입 대기)',
    'removed_seat_reclaimed': '제거·좌석 반환',
    'removed_roster_only': '명단 행만 제거',
}


class _RosterParseError(Exception):
    pass


def _portal_guard():
    """포털 API 공통 게이트: 로그인 사용자가 자기 기관 관리자인지 재확인.
    반환 (inst_id, None) 또는 (None, (json_error, status))."""
    inst_id = getattr(g, 'institution_id', None)
    if not inst_id or not _is_institution_admin(g.user_id, inst_id):
        return None, (jsonify({'success': False, 'error': 'FORBIDDEN',
                               'message': '기관 관리자 권한이 없습니다'}), 403)
    return inst_id, None


def _subscribed_subjects(cur, institution_id):
    """그 기관이 '구독 행을 보유한' 과목만(status 무관) — 멤버 추가/업로드 과목 allowlist.
    반환 {code: name_ko}(name_ko 없으면 code 폴백)."""
    cur.execute(
        """SELECT DISTINCT sc.code, COALESCE(sc.name_ko, sc.code)
             FROM subscriptions s
             JOIN subject_codes sc ON sc.code = s.subject_code
            WHERE s.institution_id = %s
            ORDER BY sc.code""",
        (institution_id,),
    )
    return {r[0]: r[1] for r in cur.fetchall()}


def _clean_name(raw):
    """표시 이름 위생: 공백/개행/제어문자 제거 + 길이 캡(100). (수식주입은 export 시점 방어 — §8.)"""
    s = (raw or '').replace('\r', ' ').replace('\n', ' ').strip()
    s = ''.join(ch for ch in s if ord(ch) >= 32)
    return s[:100]


def _sync_member(cur, institution_id, subject_code, email, name, position, today, seat_cache):
    """과목 명단 행 1건 upsert + 동일 이메일 기존 user 동기화(§3 D17 해결).

    판정식은 register()와 동일한 공통 헬퍼만 재사용(§0). role 은 절대 건드리지 않는다(겸직 admin 보존).
    seat_cache: {subject_code: [max_seats(int|None), used(int), found(bool)]}
      — 일괄 업로드 좌석 직렬화 캐시(첫 사용 시 구독행 FOR UPDATE 잠금+카운트, 이후 메모리 차감).
    반환 outcome ∈ {'synced','no_change','seat_full','pending_window','multi_subject_hold','added_no_user'}.
    """
    email = email.strip().lower()
    name = _clean_name(name)

    # 1) 동일 이메일·동일 기관 기존 user 조회.
    #    [v3.9 Codex/Gemini High#1 IDOR] institution_id 로 스코프한다 — 이 누락 시 A기관 관리자가
    #    B기관 사용자 이메일을 자기 roster에 넣어 B기관 user의 subject_code/position을 바꾸거나
    #    '추가 후 삭제'로 B기관 user의 active 과목을 NULL 회수해 슬라이드 접근을 끊을 수 있다(§9 위반).
    #    타 기관 이메일은 '현재 기관에 기존 user 없음'으로 취급 → roster-only(분기 D).
    #    [v3.9 Codex 2차 Med#1] status 포함 — 좌석 카운팅은 active 사용자에만 적용(아래 2)).
    cur.execute(
        "SELECT id, subject_code, status FROM users WHERE lower(email) = %s AND institution_id = %s",
        (email, institution_id),
    )
    urow = cur.fetchone()
    u_status = urow[2] if urow is not None else None

    # 2) 좌석 판정을 roster 쓰기 '앞'에 둔다(Med#3): seat_full이면 roster upsert도 user 변경도 하지 않고
    #    그 행만 skip(아무 것도 안 바뀜). 좌석 검사는 admin-only(subject NULL) 승격(분기 A)에서만 의미.
    #    판정식은 register/verify 와 동일한 공통 헬퍼만 사용(§0).
    #    [v3.9 Codex 2차 Med#1] '실제 active 좌석을 만드는 경우(status='active')'에만 좌석을 검사·가산한다 —
    #      메모리 캐시 증분 기준을 active_seat_count(status='active')와 정확히 일치시킨다(§0 단일판정식).
    #      pending user 승격은 active 좌석을 점유하지 않으므로 seat_full로 막지 않고(빈 좌석 오거부 방지),
    #      그 좌석 검사는 verify_email 의 FOR UPDATE 재검사에 위임한다(활성화 시점에 권위 판정).
    #      단 접근창(found) 판정에 쓰려고 pending 도 st 는 계산해 둔다(좌석 가산은 안 함).
    if urow is not None and urow[1] is None:
        st = seat_cache.get(subject_code)
        if st is None:
            found, max_seats = active_window_subscription(
                cur, institution_id, subject_code, today, for_update=True)
            used = active_seat_count(cur, institution_id, subject_code) if found else 0
            st = [max_seats, used, found]
            seat_cache[subject_code] = st
        _max, _used, _found = st
        if u_status == 'active' and _found and _max is not None and _used >= _max:
            return 'seat_full'        # active 승격이 정원 초과 — roster 미생성, user 불변(전체 롤백 아님).

    # 3) roster 행 upsert (role='viewer' 고정 — role 은 sync 대상 아님). is_verified 기존값 유지.
    #    seat_full 을 제외한 모든 분기(synced/no_change/multi/pending/added)에서 roster 행은 남는다.
    cur.execute(
        """INSERT INTO institution_rosters (institution_id, subject_code, email, name, role, position)
           VALUES (%s, %s, %s, %s, 'viewer', %s)
           ON CONFLICT (institution_id, subject_code, email)
           DO UPDATE SET name = EXCLUDED.name, position = EXCLUDED.position""",
        (institution_id, subject_code, email, name, position),
    )

    # 4) user 동기화 분기.
    if urow is None:
        return 'added_no_user'        # 분기 D — 실제 좌석 점유·채번은 그 사람이 가입할 때(또는 타 기관 이메일).
    user_id, u_subject = urow[0], urow[1]

    if u_subject == subject_code:
        # 이미 이 과목 — position 만 동기화(좌석 변화 없음, subject·role 불변).
        cur.execute("UPDATE users SET position = %s WHERE id = %s AND institution_id = %s",
                    (position, user_id, institution_id))
        return 'no_change'
    if u_subject is not None:
        # 분기 B — 이미 다른 과목. v1.0 단일 subject_code라 덮어쓰지 않음(D12). roster 행만 추가됨.
        return 'multi_subject_hold'

    # u_subject IS NULL = admin-only. 분기 A/C — st 는 위 2)에서 채워짐(접근창 판정).
    _max, _used, _found = seat_cache[subject_code]
    if not _found:
        return 'pending_window'       # 분기 C — 접근창 닫힘(미래학기/창 전). admin-only 유지(fail-closed).
    # 분기 A — NULL→과목 전환 + position 갱신. role 불변(겸직 admin 보존).
    cur.execute(
        "UPDATE users SET subject_code = %s, position = %s WHERE id = %s AND institution_id = %s",
        (subject_code, position, user_id, institution_id),
    )
    # active 사용자를 실제 좌석으로 만든 경우에만 메모리 좌석 +1(§0: active_seat_count 기준과 일치).
    #   pending user 전환은 좌석 미점유 — verify 의 FOR UPDATE 재검사가 활성화 시점에 정원을 집행한다.
    if u_status == 'active':
        seat_cache[subject_code][1] = _used + 1
    return 'synced'


def _remove_member(cur, institution_id, subject_code, email):
    """과목 명단 행 1건 제거 + 좌석 회수(§3). __ADMIN__ 행은 제거 불가(읽기전용, 슈퍼관리자 관할).
    반환 outcome ∈ {'removed_seat_reclaimed','removed_roster_only','not_found','admin_row_protected'}."""
    email = email.strip().lower()
    if subject_code == ADMIN_ROSTER_SUBJECT:
        return 'admin_row_protected'
    cur.execute(
        "DELETE FROM institution_rosters "
        "WHERE institution_id = %s AND subject_code = %s AND lower(email) = %s",
        (institution_id, subject_code, email),
    )
    if cur.rowcount == 0:
        return 'not_found'
    # 동일 이메일·동일 기관 user 가 이 과목을 현재 active 과목으로 점유 중이면 회수.
    #   [v3.9 High#1 IDOR] institution_id 스코프 — 타 기관 user 행을 NULL 회수하지 못하게(§9).
    cur.execute(
        "SELECT id, subject_code FROM users WHERE lower(email) = %s AND institution_id = %s",
        (email, institution_id),
    )
    urow = cur.fetchone()
    if urow and urow[1] == subject_code:
        # subject_code NULL 회수(좌석 1석 반환 + 슬라이드 접근 즉시 차단=단일 게이트 자동) + position NULL.
        #   role 불변 → 겸직(__ADMIN__ 보유)이면 admin-only 복귀, 계정 자체는 삭제 안 함(§9 계정 불가침).
        cur.execute(
            "UPDATE users SET subject_code = NULL, position = NULL WHERE id = %s AND institution_id = %s",
            (urow[0], institution_id),
        )
        return 'removed_seat_reclaimed'
    return 'removed_roster_only'


def _looks_like_header(cells):
    joined = ' '.join((c or '').strip() for c in cells[:4]).lower()
    return ('이름' in joined or 'name' in joined) and ('이메일' in joined or 'email' in joined)


def _read_capped(stream, max_bytes):
    """업로드 스트림을 청크로 읽되 실측 바이트가 상한을 넘으면 즉시 중단(Content-Length 헤더 비신뢰).
       chunked 인코딩으로 헤더를 우회해도 실제 바이트 계수로 차단된다."""
    chunks, total = [], 0
    while True:
        chunk = stream.read(65536)
        if not chunk:
            break
        total += len(chunk)
        if total > max_bytes:
            raise _RosterParseError(f'파일이 너무 큽니다(최대 {max_bytes // (1024 * 1024)}MB)')
        chunks.append(chunk)
    return b''.join(chunks)


def _xlsx_zip_guard(data):
    """xlsx(=zip) 선검사: load 전에 압축해제 총합·entry 수를 보고 압축폭탄을 차단(load_workbook은
       내부적으로 압축을 풀므로 반드시 load 전에 검사한다)."""
    import zipfile as _zip, io as _io
    try:
        zf = _zip.ZipFile(_io.BytesIO(data))
    except _zip.BadZipFile:
        raise _RosterParseError('엑셀 파일을 읽을 수 없습니다')
    try:
        infos = zf.infolist()
        if len(infos) > _PORTAL_XLSX_MAX_ENTRIES:
            raise _RosterParseError('엑셀 구조가 비정상입니다(entry 과다)')
        total_unc = sum(i.file_size for i in infos)
        if total_unc > _PORTAL_XLSX_MAX_UNCOMPRESSED:
            raise _RosterParseError('엑셀 압축 해제 크기가 과도합니다(압축폭탄 의심)')
    finally:
        zf.close()


def _rows_from_iter(row_iter, max_rows):
    """공통: (4컬럼 정규화 + 헤더/빈행 스킵 + 스트리밍 중 행상한 적용). 상한 초과 시 즉시 거부.
       빈 행이 과도해도 절대 무한루프하지 않도록 전체 순회 횟수도 backstop 으로 제한."""
    out = []
    scanned = 0
    scan_limit = max_rows * 50 + 1000   # 빈 행/forged dimension 대비 backstop
    for i, row in enumerate(row_iter):
        scanned += 1
        if scanned > scan_limit:
            raise _RosterParseError('엑셀 행 수가 과도합니다')
        # 셀 길이 캡(보안검증 권고): 거대 단일 셀이 다운스트림으로 전파/저장되지 않게 일관 제한.
        #   정상 이름/이메일/과목/지위는 수십자 이내 — 512자 캡은 유효 데이터에 영향 없고, 초과분은
        #   어차피 validator/allowlist/DB 제약에서 거부된다(이름은 _clean_name 에서 100자 재캡).
        cells = ['' if c is None else str(c)[:512] for c in (list(row) + ['', '', '', ''])[:4]]
        if i == 0 and _looks_like_header(cells):
            continue
        if not any(c.strip() for c in cells):
            continue
        out.append((cells[0], cells[1], cells[2], cells[3]))
        if len(out) > max_rows:
            raise _RosterParseError(f'최대 {max_rows}행까지 처리합니다')
    return out


def _parse_xlsx_roster(data, max_rows):
    """xlsx 안전 파싱(★ 포맷 유지 — CSV 전용 전환 금지, 학교 실무는 xlsx).
       (a) read_only=True 스트리밍 (b) zip 압축폭탄 선검사 (c) 행상한 스트리밍 적용."""
    try:
        import openpyxl
    except ImportError:
        raise _RosterParseError('서버에 openpyxl 미설치')
    import io as _io
    _xlsx_zip_guard(data)
    try:
        wb = openpyxl.load_workbook(_io.BytesIO(data), read_only=True, data_only=True)
    except _RosterParseError:
        raise
    except Exception:
        raise _RosterParseError('엑셀 파일을 읽을 수 없습니다')
    try:
        return _rows_from_iter(wb.active.iter_rows(values_only=True), max_rows)
    finally:
        wb.close()


def _parse_csv_roster(data, max_rows):
    import csv as _csv, io as _io
    text = None
    for enc in ('utf-8-sig', 'cp949', 'utf-8'):
        try:
            text = data.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise _RosterParseError('CSV 인코딩을 해석할 수 없습니다')
    return _rows_from_iter(_csv.reader(_io.StringIO(text)), max_rows)


@app.route('/portal/api/roster', methods=['GET'])
@login_required
def portal_roster_list():
    inst_id, err = _portal_guard()
    if err:
        return err
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            subjects = _subscribed_subjects(cur, inst_id)
            cur.execute(
                """SELECT r.name, r.position, r.subject_code, r.email, r.is_verified
                     FROM institution_rosters r
                    WHERE r.institution_id = %s
                    ORDER BY (r.subject_code = %s) ASC, r.subject_code, r.name""",
                (inst_id, ADMIN_ROSTER_SUBJECT),
            )
            members, admins = [], []
            for name, position, subj, email, verified in cur.fetchall():
                if subj == ADMIN_ROSTER_SUBJECT:
                    admins.append({'name': name or '', 'email': email,
                                   'is_verified': bool(verified)})   # 읽기전용 표시
                else:
                    members.append({'name': name or '', 'position': position,
                                    'subject_code': subj,
                                    'subject_name': subjects.get(subj, subj),
                                    'email': email, 'is_verified': bool(verified)})
        return jsonify({'success': True, 'institution_id': inst_id,
                        'subjects': [{'code': c, 'name_ko': n} for c, n in subjects.items()],
                        'members': members, 'admins': admins})
    finally:
        release_db_conn(conn)


@app.route('/portal/api/roster', methods=['POST'])
@login_required
def portal_roster_add():
    inst_id, err = _portal_guard()
    if err:
        return err
    body = request.get_json(silent=True) or {}
    name = (body.get('name') or '').strip()
    position = (body.get('position') or '').strip()
    subject_code = (body.get('subject_code') or '').strip()
    email = (body.get('email') or '').strip().lower()
    if not email or not _PORTAL_EMAIL_RE.match(email):
        return jsonify({'success': False, 'error': 'INVALID_EMAIL',
                        'message': '이메일 형식이 올바르지 않습니다'}), 400
    if position not in _PORTAL_POSITIONS:
        return jsonify({'success': False, 'error': 'INVALID_POSITION',
                        'message': '지위는 학생/조교/교수만 가능합니다'}), 400

    conn = get_db_conn()
    conn.autocommit = False
    try:
        with conn.cursor() as cur:
            subjects = _subscribed_subjects(cur, inst_id)
            if subject_code not in subjects:
                conn.rollback()
                return jsonify({'success': False, 'error': 'INVALID_SUBJECT',
                                'message': '구독 중인 과목만 등록할 수 있습니다'}), 400
            outcome = _sync_member(cur, inst_id, subject_code, email, name,
                                   position, _today_kst(), {})
        conn.commit()
    except Exception:
        conn.rollback()
        release_db_conn(conn)
        return jsonify({'success': False, 'error': 'SERVER_ERROR',
                        'message': '처리 중 오류가 발생했습니다'}), 500
    finally:
        try:
            conn.autocommit = True
        except Exception:
            pass
    release_db_conn(conn)
    return jsonify({'success': True, 'outcome': outcome,
                    'message': _PORTAL_OUTCOME_MSG.get(outcome, outcome)})


@app.route('/portal/api/roster', methods=['DELETE'])
@login_required
def portal_roster_delete():
    inst_id, err = _portal_guard()
    if err:
        return err
    body = request.get_json(silent=True) or {}
    subject_code = (body.get('subject_code') or '').strip()
    email = (body.get('email') or '').strip().lower()
    if not email or not subject_code:
        return jsonify({'success': False, 'error': 'MISSING_FIELDS',
                        'message': '과목과 이메일이 필요합니다'}), 400
    conn = get_db_conn()
    conn.autocommit = False
    try:
        with conn.cursor() as cur:
            outcome = _remove_member(cur, inst_id, subject_code, email)
        conn.commit()
    except Exception:
        conn.rollback()
        release_db_conn(conn)
        return jsonify({'success': False, 'error': 'SERVER_ERROR',
                        'message': '처리 중 오류가 발생했습니다'}), 500
    finally:
        try:
            conn.autocommit = True
        except Exception:
            pass
    release_db_conn(conn)
    if outcome == 'not_found':
        return jsonify({'success': False, 'error': 'NOT_FOUND',
                        'message': '해당 명단 행이 없습니다'}), 404
    if outcome == 'admin_row_protected':
        return jsonify({'success': False, 'error': 'ADMIN_ROW_PROTECTED',
                        'message': '기관 관리자 행은 포털에서 제거할 수 없습니다'}), 403
    return jsonify({'success': True, 'outcome': outcome,
                    'message': _PORTAL_OUTCOME_MSG.get(outcome, outcome)})


@app.route('/portal/api/roster/upload', methods=['POST'])
@login_required
def portal_roster_upload():
    inst_id, err = _portal_guard()
    if err:
        return err
    # 명단 파일은 텍스트(수천 행이라도 수백 KB). 전역 MAX_CONTENT_LENGTH 는 어드민 슬라이드(GB급 SVS)
    #   업로드를 깨뜨리므로 두지 않고, 이 라우트에서만 국소로 막는다(High#4):
    #   · content_length 는 '빠른 사전 거부'로만 쓰고(헤더 비신뢰) 권위 검사는 실측 바이트(_read_capped).
    #   · xlsx 는 load 전에 zip 압축폭탄 선검사 + read_only 스트리밍 + 행상한.
    if request.content_length and request.content_length > _PORTAL_UPLOAD_MAX_BYTES:
        return jsonify({'success': False, 'error': 'FILE_TOO_LARGE',
                        'message': f'파일이 너무 큽니다(최대 {_PORTAL_UPLOAD_MAX_BYTES // (1024*1024)}MB)'}), 413
    f = request.files.get('file')
    if f is None or not f.filename:
        return jsonify({'success': False, 'error': 'NO_FILE', 'message': '파일이 없습니다'}), 400
    fname = f.filename.lower()
    if not (fname.endswith('.xlsx') or fname.endswith('.csv')):
        return jsonify({'success': False, 'error': 'BAD_FORMAT',
                        'message': '.xlsx 또는 .csv 파일만 지원합니다'}), 400
    try:
        data = _read_capped(f.stream, _PORTAL_UPLOAD_MAX_BYTES)   # 실측 바이트 상한
        if fname.endswith('.xlsx'):
            rows = _parse_xlsx_roster(data, _PORTAL_UPLOAD_MAX_ROWS)
        else:
            rows = _parse_csv_roster(data, _PORTAL_UPLOAD_MAX_ROWS)
    except _RosterParseError as e:
        return jsonify({'success': False, 'error': 'PARSE_ERROR', 'message': str(e)}), 400
    if not rows:
        return jsonify({'success': False, 'error': 'EMPTY', 'message': '데이터 행이 없습니다'}), 400

    conn = get_db_conn()
    conn.autocommit = False
    results = []
    try:
        with conn.cursor() as cur:
            subjects = _subscribed_subjects(cur, inst_id)
            # 과목 셀은 code 또는 name_ko 둘 다 허용 → code 로 정규화.
            name_to_code = {}
            for c, n in subjects.items():
                name_to_code[c.lower()] = c
                name_to_code[(n or '').strip().lower()] = c
            today = _today_kst()
            seat_cache, seen = {}, set()
            for idx, (rname, rpos, rsubj, remail) in enumerate(rows, start=1):
                remail = (remail or '').strip().lower()
                rpos = (rpos or '').strip()
                rsubj = (rsubj or '').strip()
                # 행별 검증 — '예상된 거절'은 skip-and-report(전체 롤백 아님, §3).
                if not remail or not _PORTAL_EMAIL_RE.match(remail):
                    results.append({'row': idx, 'email': remail, 'outcome': 'invalid_email'})
                    continue
                if rpos not in _PORTAL_POSITIONS:
                    results.append({'row': idx, 'email': remail, 'outcome': 'invalid_position'})
                    continue
                code = name_to_code.get(rsubj.lower())
                if code is None:
                    results.append({'row': idx, 'email': remail, 'outcome': 'invalid_subject'})
                    continue
                key = (code, remail)
                if key in seen:
                    results.append({'row': idx, 'email': remail, 'outcome': 'duplicate_row'})
                    continue
                seen.add(key)
                outcome = _sync_member(cur, inst_id, code, remail, rname, rpos, today, seat_cache)
                results.append({'row': idx, 'email': remail, 'subject_code': code, 'outcome': outcome})
        conn.commit()
    except Exception:
        conn.rollback()
        release_db_conn(conn)
        return jsonify({'success': False, 'error': 'SERVER_ERROR',
                        'message': '처리 중 오류로 전체 취소되었습니다'}), 500
    finally:
        try:
            conn.autocommit = True
        except Exception:
            pass
    release_db_conn(conn)
    counts = {}
    for r in results:
        counts[r['outcome']] = counts.get(r['outcome'], 0) + 1
    return jsonify({'success': True, 'total': len(results), 'counts': counts, 'results': results})


# 업로드 파서(_parse_*_roster)가 기대하는 헤더 — 양식과 파서가 어긋나면 업로드가 자기 양식을
#   거부하는 모순이 생긴다. 그래서 헤더는 '문자열 상수'로 한곳에 두고 양식·검증이 함께 쓴다.
#   · 순서는 파서 위치읽기(cells[0..3] = 이름·지위·과목·이메일)와 동일.
#   · 헤더 검출(_looks_like_header)은 1행에 '이름'·'이메일'이 있으면 헤더로 보고 건너뛴다 → 이 헤더는
#     재업로드 시 자동 스킵된다.
_PORTAL_ROSTER_HEADER = ('이름', '지위', '과목', '이메일')


@app.route('/portal/api/roster/template', methods=['GET'])
@login_required
def portal_roster_template():
    """명단 일괄 업로드용 빈 양식(xlsx/csv) 다운로드.

    업로드(POST /portal/api/roster/upload)와 '정확히 같은' 컬럼·순서·검증 기준으로 생성한다:
      · 헤더 = _PORTAL_ROSTER_HEADER(파서 기대 문자열 그대로) → 재업로드 시 자동 헤더 스킵.
      · 지위 안내 = _PORTAL_POSITIONS(파서 allowlist 그대로). 행정직원은 과목 명단 대상이 아니라
        admin-only(슈퍼관리자 관할, 과목행 없음 §21-1)라 양식 지위에서 제외 — 넣으면 파서가 거절.
      · 과목 안내 = _subscribed_subjects(g.institution_id) — 그 기관 구독 과목만(타 기관 노출 0, §9 scope).
      · 모든 셀 _xlsx_safe(§8·§18 D9 수식주입 방어).
    """
    inst_id, err = _portal_guard()
    if err:
        return err
    fmt = (request.args.get('format') or 'xlsx').strip().lower()
    if fmt not in ('xlsx', 'csv'):
        return jsonify({'success': False, 'error': 'BAD_FORMAT',
                        'message': 'xlsx 또는 csv만 지원합니다'}), 400

    # 과목 allowlist 는 업로드 파서와 동일 헬퍼(_subscribed_subjects) — scope=g.institution_id 강제.
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            subjects = _subscribed_subjects(cur, inst_id)   # {code: name_ko}
    finally:
        release_db_conn(conn)

    # 안내 문자열(양식↔파서 단일 출처에서 파생)
    positions = list(_PORTAL_POSITIONS)                     # 파서 allowlist 그대로(학생/조교/교수)
    subj_lines = [f'{c}({n})' for c, n in subjects.items()] # 예: "HST(조직학)"
    example_subject = next(iter(subjects), '')              # 구독 과목 중 첫 코드(없으면 빈칸)
    guide_lines = [
        '◆ 이용자 명단 업로드 양식 (포털 명단 관리)',
        '',
        '1) 첫 행(이름·지위·과목·이메일)은 머리글입니다 — 지우지 말고 그 아래부터 입력하세요.',
        '2) 예시 행(홍길동…)은 삭제한 뒤 실제 명단을 입력하세요.',
        '',
        '• 지위: ' + ' / '.join(positions) + '  (이 외 값은 업로드 시 거절됩니다)',
        '   ※ 행정직원은 과목 명단이 아니라 기관 관리자(admin) 등록 대상이라 이 양식에 넣지 않습니다.',
        '• 과목(아래 구독 과목 코드만 허용 — 비구독 과목은 업로드 거절):',
    ]
    if subj_lines:
        guide_lines += ['   - ' + s for s in subj_lines]
    else:
        guide_lines += ['   - (현재 구독 중인 과목이 없습니다 — 구독 등록 후 명단을 업로드하세요.)']
    guide_lines += [
        '• 이메일: 유효한 이메일 형식(예: gildong@univ.ac.kr).',
        '• 참고: 해당 과목에 구독이 있어야 학생이 가입·접근할 수 있습니다(§6-3).',
    ]
    example_row = ['홍길동', positions[-1] if positions else '교수', example_subject, 'gildong@univ.ac.kr']

    import io
    from flask import send_file as _sf
    fname_base = f'SlideAtlas_{inst_id}_roster_template'

    if fmt == 'csv':
        # CSV 는 시트·코멘트가 없어 안내를 머리글 위 주석으로 넣으면 재업로드 시 데이터 행으로 오인되므로,
        #   round-trip 안전을 위해 헤더+예시 행만 출력한다(안내는 xlsx 양식이 제공).
        import csv as _csv
        buf = io.StringIO()
        buf.write('﻿')   # BOM — Excel 한글 깨짐 방지
        w = _csv.writer(buf)
        w.writerow([_xlsx_safe(h) for h in _PORTAL_ROSTER_HEADER])
        w.writerow([_xlsx_safe(v) for v in example_row])
        data = buf.getvalue().encode('utf-8')
        return _sf(io.BytesIO(data), as_attachment=True,
                   download_name=f'{fname_base}.csv', mimetype='text/csv; charset=utf-8')

    # xlsx — 활성 시트='명단'(헤더+예시, 재업로드 시 wb.active 가 이 시트), 별도 '안내' 시트.
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.comments import Comment
    except ImportError:
        return jsonify({'success': False, 'error': 'NO_OPENPYXL',
                        'message': 'openpyxl 미설치'}), 500
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '명단'
    ws.append([_xlsx_safe(h) for h in _PORTAL_ROSTER_HEADER])   # row1 = 헤더(파서 자동 스킵)
    HDR = Font(bold=True, color='FFFFFF')
    FILL = PatternFill('solid', fgColor='1A2238')
    for cell in ws[1]:
        cell.font = HDR
        cell.fill = FILL
    ws.append([_xlsx_safe(v) for v in example_row])            # row2 = 예시(삭제 후 입력)
    for cell in ws[2]:
        cell.font = Font(color='888888', italic=True)
    ws['A2'].comment = Comment('예시 행입니다. 삭제한 뒤 실제 명단을 입력하세요.', 'SlideAtlas')
    for col, wdt in zip('ABCD', (16, 10, 12, 30)):
        ws.column_dimensions[col].width = wdt

    ws2 = wb.create_sheet('안내')
    for line in guide_lines:
        ws2.append([_xlsx_safe(line)])
    ws2['A1'].font = Font(bold=True, size=13)
    ws2.column_dimensions['A'].width = 70
    for row in ws2.iter_rows():
        row[0].alignment = Alignment(wrap_text=True, vertical='top')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return _sf(buf, as_attachment=True, download_name=f'{fname_base}.xlsx',
               mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ═════════════════════════════════════════════════════════════════════════════
# 기관 포털 P2 — 구독 플랜 (읽기 전용, §9 멀티테넌시 / §0 단일 진실 / §16)
#   포털 관리자가 자기 기관의 (기관×과목) 구독 카드·좌석 현황·과목별 배포 슬라이드 목록을 본다.
#   · 슈퍼관리자 엔드포인트(api_institution_detail 등) 직접 호출 금지 — 포털 전용 읽기 래퍼.
#   · scope 는 g.institution_id 강제(_portal_guard) — inst_id 를 body/쿼리로 받지 않음(IDOR 불가, §9).
#   · 구독 데이터 원천은 subscriptions(기관×과목)만 — institutions 옛 구독 컬럼 참조 0건(§0).
#   · 좌석 현황 = active_seat_count(status='active') — P1·리포트와 동일 단일판정식(§0, pending 미점유).
#   · 접근창·만료·D-day = 기존 _sub_status/_sem_dates 재사용(새 계산식 금지).
#   · 슬라이드 목록은 '배포 메타데이터 카탈로그'(타일 없음). "열람"은 /viewer/<id> → 표준
#     _slide_access_allowed 게이트 판정(관리자도 과목 좌석 필요). 포털이 게이트를 우회하지 않는다.
#   · 모든 slides 경로(목록·export)는 _subscribed_subjects allowlist 로 과목 격리 —
#     비구독 subject_code 는 빈 목록이 아니라 403(명시적 거부, CEO 확정).
# ═════════════════════════════════════════════════════════════════════════════
def _portal_subject_slides(cur, subject_code):
    """그 과목의 배포(deploy_status='deployed') 슬라이드 메타데이터(읽기).

    콘텐츠는 institution_id='SA' 단일 채번이며 접근 격리는 과목 구독으로 한다(§6-1·§8) —
    고객 기관 id 로 필터하지 않는다. 과목 격리(비구독 거부)는 호출 측 allowlist 검사에서.
    """
    cur.execute(
        """SELECT id, title_ko, subject_code, stain
             FROM slides
            WHERE subject_code = %s AND deploy_status = 'deployed'
            ORDER BY id""",
        (subject_code,),
    )
    return [{'id': r[0], 'title_ko': r[1] or '', 'subject_code': r[2],
             'stain': r[3] or ''} for r in cur.fetchall()]


@app.route('/portal/api/plans', methods=['GET'])
@login_required
def portal_plans_list():
    inst_id, err = _portal_guard()
    if err:
        return err
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT s.id, s.subject_code, s.plan, s.max_seats,
                          s.start_term, s.term_count,
                          s.access_open_date, s.subscription_end,
                          s.fee, s.payment_method, s.status,
                          COALESCE(sc.name_ko, s.subject_code)
                     FROM subscriptions s
                     LEFT JOIN subject_codes sc ON sc.code = s.subject_code
                    WHERE s.institution_id = %s
                    ORDER BY s.subscription_end DESC""",
                (inst_id,),
            )
            rows = cur.fetchall()
            today = _today_kst()    # [item4] 대시보드 날짜도 게이트와 동일 KST 기준(§18 D10)
            plans = []
            for (sid, subj, plan, max_seats, start_term, term_count,
                 open_d, end_d, fee, pay, status, subj_name) in rows:
                # 좌석 현황: register/verify/P1 과 동일한 active_seat_count(§0 단일판정식).
                used = active_seat_count(cur, inst_id, subj)
                status_key, status_label = _sub_status(open_d, end_d)
                if status_key == 'active':
                    dday = f'D-{(end_d - today).days} 만료'
                elif status_key in ('upcoming', 'pending'):
                    dday = f'D-{(open_d - today).days} 오픈'
                else:
                    dday = f'D+{(today - end_d).days} 만료'
                plans.append({
                    'id': sid, 'subject_code': subj, 'subject_name': subj_name,
                    'plan': plan, 'max_seats': max_seats, 'used_seats': used,
                    'seat_rate': round(used / max_seats * 100) if max_seats else 0,
                    'start_term': start_term, 'term_count': term_count,
                    'access_open_date': str(open_d) if open_d else None,
                    'subscription_end': str(end_d) if end_d else None,
                    'fee': fee, 'payment_method': pay,
                    'status': status, 'status_key': status_key,
                    'status_label': status_label, 'dday': dday,
                })
        return jsonify({'success': True, 'institution_id': inst_id, 'plans': plans})
    finally:
        release_db_conn(conn)


@app.route('/portal/api/plans/slides', methods=['GET'])
@login_required
def portal_plan_slides():
    inst_id, err = _portal_guard()
    if err:
        return err
    subject_code = (request.args.get('subject_code') or '').strip()
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            subjects = _subscribed_subjects(cur, inst_id)
            if subject_code not in subjects:
                # 비구독 과목 = 명시적 거부(빈 목록 아님, CEO 확정). 과목 격리(§8·§9).
                return jsonify({'success': False, 'error': 'SUBJECT_NOT_SUBSCRIBED',
                                'message': '구독하지 않은 과목입니다'}), 403
            slides = _portal_subject_slides(cur, subject_code)
        return jsonify({'success': True, 'subject_code': subject_code,
                        'subject_name': subjects.get(subject_code, subject_code),
                        'slides': slides})
    finally:
        release_db_conn(conn)


@app.route('/portal/api/plans/slides/export', methods=['GET'])
@login_required
def portal_plan_slides_export():
    inst_id, err = _portal_guard()
    if err:
        return err
    subject_code = (request.args.get('subject_code') or '').strip()
    fmt = (request.args.get('format') or 'xlsx').strip().lower()
    if fmt not in ('xlsx', 'csv'):
        return jsonify({'success': False, 'error': 'BAD_FORMAT',
                        'message': 'xlsx 또는 csv만 지원합니다'}), 400
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            subjects = _subscribed_subjects(cur, inst_id)
            if subject_code not in subjects:
                # export 도 동일 allowlist — 비구독 과목 403(CEO 보완 #1·#2).
                return jsonify({'success': False, 'error': 'SUBJECT_NOT_SUBSCRIBED',
                                'message': '구독하지 않은 과목입니다'}), 403
            subj_name = subjects.get(subject_code, subject_code)
            slides = _portal_subject_slides(cur, subject_code)
    finally:
        release_db_conn(conn)

    import io
    from flask import send_file as _sf
    headers = ['슬라이드 ID', '제목', '과목', '염색']
    fname_base = f'SlideAtlas_{inst_id}_{subject_code}_slides'
    # 수식 주입 방어: 모든 셀 _xlsx_safe 재사용(§8·§18 D9). CSV·XLSX 공통.
    if fmt == 'csv':
        import csv as _csv
        buf = io.StringIO()
        buf.write('﻿')   # BOM — Excel 한글 깨짐 방지
        w = _csv.writer(buf)
        w.writerow(headers)
        for s in slides:
            w.writerow([_xlsx_safe(s['id']), _xlsx_safe(s['title_ko']),
                        _xlsx_safe(subject_code), _xlsx_safe(s['stain'])])
        data = buf.getvalue().encode('utf-8')
        return _sf(io.BytesIO(data), as_attachment=True,
                   download_name=f'{fname_base}.csv', mimetype='text/csv; charset=utf-8')
    # xlsx
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return jsonify({'success': False, 'error': 'NO_OPENPYXL',
                        'message': 'openpyxl 미설치'}), 500
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '슬라이드 목록'
    ws.append([_xlsx_safe(f'{subj_name} 슬라이드 목록')])
    ws['A1'].font = Font(bold=True, size=13)
    ws.append([])
    ws.append(headers)
    HDR = Font(bold=True, color='FFFFFF')
    FILL = PatternFill('solid', fgColor='1A2238')
    for cell in ws[3]:
        cell.font = HDR
        cell.fill = FILL
    for s in slides:
        ws.append([_xlsx_safe(s['id']), _xlsx_safe(s['title_ko']),
                   _xlsx_safe(subject_code), _xlsx_safe(s['stain'])])
    for col, wdt in zip('ABCD', (20, 36, 10, 16)):
        ws.column_dimensions[col].width = wdt
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return _sf(buf, as_attachment=True, download_name=f'{fname_base}.xlsx',
               mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ═════════════════════════════════════════════════════════════════════════════
# 기관 포털 P3 — 이용 리포트 (읽기 전용, §9 멀티테넌시 / §0 단일 진실 / §15-7·§18 D9)
#   포털 관리자가 자기 기관의 이용 집계(KPI·구성원활동·월별조회·Top10·AI)를 본다.
#   · 슈퍼관리자 reports 엔드포인트 직접 호출 안 함 — SQL·집계 로직만 재사용한 포털 전용 래퍼.
#   · scope 는 g.institution_id 강제(_portal_guard) — inst_id 를 body/쿼리로 안 받음(IDOR 불가, §9).
#     슈퍼관리자엔 '학교 선택 드롭다운'이 있으나 포털엔 없다(자기 기관 고정).
#   · 과목 격리: subject_code='all'(구독과목 합산) 또는 _subscribed_subjects 중 하나 — 비구독은 403(P2 동일).
#   · 단일 진실(§0): 집계 원천 = access_logs·chat_logs·users·subscriptions 만(institutions 옛 컬럼 0건).
#     '활성 사용자' = status='active'(P1·P2·active_seat_count 일치). util/per_user 0나눗셈 가드.
#   · 과목축 분리→기관 롤업(§18 D9): active_users·max_seats·소진율은 (기관×과목) 산출, all 은 과목별 합(SUM).
#     단일 사용자=단일 subject_code 라 합산 중복 없음. 과목축을 섞지 않는다.
#   · 데이터 현실: access_logs·chat_logs 실데이터가 거의 없어 0/빈 차트가 정상 — graceful 표시.
# ═════════════════════════════════════════════════════════════════════════════
_PORTAL_REPORT_PERIODS = {'1m': 30, '3m': 90, '6m': 180}   # 'all' = 무필터(전체)
_VALID_REPORT_PERIODS = {'1m', '3m', '6m', 'all'}          # [Codex Low item5] allowlist
_DEFAULT_REPORT_PERIOD = '3m'


def _norm_report_period(period):
    """[item5] 허용값 외 period 는 조용히 '전체'로 확장되지 않게 기본값('3m')으로 고정."""
    return period if period in _VALID_REPORT_PERIODS else _DEFAULT_REPORT_PERIOD


def _portal_report_range(period):
    """기간 코드 → (start_date, end_date) | (None, None). 날짜 필터일 뿐(집계식 신규 아님)."""
    from datetime import timedelta
    days = _PORTAL_REPORT_PERIODS.get(period)
    if days is None:
        return None, None        # 'all'/미지정 → 전체 기간
    today = _today_kst()         # [item4] 리포트 기간 경계도 KST 기준(게이트와 일관, §18 D10)
    return today - timedelta(days=days), today


def _empty_report():
    """구독 과목이 하나도 없을 때(또는 빈 데이터) 반환할 0 구조 — ANY(빈배열) 회피."""
    return {
        'registered_users': 0, 'by_position': {},
        'members': {'active': 0, 'unverified': 0, 'inactive': 0},
        'active_users': 0, 'max_seats': 0, 'util_pct': 0,
        'total_views': 0, 'ai_questions': 0, 'per_user_views': 0,
        'monthly_views': [], 'top_slides': [], 'ai_monthly': [],
    }


def _portal_report_data(cur, inst_id, subject_code, subjects, start, end):
    """이용 리포트 집계 1회 산출(JSON·export 공통). 모든 쿼리는 inst_id 로 스코프(§9).

    과목 필터: 'all'→구독과목 ANY(codes), 특정과목→= code. 날짜: access=accessed_at, chat=created_at.
    chat_logs 는 구버전 DB 부재 가능 → 마지막에 try 로 감싸 graceful 0/[] (다른 집계 보호).
    """
    # [Codex High#1] 조회수 계열은 access_logs 스냅샷(al.institution_id·al.subject_code)으로 필터한다.
    #   JOIN users/slides 의 '현재값'(u.institution_id·s.subject_code)으로 거르면 사용자 과목·기관 이동이나
    #   슬라이드 과목 변경 시 과거 로그가 현재 기준으로 재분류돼 타 기관/타 과목 리포트에 섞인다(시간축 오염).
    #   → al_filt = 로그 스냅샷 과목(조회수 집계용), u_filt = users 현재 상태(등록/활성 집계용, 정상),
    #     c_filt = chat_logs 스냅샷(AI 호출, 이미 스냅샷 사용 중이라 정상).
    if subject_code == 'all':
        codes = list(subjects.keys())
        u_filt, u_p = "AND u.subject_code = ANY(%s)", [codes]
        al_filt, al_p = "AND al.subject_code = ANY(%s)", [codes]
        c_filt, c_p = "AND cl.subject_code = ANY(%s)", [codes]
        seat_codes = codes
    else:
        u_filt, u_p = "AND u.subject_code = %s", [subject_code]
        al_filt, al_p = "AND al.subject_code = %s", [subject_code]
        c_filt, c_p = "AND cl.subject_code = %s", [subject_code]
        seat_codes = [subject_code]

    # [Codex 타임존 item4] 날짜 경계는 KST(_today_kst) 기준 [start, end] — '>= start AND < end+1day'.
    al_date, al_dp, cl_date, cl_dp = "", [], "", []
    if start and end:
        al_date = "AND al.accessed_at >= %s AND al.accessed_at < %s + INTERVAL '1 day'"
        al_dp = [start, end]
        cl_date = "AND cl.created_at >= %s AND cl.created_at < %s + INTERVAL '1 day'"
        cl_dp = [start, end]

    # 1) 등록 이용자(지위별) — 전체 구독 과목 기준(등록 명단 총원, 현재 접근창 무관).
    #    [Codex 2R#1 §0] is_special 제외절 제거 — 특별계정은 승격 시 subject_code=NULL 이라 subject 필터
    #    (u_filt)로 이미 빠진다. active_seat_count(is_special 절 없음)와 '같은 집합' 기준으로 정렬한다.
    cur.execute(f"""
        SELECT u.position, COUNT(*) FROM users u
        WHERE u.institution_id = %s {u_filt}
        GROUP BY u.position
    """, [inst_id] + u_p)
    by_position, registered = {}, 0
    for pos, cnt in cur.fetchall():
        by_position[pos or '기타'] = cnt
        registered += cnt

    # 2-A) 현재 접근창이 열린 active 구독(과목·정원) — 소진율 분자/분모의 '같은 행 집합'(기준 A).
    #    [Codex 2R#2] 분모(max_seats)뿐 아니라 분자(active_users)도 이 창에 속한 과목만 본다.
    #    만료 과목의 active 사용자는 _authenticate 가 이미 접근 차단한 유령 → 분자에서도 제외(현실 정합).
    #    today=_today_kst(1R). ★ active_seat_count(P1·P2 점유 카운트)는 불변 — 여긴 리포트 소진율 산출만.
    #    [Codex 3R §0] 과목별 '권위 row 1개'로 정규화 — 같은 (기관,과목)에 접근창 겹치는 active 구독이
    #    2개+이면 SUM 이 중복 합산(150+150=300)해 인증 게이트(active_window_subscription: 과목당
    #    ORDER BY subscription_end DESC LIMIT 1, auth/auth.py)보다 분모가 커진다(§0 위반). 새 규칙을 만들지
    #    않고 게이트와 동일한 선택을 DISTINCT ON 으로 따른다 → 분모의 과목집합·과목별 정원 = 분자(window_codes)
    #    = 인증 게이트가 보는 권위 구독, 셋이 같은 행 집합. (동률 subscription_end 코너는 D22 추적 유지.)
    today = _today_kst()
    cur.execute("""
        SELECT DISTINCT ON (subject_code) subject_code, max_seats FROM subscriptions
        WHERE institution_id=%s AND subject_code = ANY(%s) AND status='active'
          AND access_open_date <= %s AND subscription_end >= %s
        ORDER BY subject_code, subscription_end DESC
    """, [inst_id, seat_codes, today, today])
    window_rows = cur.fetchall()                              # 과목당 1행(권위 row = 인증 게이트와 동일 선택)
    window_codes = sorted({r[0] for r in window_rows})        # 접근창 열린 과목(분자=분모 정합)
    max_seats = sum(r[1] for r in window_rows if r[1] is not None)

    # 2-B) 구성원 활동 + 활성 사용자 — 현재 접근창 과목(window_codes)만(분자=분모 같은 집합, §0).
    #    [Codex 2R#1] is_special 제외절 없음: 특별계정은 subject_code=NULL 이라 window_codes ANY 필터로
    #    자연 제외 → active_seat_count(status='active' AND subject_code=…)와 '글자까지 같은 집합'.
    #    활성=status='active'(NULL 제외, 1R Med#2). 미인증=pending. 비활성=그 외(NULL 포함, IS DISTINCT FROM).
    if window_codes:
        cur.execute("""
            SELECT
              COUNT(*) FILTER (WHERE u.status='active'),
              COUNT(*) FILTER (WHERE u.status='pending_verification'),
              COUNT(*) FILTER (WHERE u.status IS DISTINCT FROM 'active'
                                 AND u.status IS DISTINCT FROM 'pending_verification')
            FROM users u
            WHERE u.institution_id = %s AND u.subject_code = ANY(%s)
        """, [inst_id, window_codes])
        a, p, i = cur.fetchone()
        members = {'active': a or 0, 'unverified': p or 0, 'inactive': i or 0}
    else:
        # 접근창 열린 구독 없음 → 좌석 집계 대상 없음(만료/미래만). 분자·분모 양쪽 0 = '집계 제외'.
        members = {'active': 0, 'unverified': 0, 'inactive': 0}
    active_users = members['active']
    util_pct = round(active_users / max_seats * 100) if max_seats > 0 else 0

    # 4) 총 슬라이드 조회수 — access_logs 스냅샷 기준(Codex High#1). users/slides 조인 불요.
    cur.execute(f"""
        SELECT COUNT(al.id) FROM access_logs al
        WHERE al.institution_id = %s {al_filt} {al_date}
    """, [inst_id] + al_p + al_dp)
    total_views = cur.fetchone()[0] or 0
    per_user_views = round(total_views / active_users, 1) if active_users > 0 else 0

    # 5) 월별 슬라이드 조회수(최근 12개월) — 스냅샷 기준.
    cur.execute(f"""
        SELECT TO_CHAR(DATE_TRUNC('month', al.accessed_at),'YYYY-MM'), COUNT(al.id)
        FROM access_logs al
        WHERE al.institution_id = %s {al_filt} {al_date}
        GROUP BY 1 ORDER BY 1 LIMIT 12
    """, [inst_id] + al_p + al_dp)
    monthly_views = [{'label': r[0], 'count': r[1]} for r in cur.fetchall()]
    mv_max = max((m['count'] for m in monthly_views), default=0)
    for m in monthly_views:
        m['pct'] = round(m['count'] / mv_max * 100) if mv_max else 0

    # 6) 인기 슬라이드 Top 10 — 스냅샷(al.subject_code)으로 필터(High#1).
    #    [Codex 2R#4] LEFT JOIN — slides 행이 없거나 깨진 참조여도 집계에서 떨구지 않게(total/monthly 와
    #    같은 기준). 제목은 COALESCE(s.title_ko, al.slide_id) 폴백. 조인은 표시용, 필터는 al.* 만.
    cur.execute(f"""
        SELECT al.slide_id, COALESCE(s.title_ko, al.slide_id), s.stain, COUNT(al.id) AS views
        FROM access_logs al
        LEFT JOIN slides s ON al.slide_id = s.id
        WHERE al.institution_id = %s {al_filt} {al_date}
        GROUP BY al.slide_id, s.title_ko, s.stain
        ORDER BY views DESC LIMIT 10
    """, [inst_id] + al_p + al_dp)
    top_slides = [{'id': r[0], 'title': r[1] or '', 'stain': r[2] or '', 'views': r[3]}
                  for r in cur.fetchall()]
    ts_max = top_slides[0]['views'] if top_slides else 0
    for t in top_slides:
        t['pct'] = round(t['views'] / ts_max * 100) if ts_max else 0

    # 7) AI 튜터(호출수·월별) — chat_logs 부재/오류 시 graceful 0/[] (다른 집계 보호 위해 마지막).
    ai_questions, ai_monthly = 0, []
    try:
        cur.execute(f"""
            SELECT COUNT(cl.id) FROM chat_logs cl
            WHERE cl.institution_id = %s {c_filt} {cl_date}
        """, [inst_id] + c_p + cl_dp)
        ai_questions = cur.fetchone()[0] or 0
        cur.execute(f"""
            SELECT TO_CHAR(DATE_TRUNC('month', cl.created_at),'YYYY-MM'), COUNT(cl.id)
            FROM chat_logs cl
            WHERE cl.institution_id = %s {c_filt} {cl_date}
            GROUP BY 1 ORDER BY 1 LIMIT 12
        """, [inst_id] + c_p + cl_dp)
        ai_monthly = [{'label': r[0], 'count': r[1]} for r in cur.fetchall()]
        am_max = max((x['count'] for x in ai_monthly), default=0)
        for x in ai_monthly:
            x['pct'] = round(x['count'] / am_max * 100) if am_max else 0
    except Exception:
        try:
            cur.connection.rollback()
        except Exception:
            pass
        ai_questions, ai_monthly = 0, []

    return {
        'registered_users': registered, 'by_position': by_position,
        'members': members,
        'active_users': active_users, 'max_seats': max_seats, 'util_pct': util_pct,
        'total_views': total_views, 'ai_questions': ai_questions,
        'per_user_views': per_user_views,
        'monthly_views': monthly_views, 'top_slides': top_slides, 'ai_monthly': ai_monthly,
    }


@app.route('/portal/api/report', methods=['GET'])
@login_required
def portal_report():
    inst_id, err = _portal_guard()
    if err:
        return err
    period = _norm_report_period((request.args.get('period') or _DEFAULT_REPORT_PERIOD).strip())
    subject_code = (request.args.get('subject_code') or 'all').strip()
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            subjects = _subscribed_subjects(cur, inst_id)
            if subject_code != 'all' and subject_code not in subjects:
                # 비구독 과목 = 명시적 거부(빈 결과 아님, P2 동일 §9).
                return jsonify({'success': False, 'error': 'SUBJECT_NOT_SUBSCRIBED',
                                'message': '구독하지 않은 과목입니다'}), 403
            start, end = _portal_report_range(period)
            if subject_code == 'all' and not subjects:
                data = _empty_report()
            else:
                data = _portal_report_data(cur, inst_id, subject_code, subjects, start, end)
        return jsonify({'success': True, 'institution_id': inst_id,
                        'period': period, 'subject_code': subject_code,
                        'subjects': [{'code': c, 'name_ko': n} for c, n in subjects.items()],
                        **data})
    finally:
        release_db_conn(conn)


@app.route('/portal/api/report/export', methods=['GET'])
@login_required
def portal_report_export():
    inst_id, err = _portal_guard()
    if err:
        return err
    period = _norm_report_period((request.args.get('period') or _DEFAULT_REPORT_PERIOD).strip())
    subject_code = (request.args.get('subject_code') or 'all').strip()
    fmt = (request.args.get('format') or 'xlsx').strip().lower()
    if fmt != 'xlsx':
        # PDF 는 클라이언트 window.print() (서버 PDF 생성 금지 §13-1 한국어 폰트 한계).
        return jsonify({'success': False, 'error': 'BAD_FORMAT',
                        'message': 'xlsx만 지원합니다(PDF는 인쇄/PDF 저장 사용)'}), 400
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            subjects = _subscribed_subjects(cur, inst_id)
            if subject_code != 'all' and subject_code not in subjects:
                return jsonify({'success': False, 'error': 'SUBJECT_NOT_SUBSCRIBED',
                                'message': '구독하지 않은 과목입니다'}), 403
            inst_name = inst_id
            cur.execute("SELECT name_ko FROM institutions WHERE id = %s", (inst_id,))
            row = cur.fetchone()
            if row and row[0]:
                inst_name = row[0]
            start, end = _portal_report_range(period)
            data = (_empty_report() if (subject_code == 'all' and not subjects)
                    else _portal_report_data(cur, inst_id, subject_code, subjects, start, end))
    finally:
        release_db_conn(conn)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return jsonify({'success': False, 'error': 'NO_OPENPYXL', 'message': 'openpyxl 미설치'}), 500

    import io
    from flask import send_file as _sf
    period_label = {'1m': '최근 1개월', '3m': '최근 3개월', '6m': '최근 6개월',
                    'all': '전체 기간'}.get(period, period)
    subj_label = '전체 과목' if subject_code == 'all' else f"{subjects.get(subject_code, subject_code)}({subject_code})"
    HDR = Font(bold=True, color='FFFFFF')
    FILL = PatternFill('solid', fgColor='1A2238')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '요약'
    ws.append([_xlsx_safe(f'{inst_name} 이용 리포트')])
    ws['A1'].font = Font(bold=True, size=13)
    ws.append([_xlsx_safe(f'기간: {period_label}  /  과목: {subj_label}')])
    ws.append([])
    ws.append(['지표', '수치'])
    for cell in ws[4]:
        cell.font = HDR
        cell.fill = FILL
    pos = data['by_position']
    pos_txt = ' · '.join(f'{k} {v}' for k, v in pos.items()) if pos else '—'
    ws.append(['등록 이용자', data['registered_users']])
    ws.append(['  지위별', _xlsx_safe(pos_txt)])
    ws.append(['활성 사용자', data['active_users']])
    ws.append(['최대 좌석', data['max_seats']])
    ws.append(['좌석 소진율 (%)', data['util_pct']])
    ws.append(['구성원 — 활성/미인증/비활성',
               _xlsx_safe(f"{data['members']['active']} / {data['members']['unverified']} / {data['members']['inactive']}")])
    ws.append(['슬라이드 총 조회수', data['total_views']])
    ws.append(['1인당 평균 조회수', data['per_user_views']])
    ws.append(['AI 튜터 호출수', data['ai_questions']])
    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 22

    ws2 = wb.create_sheet('월별 조회수')
    ws2.append(['월', '조회수'])
    for cell in ws2[1]:
        cell.font = HDR
        cell.fill = FILL
    for m in data['monthly_views']:
        ws2.append([_xlsx_safe(m['label']), m['count']])
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 12

    ws3 = wb.create_sheet('인기 슬라이드')
    ws3.append(['슬라이드 ID', '제목', '염색', '조회수'])
    for cell in ws3[1]:
        cell.font = HDR
        cell.fill = FILL
    for t in data['top_slides']:
        ws3.append([_xlsx_safe(t['id']), _xlsx_safe(t['title']), _xlsx_safe(t['stain']), t['views']])
    for col, wdt in zip('ABCD', (20, 32, 12, 10)):
        ws3.column_dimensions[col].width = wdt

    ws4 = wb.create_sheet('AI 월별 호출')
    ws4.append(['월', '호출수'])
    for cell in ws4[1]:
        cell.font = HDR
        cell.fill = FILL
    for x in data['ai_monthly']:
        ws4.append([_xlsx_safe(x['label']), x['count']])
    ws4.column_dimensions['A'].width = 12
    ws4.column_dimensions['B'].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'SlideAtlas_report_{inst_id}_{subject_code}_{period}.xlsx'
    return _sf(buf, as_attachment=True, download_name=fname,
               mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ═════════════════════════════════════════════════════════════════════════════
# 교수 수업 페이지(LMS) — 2단계 백엔드 (§21 / §8 단일 게이트 / §15-7 개인정보)
#   ★ 불변 원칙(어기면 reject):
#   1) 권한 분기는 users.position 기반. 수업 개설=교수만, 편집=교수(소유)·조교(위임)만.
#      role(viewer/admin)은 LMS 권한에 쓰지 않는다. 학생/행정직원은 개설·편집 불가.
#   2) 수업은 접근 게이트가 아니다 — 어떤 course API도 슬라이드 접근을 새로 부여하지 않는다.
#      슬라이드 열람 판정은 오직 _slide_access_allowed(불변). 슬라이드 '배치' 시 그 슬라이드가
#      현재 사용자(교수)의 과목 구독 범위인지 _slide_access_allowed로 검증해 비구독/미배포 거부(403).
#   3) scope 강제: institution_id·subject_code는 g.* 에서만. body/쿼리 미참조(IDOR 차단).
#      course 소유·위임 검증은 매 요청 DB 재조회.
#   4) 개인정보: 수업 통계는 익명 집계만 — 학생 개별 활동(이름+접속)을 같은 행에 묶지 않는다(§15-7).
# ═════════════════════════════════════════════════════════════════════════════

def _course_position(cur, user_id):
    """현재 user의 지위(users.position)를 DB에서 재조회(권위, 매 요청). LMS 권한 분기 근거(§6-4·§21)."""
    cur.execute("SELECT position FROM users WHERE id = %s", (user_id,))
    row = cur.fetchone()
    return (row[0] if row else None)


def _course_owner_or_assistant(cur, course_id):
    """course 편집 권한 판정(매 요청 DB 재조회). 반환 (ok, role_in_course, err).

    - course가 g.institution_id·g.subject_code 소속이 아니면 403(scope 격리·IDOR 차단, §9).
    - 현재 user가 교수(courses.professor_user_id 일치 AND users.position=='교수') → ('professor'),
      또는 위임 조교(course_assistants 행 존재 AND users.position=='조교') → ('assistant'). 아니면 403.
    err 는 (json_response, status) 튜플 또는 None.

    ★ [외부검증 수정1] 현재 users.position 을 같은 cursor(같은 트랜잭션)에서 재조회해 교차 검증한다.
      소유/위임 행이 남아 있어도 지위가 강등(교수→학생, 조교 박탈)되면 즉시 차단 — 인증 DB 권위 원칙(§8)과
      정합. 상태변경 라우트는 autocommit=False 트랜잭션 안에서 이 판정 후 같은 conn 으로 UPDATE/DELETE 하므로
      권한 SELECT~변경 사이 TOCTOU 창이 줄어든다(완전 제거는 아님 — 아래 보고 참조).
    """
    cur.execute(
        "SELECT institution_id, subject_code, professor_user_id FROM courses WHERE id = %s",
        (course_id,),
    )
    row = cur.fetchone()
    if not row:
        return False, None, (jsonify({'success': False, 'error': 'NOT_FOUND',
                                      'message': '수업을 찾을 수 없습니다'}), 404)
    c_inst, c_subj, prof_id = row
    # scope: institution_id·subject_code 는 g 에서만 — 다른 기관/과목 course 는 존재를 숨기고 403.
    if c_inst != getattr(g, 'institution_id', None) or c_subj != getattr(g, 'subject_code', None):
        return False, None, (jsonify({'success': False, 'error': 'FORBIDDEN',
                                      'message': '권한이 없습니다'}), 403)
    # 현재 지위 재조회(권위) — 강등 즉시 반영.
    pos = _course_position(cur, g.user_id)
    # 교수: professor_user_id 일치 + 현재 position=='교수'.
    if prof_id is not None and str(prof_id) == str(g.user_id) and pos == '교수':
        return True, 'professor', None
    # 조교: course_assistants 위임 행 존재 + 현재 position=='조교'.
    cur.execute("SELECT 1 FROM course_assistants WHERE course_id = %s AND user_id = %s",
                (course_id, g.user_id))
    if cur.fetchone() and pos == '조교':
        return True, 'assistant', None
    return False, None, (jsonify({'success': False, 'error': 'FORBIDDEN',
                                  'message': '권한이 없습니다'}), 403)


def _course_in_scope(cur, course_id):
    """course가 g.institution_id·g.subject_code 소속인지 확인(열람 자격 — 수업은 게이트 아님, §21-6).

    반환 (row 또는 None). row = (id, title, semester, professor_user_id).
    같은 기관·같은 과목 좌석 사용자면 등록 여부 무관하게 상세 열람 가능(수업=게이트 아님).
    """
    cur.execute(
        """SELECT id, title, semester, professor_user_id
             FROM courses
            WHERE id = %s AND institution_id = %s AND subject_code = %s""",
        (course_id, getattr(g, 'institution_id', None), getattr(g, 'subject_code', None)),
    )
    return cur.fetchone()


def _forbidden_json(msg='권한이 없습니다'):
    return jsonify({'success': False, 'error': 'FORBIDDEN', 'message': msg}), 403


# ── 교수/조교: 편집 ───────────────────────────────────────────────────────────

@app.route('/api/courses', methods=['POST'])
@login_required
def api_course_create():
    """수업 개설 — position=='교수'만(조교 신규개설 불가, 위임 시 편집만, §21-3).
    subject_code=g.subject_code 고정, professor_user_id=g.user_id (body 미참조, §9)."""
    body = request.get_json(silent=True) or {}
    title = (body.get('title') or '').strip()
    semester = (body.get('semester') or '').strip()
    if not title:
        return jsonify({'success': False, 'error': 'MISSING_FIELDS', 'message': '수업명을 입력하세요'}), 400
    if getattr(g, 'subject_code', None) is None:
        # 과목 좌석이 없는 계정(admin-only 등)은 수업을 가질 과목이 없음.
        return _forbidden_json('과목 소속이 없어 수업을 개설할 수 없습니다')
    conn = get_db_conn()
    try:
        conn.autocommit = False
        with conn.cursor() as cur:
            if _course_position(cur, g.user_id) != '교수':
                conn.rollback()
                return _forbidden_json('수업 개설은 교수만 가능합니다')
            cur.execute(
                """INSERT INTO courses (institution_id, subject_code, professor_user_id, title, semester)
                   VALUES (%s, %s, %s, %s, %s) RETURNING id""",
                (g.institution_id, g.subject_code, g.user_id, title[:200], semester[:20]),
            )
            cid = cur.fetchone()[0]
        conn.commit()
        return jsonify({'success': True, 'course_id': cid})
    except Exception:
        conn.rollback()
        return jsonify({'success': False, 'error': 'SERVER_ERROR', 'message': '처리 중 오류'}), 500
    finally:
        try:
            conn.autocommit = True
        except Exception:
            pass
        release_db_conn(conn)


@app.route('/api/courses/mine', methods=['GET'])
@login_required
def api_courses_mine():
    """내가 개설(교수)했거나 위임받은(조교) 수업 목록 — g.institution_id·g.subject_code scope."""
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT c.id, c.title, c.semester,
                          (c.professor_user_id = %s) AS is_owner,
                          (SELECT COUNT(*) FROM course_enrollments e WHERE e.course_id = c.id)
                     FROM courses c
                    WHERE c.institution_id = %s AND c.subject_code = %s
                      AND (c.professor_user_id = %s
                           OR EXISTS (SELECT 1 FROM course_assistants a
                                       WHERE a.course_id = c.id AND a.user_id = %s))
                    ORDER BY c.created_at DESC""",
                (g.user_id, g.institution_id, g.subject_code, g.user_id, g.user_id),
            )
            courses = [{'id': r[0], 'title': r[1] or '', 'semester': r[2] or '',
                        'role': 'professor' if r[3] else 'assistant',
                        'enrolled_count': r[4] or 0} for r in cur.fetchall()]
        return jsonify({'success': True, 'courses': courses})
    finally:
        release_db_conn(conn)


@app.route('/api/courses/<int:cid>', methods=['PUT'])
@login_required
def api_course_update(cid):
    """수업명/학기 수정 — 교수(소유) 또는 위임 조교."""
    body = request.get_json(silent=True) or {}
    title = (body.get('title') or '').strip()
    semester = (body.get('semester') or '').strip()
    if not title:
        return jsonify({'success': False, 'error': 'MISSING_FIELDS', 'message': '수업명을 입력하세요'}), 400
    conn = get_db_conn()
    try:
        conn.autocommit = False
        with conn.cursor() as cur:
            ok, _role, err = _course_owner_or_assistant(cur, cid)
            if not ok:
                conn.rollback()
                return err
            cur.execute("UPDATE courses SET title = %s, semester = %s WHERE id = %s",
                        (title[:200], semester[:20], cid))
        conn.commit()
        return jsonify({'success': True})
    except Exception:
        conn.rollback()
        return jsonify({'success': False, 'error': 'SERVER_ERROR', 'message': '처리 중 오류'}), 500
    finally:
        try:
            conn.autocommit = True
        except Exception:
            pass
        release_db_conn(conn)


@app.route('/api/courses/<int:cid>', methods=['DELETE'])
@login_required
def api_course_delete(cid):
    """수업 삭제 — 교수만(조교 불가). weeks/week_slides/assistants/enrollments 는 FK ON DELETE CASCADE."""
    conn = get_db_conn()
    try:
        conn.autocommit = False
        with conn.cursor() as cur:
            ok, role_in_course, err = _course_owner_or_assistant(cur, cid)
            if not ok:
                conn.rollback()
                return err
            if role_in_course != 'professor':
                conn.rollback()
                return _forbidden_json('수업 삭제는 교수만 가능합니다')
            # course_week_slides 는 course_weeks CASCADE 로, 그 외(weeks/assistants/enrollments)는
            # courses CASCADE 로 정리된다(§7 스키마 FK). 안전을 위해 명시 순서로도 정리.
            cur.execute("""DELETE FROM course_week_slides WHERE course_week_id IN
                             (SELECT id FROM course_weeks WHERE course_id = %s)""", (cid,))
            cur.execute("DELETE FROM course_weeks WHERE course_id = %s", (cid,))
            cur.execute("DELETE FROM course_assistants WHERE course_id = %s", (cid,))
            cur.execute("DELETE FROM course_enrollments WHERE course_id = %s", (cid,))
            cur.execute("DELETE FROM courses WHERE id = %s", (cid,))
        conn.commit()
        return jsonify({'success': True})
    except Exception:
        conn.rollback()
        return jsonify({'success': False, 'error': 'SERVER_ERROR', 'message': '처리 중 오류'}), 500
    finally:
        try:
            conn.autocommit = True
        except Exception:
            pass
        release_db_conn(conn)


@app.route('/api/courses/<int:cid>/weeks', methods=['POST'])
@login_required
def api_course_week_add(cid):
    """주차 추가 — 교수·위임 조교. 빈 주차 허용(empty_reason 메모)."""
    body = request.get_json(silent=True) or {}
    try:
        week_number = int(body.get('week_number'))
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': 'MISSING_FIELDS', 'message': '주차 번호가 필요합니다'}), 400
    title = (body.get('title') or '').strip()
    empty_reason = (body.get('empty_reason') or '').strip() or None
    conn = get_db_conn()
    try:
        conn.autocommit = False
        with conn.cursor() as cur:
            ok, _role, err = _course_owner_or_assistant(cur, cid)
            if not ok:
                conn.rollback()
                return err
            cur.execute(
                """INSERT INTO course_weeks (course_id, week_number, title, empty_reason)
                   VALUES (%s, %s, %s, %s) RETURNING id""",
                (cid, week_number, title[:200], empty_reason),
            )
            wid = cur.fetchone()[0]
        conn.commit()
        return jsonify({'success': True, 'week_id': wid})
    except Exception:
        conn.rollback()
        return jsonify({'success': False, 'error': 'SERVER_ERROR', 'message': '처리 중 오류'}), 500
    finally:
        try:
            conn.autocommit = True
        except Exception:
            pass
        release_db_conn(conn)


@app.route('/api/courses/<int:cid>/weeks/<int:wid>', methods=['DELETE'])
@login_required
def api_course_week_delete(cid, wid):
    """주차 삭제 — 교수·위임 조교. 그 주차의 슬라이드 배치(course_week_slides)는 CASCADE."""
    conn = get_db_conn()
    try:
        conn.autocommit = False
        with conn.cursor() as cur:
            ok, _role, err = _course_owner_or_assistant(cur, cid)
            if not ok:
                conn.rollback()
                return err
            # 주차가 그 course 소속인지 확인(다른 course의 주차 wid 삭제 방지).
            cur.execute("SELECT 1 FROM course_weeks WHERE id = %s AND course_id = %s", (wid, cid))
            if not cur.fetchone():
                conn.rollback()
                return jsonify({'success': False, 'error': 'NOT_FOUND', 'message': '주차를 찾을 수 없습니다'}), 404
            cur.execute("DELETE FROM course_week_slides WHERE course_week_id = %s", (wid,))
            cur.execute("DELETE FROM course_weeks WHERE id = %s AND course_id = %s", (wid, cid))
        conn.commit()
        return jsonify({'success': True})
    except Exception:
        conn.rollback()
        return jsonify({'success': False, 'error': 'SERVER_ERROR', 'message': '처리 중 오류'}), 500
    finally:
        try:
            conn.autocommit = True
        except Exception:
            pass
        release_db_conn(conn)


@app.route('/api/courses/<int:cid>/weeks/<int:wid>/slides', methods=['POST'])
@login_required
def api_course_week_slide_add(cid, wid):
    """주차에 슬라이드 배치 — 교수·위임 조교. 주차 내 중복 허용(§21-4).

    ★ 배치 전 _slide_access_allowed(slide_id)로 그 슬라이드가 현재 사용자(교수/조교)의 과목 구독·배포
      범위인지 검증한다. 수업은 접근을 새로 부여하지 않으므로(§8), 비구독/미배포 슬라이드는 배치 거부(403).
      _slide_access_allowed 는 g.subject_code·g.is_special 로 판정 — 편집자도 그 과목 좌석이어야 한다.
    """
    body = request.get_json(silent=True) or {}
    slide_id = (body.get('slide_id') or '').strip()
    if not slide_id:
        return jsonify({'success': False, 'error': 'MISSING_SLIDE', 'message': '슬라이드 ID가 필요합니다'}), 400
    try:
        display_order = int(body.get('display_order', 0))
    except (TypeError, ValueError):
        display_order = 0
    conn = get_db_conn()
    try:
        conn.autocommit = False
        with conn.cursor() as cur:
            ok, _role, err = _course_owner_or_assistant(cur, cid)
            if not ok:
                conn.rollback()
                return err
            cur.execute("SELECT 1 FROM course_weeks WHERE id = %s AND course_id = %s", (wid, cid))
            if not cur.fetchone():
                conn.rollback()
                return jsonify({'success': False, 'error': 'NOT_FOUND', 'message': '주차를 찾을 수 없습니다'}), 404
            # ★ [외부검증 수정1] 과목 정합성 가드 — 접근 게이트(_slide_access_allowed)와 별개 축이다(§21-2
            #   "수업은 특정 과목 안에서만"). is_special 편집자는 _slide_access_allowed/_visible_slides 의
            #   institution·subject 우회로 타 과목(PATH 등) 슬라이드를 HST 수업에 배치할 수 있는데, 그러면
            #   일반 학생 화면에선 단일 게이트에 막혀 깨진 카드/403 비대칭이 생긴다. 따라서 배치 슬라이드의
            #   과목이 이 수업 과목과 같은지 먼저 검사한다(course.subject_code == g.subject_code 는
            #   _course_owner_or_assistant 가 강제하므로 g.subject_code 와 대조). _slide_access_allowed 무수정.
            _sinfo = get_slide_institution(slide_id)
            if _sinfo is None:
                conn.rollback()
                return jsonify({'success': False, 'error': 'SLIDE_NOT_FOUND', 'message': '슬라이드를 찾을 수 없습니다'}), 404
            if _sinfo[1] != getattr(g, 'subject_code', None):
                conn.rollback()
                return jsonify({'success': False, 'error': 'SLIDE_SUBJECT_MISMATCH',
                                'message': '이 수업 과목의 슬라이드만 배치할 수 있습니다'}), 403
            # ★ 슬라이드 접근 단일 게이트로 배치 가능 여부 검증(수업은 게이트 아님 §8).
            allowed, _aerr = _slide_access_allowed(slide_id)
            if not allowed:
                conn.rollback()
                return jsonify({'success': False, 'error': 'SLIDE_NOT_ALLOWED',
                                'message': '구독·배포된 과목 슬라이드만 배치할 수 있습니다'}), 403
            cur.execute(
                """INSERT INTO course_week_slides (course_week_id, slide_id, display_order)
                   VALUES (%s, %s, %s) RETURNING id""",
                (wid, slide_id, display_order),
            )
            link_id = cur.fetchone()[0]
        conn.commit()
        return jsonify({'success': True, 'id': link_id})
    except Exception:
        conn.rollback()
        return jsonify({'success': False, 'error': 'SERVER_ERROR', 'message': '처리 중 오류'}), 500
    finally:
        try:
            conn.autocommit = True
        except Exception:
            pass
        release_db_conn(conn)


@app.route('/api/courses/<int:cid>/weeks/<int:wid>/slides/<int:sid>', methods=['DELETE'])
@login_required
def api_course_week_slide_delete(cid, wid, sid):
    """배치 제거 — 교수·위임 조교. sid = course_week_slides.id(배치 행 id)."""
    conn = get_db_conn()
    try:
        conn.autocommit = False
        with conn.cursor() as cur:
            ok, _role, err = _course_owner_or_assistant(cur, cid)
            if not ok:
                conn.rollback()
                return err
            # 그 배치 행이 이 course의 이 주차 소속인지 확인 후 삭제(IDOR 방지).
            cur.execute(
                """DELETE FROM course_week_slides
                    WHERE id = %s AND course_week_id = %s
                      AND course_week_id IN (SELECT id FROM course_weeks WHERE course_id = %s)""",
                (sid, wid, cid),
            )
            removed = cur.rowcount
        conn.commit()
        if removed == 0:
            return jsonify({'success': False, 'error': 'NOT_FOUND', 'message': '배치를 찾을 수 없습니다'}), 404
        return jsonify({'success': True})
    except Exception:
        conn.rollback()
        return jsonify({'success': False, 'error': 'SERVER_ERROR', 'message': '처리 중 오류'}), 500
    finally:
        try:
            conn.autocommit = True
        except Exception:
            pass
        release_db_conn(conn)


@app.route('/api/courses/<int:cid>/assistants', methods=['POST'])
@login_required
def api_course_assistant_add(cid):
    """조교 위임 — 교수만. 대상이 같은 기관·같은 과목·position=='조교'인지 검증(§21-4)."""
    body = request.get_json(silent=True) or {}
    target = body.get('user_id')
    if target is None:
        return jsonify({'success': False, 'error': 'MISSING_FIELDS', 'message': 'user_id가 필요합니다'}), 400
    conn = get_db_conn()
    try:
        conn.autocommit = False
        with conn.cursor() as cur:
            ok, role_in_course, err = _course_owner_or_assistant(cur, cid)
            if not ok:
                conn.rollback()
                return err
            if role_in_course != 'professor':
                conn.rollback()
                return _forbidden_json('조교 지정은 교수만 가능합니다')
            # 대상 검증: 같은 기관·같은 과목·지위 조교·활성 계정 (scope 는 g 기준, 대상 user 만 id 로 조회).
            #   [외부검증 수정2] status='active' 추가 — locked/pending 계정 위임 거부(활성 정의 통일 §0).
            cur.execute(
                """SELECT 1 FROM users
                    WHERE id = %s AND institution_id = %s AND subject_code = %s
                      AND position = '조교' AND status = 'active'""",
                (target, g.institution_id, g.subject_code),
            )
            if not cur.fetchone():
                conn.rollback()
                return jsonify({'success': False, 'error': 'INVALID_TARGET',
                                'message': '같은 과목의 조교만 지정할 수 있습니다'}), 400
            cur.execute(
                """INSERT INTO course_assistants (course_id, user_id) VALUES (%s, %s)
                   ON CONFLICT (course_id, user_id) DO NOTHING""",
                (cid, target),
            )
        conn.commit()
        return jsonify({'success': True})
    except Exception:
        conn.rollback()
        return jsonify({'success': False, 'error': 'SERVER_ERROR', 'message': '처리 중 오류'}), 500
    finally:
        try:
            conn.autocommit = True
        except Exception:
            pass
        release_db_conn(conn)


@app.route('/api/courses/<int:cid>/assistants/<int:uid>', methods=['DELETE'])
@login_required
def api_course_assistant_remove(cid, uid):
    """위임 해제 — 교수만."""
    conn = get_db_conn()
    try:
        conn.autocommit = False
        with conn.cursor() as cur:
            ok, role_in_course, err = _course_owner_or_assistant(cur, cid)
            if not ok:
                conn.rollback()
                return err
            if role_in_course != 'professor':
                conn.rollback()
                return _forbidden_json('위임 해제는 교수만 가능합니다')
            cur.execute("DELETE FROM course_assistants WHERE course_id = %s AND user_id = %s",
                        (cid, uid))
        conn.commit()
        return jsonify({'success': True})
    except Exception:
        conn.rollback()
        return jsonify({'success': False, 'error': 'SERVER_ERROR', 'message': '처리 중 오류'}), 500
    finally:
        try:
            conn.autocommit = True
        except Exception:
            pass
        release_db_conn(conn)


# ── 교수/조교: 수업 대시보드(명단 + 익명 집계) ────────────────────────────────

@app.route('/api/courses/<int:cid>/roster', methods=['GET'])
@login_required
def api_course_roster(cid):
    """등록 학생 명단 — 교수·위임 조교만. 표시명(roster.name)·이메일·등록일만.
    ★ 개인 활동/접속 데이터는 절대 포함하지 않는다(이름과 활동을 같은 행에 묶지 않음, §15-7)."""
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            ok, _role, err = _course_owner_or_assistant(cur, cid)
            if not ok:
                return err
            # 표시명은 institution_rosters.name(가입 표시명). 활동/접속 컬럼은 SELECT 자체에 없음.
            cur.execute(
                """SELECT COALESCE(r.name, ''), u.email, e.enrolled_at
                     FROM course_enrollments e
                     JOIN users u ON u.id = e.user_id
                     LEFT JOIN institution_rosters r
                       ON lower(r.email) = lower(u.email)
                      AND r.institution_id = u.institution_id
                      AND r.subject_code = u.subject_code
                    WHERE e.course_id = %s
                    ORDER BY r.name, u.email""",
                (cid,),
            )
            students = [{'name': r[0] or '', 'email': r[1],
                         'enrolled_at': r[2].isoformat() if r[2] else None}
                        for r in cur.fetchall()]
        return jsonify({'success': True, 'students': students, 'count': len(students)})
    finally:
        release_db_conn(conn)


@app.route('/api/courses/<int:cid>/stats', methods=['GET'])
@login_required
def api_course_stats(cid):
    """수업 단위 익명 집계만 — 교수·위임 조교만.

    ★ 절대 원칙(§15-7): 학생별 개별 행(user_id·이름·학생별 마지막 접속)을 반환하지 않는다.
      오직 수업 전체 집계 숫자만 — 개인 지목 불가.
    ⚠ access_logs 는 '이 수업을 통한 열람'이 아니라 '등록 학생의 해당 과목 활동'이다(수업=게이트 아님 §8).
      집계는 al.institution_id·al.subject_code 스냅샷 기준(§15-7), subject_code NULL 과거 로그 제외.
    """
    from datetime import timedelta
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            ok, _role, err = _course_owner_or_assistant(cur, cid)
            if not ok:
                return err
            inst = g.institution_id
            subj = g.subject_code
            since = _today_kst() - timedelta(days=7)

            # 1) 등록 인원
            cur.execute("SELECT COUNT(*) FROM course_enrollments WHERE course_id = %s", (cid,))
            enrolled_count = cur.fetchone()[0] or 0

            # 2) 최근 7일 활동 유무 분포 — '명수'만(학생별 행 없음). 스냅샷 inst·subject 기준.
            cur.execute(
                """SELECT
                     COUNT(*) FILTER (WHERE recent),
                     COUNT(*) FILTER (WHERE NOT recent)
                   FROM (
                     SELECT EXISTS (
                         SELECT 1 FROM access_logs al
                          WHERE al.user_id = e.user_id
                            AND al.institution_id = %s
                            AND al.subject_code = %s
                            AND al.subject_code IS NOT NULL
                            AND al.accessed_at >= %s
                     ) AS recent
                       FROM course_enrollments e
                      WHERE e.course_id = %s
                   ) t""",
                (inst, subj, since, cid),
            )
            ar = cur.fetchone()
            active_recent_count = ar[0] or 0
            inactive_count = ar[1] or 0

            # 3) 슬라이드 열람률 — 배치 슬라이드 중 등록 학생이 1회+ 열람한 비율(익명 비율).
            cur.execute(
                """SELECT
                     COUNT(DISTINCT cws.slide_id),
                     COUNT(DISTINCT cws.slide_id) FILTER (WHERE EXISTS (
                       SELECT 1 FROM access_logs al
                        JOIN course_enrollments e ON e.user_id = al.user_id AND e.course_id = %s
                        WHERE al.slide_id = cws.slide_id
                          AND al.institution_id = %s
                          AND al.subject_code = %s
                          AND al.subject_code IS NOT NULL
                     ))
                   FROM course_week_slides cws
                   JOIN course_weeks cw ON cw.id = cws.course_week_id
                  WHERE cw.course_id = %s""",
                (cid, inst, subj, cid),
            )
            ps = cur.fetchone()
            placed = ps[0] or 0
            viewed = ps[1] or 0
            slide_view_rate = round(viewed / placed * 100) if placed > 0 else 0

        return jsonify({'success': True, 'stats': {
            'enrolled_count': enrolled_count,
            'active_recent_count': active_recent_count,
            'inactive_count': inactive_count,
            'placed_slide_count': placed,
            'viewed_slide_count': viewed,
            'slide_view_rate': slide_view_rate,
        }})
    finally:
        release_db_conn(conn)


# ── 학생: 수강 ────────────────────────────────────────────────────────────────

@app.route('/api/courses/available', methods=['GET'])
@login_required
def api_courses_available():
    """같은 기관·같은 과목 공개 수업 전체 + 등록 여부 플래그(§21-5)."""
    if getattr(g, 'subject_code', None) is None:
        return jsonify({'success': True, 'courses': []})
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT c.id, c.title, c.semester,
                          EXISTS (SELECT 1 FROM course_enrollments e
                                   WHERE e.course_id = c.id AND e.user_id = %s)
                     FROM courses c
                    WHERE c.institution_id = %s AND c.subject_code = %s
                    ORDER BY c.created_at DESC""",
                (g.user_id, g.institution_id, g.subject_code),
            )
            courses = [{'id': r[0], 'title': r[1] or '', 'semester': r[2] or '',
                        'enrolled': bool(r[3])} for r in cur.fetchall()]
        return jsonify({'success': True, 'courses': courses})
    finally:
        release_db_conn(conn)


@app.route('/api/courses/enrolled', methods=['GET'])
@login_required
def api_courses_enrolled():
    """내가 등록한 수업 목록(§21-5 내 수업)."""
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT c.id, c.title, c.semester
                     FROM course_enrollments e
                     JOIN courses c ON c.id = e.course_id
                    WHERE e.user_id = %s
                      AND c.institution_id = %s AND c.subject_code = %s
                    ORDER BY e.enrolled_at DESC""",
                (g.user_id, g.institution_id, g.subject_code),
            )
            courses = [{'id': r[0], 'title': r[1] or '', 'semester': r[2] or ''}
                       for r in cur.fetchall()]
        return jsonify({'success': True, 'courses': courses})
    finally:
        release_db_conn(conn)


@app.route('/api/courses/<int:cid>/enroll', methods=['POST'])
@login_required
def api_course_enroll(cid):
    """수강 등록 — 자유(승인 불필요), 같은 기관·과목 수업만. 중복 등록은 멱등(§21-5).

    ★ [외부검증 수정4] 등록은 콘텐츠 소비 지위(position∈{학생,조교})만 허용. 교수·행정직원·position NULL 거부.
      (해지 DELETE 는 기등록분 정리를 위해 position 무관 허용.)"""
    conn = get_db_conn()
    try:
        conn.autocommit = False
        with conn.cursor() as cur:
            if _course_in_scope(cur, cid) is None:
                conn.rollback()
                return jsonify({'success': False, 'error': 'NOT_FOUND', 'message': '수업을 찾을 수 없습니다'}), 404
            pos = _course_position(cur, g.user_id)
            if pos not in ('학생', '조교'):
                conn.rollback()
                return jsonify({'success': False, 'error': 'ENROLL_NOT_ALLOWED',
                                'message': '수강 등록은 학생·조교만 가능합니다'}), 403
            cur.execute(
                """INSERT INTO course_enrollments (course_id, user_id) VALUES (%s, %s)
                   ON CONFLICT (course_id, user_id) DO NOTHING""",
                (cid, g.user_id),
            )
        conn.commit()
        return jsonify({'success': True})
    except Exception:
        conn.rollback()
        return jsonify({'success': False, 'error': 'SERVER_ERROR', 'message': '처리 중 오류'}), 500
    finally:
        try:
            conn.autocommit = True
        except Exception:
            pass
        release_db_conn(conn)


@app.route('/api/courses/<int:cid>/enroll', methods=['DELETE'])
@login_required
def api_course_unenroll(cid):
    """수강 해지 — 멱등(이미 미등록이어도 성공).

    ★ [외부검증 수정2] 삭제 전 cid 가 현재 scope(g.institution_id·g.subject_code) 소속인지 확인 —
      cross-scope 수강행 삭제 차단(POST /enroll 와 동일하게 비소속이면 404). position 무관(기등록 정리)."""
    conn = get_db_conn()
    try:
        conn.autocommit = False
        with conn.cursor() as cur:
            if _course_in_scope(cur, cid) is None:
                conn.rollback()
                return jsonify({'success': False, 'error': 'NOT_FOUND', 'message': '수업을 찾을 수 없습니다'}), 404
            cur.execute("DELETE FROM course_enrollments WHERE course_id = %s AND user_id = %s",
                        (cid, g.user_id))
        conn.commit()
        return jsonify({'success': True})
    except Exception:
        conn.rollback()
        return jsonify({'success': False, 'error': 'SERVER_ERROR', 'message': '처리 중 오류'}), 500
    finally:
        try:
            conn.autocommit = True
        except Exception:
            pass
        release_db_conn(conn)


@app.route('/api/courses/<int:cid>', methods=['GET'])
@login_required
def api_course_detail(cid):
    """수업 상세(주차+슬라이드 메타) — 열람 자격 = 같은 기관·같은 과목 좌석 사용자(등록 여부 무관, §21-6).

    ★ 수업은 게이트가 아니다 — 미등록 학생도 조회 가능. 슬라이드 카드 클릭→/viewer 는 기존 게이트가
      최종 판정한다(여기서 슬라이드 접근을 새로 부여하지 않는다, §8). 슬라이드 메타는 ID·제목·염색만.
    """
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            course = _course_in_scope(cur, cid)
            if course is None:
                return jsonify({'success': False, 'error': 'NOT_FOUND', 'message': '수업을 찾을 수 없습니다'}), 404
            enrolled = False
            cur.execute("SELECT 1 FROM course_enrollments WHERE course_id = %s AND user_id = %s",
                        (cid, g.user_id))
            enrolled = cur.fetchone() is not None
            # 주차 + 배치 슬라이드 메타(타일·토큰 발급 없음 — 카탈로그 메타만).
            # ★ [외부검증 수정3] deploy_status='deployed' 슬라이드만 결과에 포함(_visible_slides 필터 원칙).
            #   배치(course_week_slides)는 남아 있어도 qc_pending/rejected 등 미배포 슬라이드의 메타데이터는
            #   노출하지 않는다. 미배포 배치는 cws LEFT JOIN 의 ON 절에서 제외해 '빈 주차'로 표시되게 한다
            #   (주차 행 자체는 유지). 일반 사용자 경로는 무조건 deployed 만(편집자 미배포 표시는 3단계).
            cur.execute(
                """SELECT cw.id, cw.week_number, cw.title, cw.empty_reason,
                          cws.id, cws.slide_id, cws.display_order, s.title_ko, s.stain, s.organ
                     FROM course_weeks cw
                     LEFT JOIN course_week_slides cws
                            ON cws.course_week_id = cw.id
                           AND cws.slide_id IN (SELECT id FROM slides WHERE deploy_status = 'deployed')
                     LEFT JOIN slides s ON s.id = cws.slide_id
                    WHERE cw.course_id = %s
                    ORDER BY cw.week_number, cws.display_order, cws.id""",
                (cid,),
            )
            weeks = {}
            order = []
            for r in cur.fetchall():
                wkid = r[0]
                if wkid not in weeks:
                    weeks[wkid] = {'id': wkid, 'week_number': r[1], 'title': r[2] or '',
                                   'empty_reason': r[3], 'slides': []}
                    order.append(wkid)
                if r[4] is not None:   # 배치 슬라이드 존재
                    weeks[wkid]['slides'].append({
                        'link_id': r[4], 'slide_id': r[5], 'display_order': r[6],
                        'title_ko': r[7] or r[5], 'stain': r[8] or '',
                        # organ(자유텍스트 = load_slides 'system' 표시축, §6-1) — 표시용 메타만(타일·토큰 없음).
                        'organ': r[9] or '',
                    })
            # 표시용 헤더 메타(교수명·과목명) — 게이트 무관 표시 필드. course[3]=professor_user_id.
            professor_name = ''
            if course[3] is not None:
                cur.execute(
                    """SELECT COALESCE(r.name, '')
                         FROM users u
                         LEFT JOIN institution_rosters r
                           ON lower(r.email) = lower(u.email)
                          AND r.institution_id = u.institution_id
                          AND r.subject_code = u.subject_code
                        WHERE u.id = %s""",
                    (course[3],),
                )
                prow = cur.fetchone()
                professor_name = (prow[0] if prow else '') or ''
            cur.execute("SELECT COALESCE(name_ko, %s) FROM subject_codes WHERE code = %s",
                        (g.subject_code, g.subject_code))
            srow = cur.fetchone()
            subject_name = (srow[0] if srow else g.subject_code) or ''
        return jsonify({'success': True, 'course': {
            'id': course[0], 'title': course[1] or '', 'semester': course[2] or '',
            'professor_name': professor_name, 'subject_code': g.subject_code,
            'subject_name': subject_name,
            'enrolled': enrolled, 'weeks': [weeks[w] for w in order],
        }})
    finally:
        release_db_conn(conn)


# ── 학생: 즐겨찾기·열람기록 (3단계-B 마이페이지) ──────────────────────────────
#   ★ scope = g.user_id 강제(본인 것만). 타인 user_id 를 body/쿼리/경로로 받지 않는다(IDOR 불가).
#     표시 목록은 deploy_status='deployed' AND 본인 과목(g.subject_code)으로 한정 —
#     단일 게이트 표시 기준과 정합(타 과목/미배포 메타 누수 차단). 슬라이드 접근 판정은 손대지 않는다.

def _uid_int():
    """g.user_id 를 INT 로 — favorites/access_logs FK 가 INT. 실패 시 None."""
    try:
        return int(getattr(g, 'user_id', None))
    except (TypeError, ValueError):
        return None


@app.route('/api/favorites', methods=['GET'])
@login_required
def api_favorites_list():
    """내 즐겨찾기(개인 북마크, 수업 무관 §21-7). 본인·배포·본인 과목만(표시 메타·타일토큰 없음)."""
    uid = _uid_int()
    if uid is None:
        return jsonify({'success': True, 'favorites': []})
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT s.id, s.title_ko, s.organ, s.stain
                     FROM favorites f
                     JOIN slides s ON s.id = f.slide_id
                    WHERE f.user_id = %s
                      AND s.deploy_status = 'deployed'
                      AND s.subject_code = %s
                    ORDER BY f.created_at DESC""",
                (uid, getattr(g, 'subject_code', None)),
            )
            favs = [{'slide_id': r[0], 'title_ko': r[1] or r[0],
                     'organ': r[2] or '', 'stain': r[3] or ''} for r in cur.fetchall()]
        return jsonify({'success': True, 'favorites': favs})
    finally:
        release_db_conn(conn)


@app.route('/api/favorites/<slide_id>', methods=['POST'])
@login_required
def api_favorite_add(slide_id):
    """즐겨찾기 추가 — 본인이 접근 가능한 슬라이드만(_slide_access_allowed 게이트 읽기, 새 권한 없음).

    게이트를 통과하지 못하는 슬라이드는 북마크 불가(존재 probing 차단). 중복은 멱등(ON CONFLICT).
    """
    uid = _uid_int()
    if uid is None:
        return jsonify({'success': False, 'error': 'TOKEN_INVALID'}), 401
    allowed, _aerr = _slide_access_allowed(slide_id)
    if not allowed:
        # ★ [외부검증 수정1·Med] 존재 oracle 차단: 게이트의 404/403(없는 ID vs 존재하나 접근불가)을
        #   그대로 흘리지 않고 항상 동일한 단일 응답으로 접는다 — probing 불가. 게이트 자체는 무수정.
        return jsonify({'success': False, 'error': 'SLIDE_NOT_ACCESSIBLE',
                        'message': '접근할 수 없는 슬라이드입니다.'}), 403
    conn = get_db_conn()
    try:
        conn.autocommit = False
        with conn.cursor() as cur:
            cur.execute(
                """INSERT INTO favorites (user_id, slide_id) VALUES (%s, %s)
                   ON CONFLICT (user_id, slide_id) DO NOTHING""",
                (uid, slide_id),
            )
        conn.commit()
        return jsonify({'success': True})
    except Exception:
        conn.rollback()
        return jsonify({'success': False, 'error': 'SERVER_ERROR', 'message': '처리 중 오류'}), 500
    finally:
        try:
            conn.autocommit = True
        except Exception:
            pass
        release_db_conn(conn)


@app.route('/api/favorites/<slide_id>', methods=['DELETE'])
@login_required
def api_favorite_remove(slide_id):
    """즐겨찾기 해제 — 본인 행만 삭제(멱등). 접근 게이트 무관(자기 북마크 정리)."""
    uid = _uid_int()
    if uid is None:
        return jsonify({'success': False, 'error': 'TOKEN_INVALID'}), 401
    conn = get_db_conn()
    try:
        conn.autocommit = False
        with conn.cursor() as cur:
            cur.execute("DELETE FROM favorites WHERE user_id = %s AND slide_id = %s",
                        (uid, slide_id))
        conn.commit()
        return jsonify({'success': True})
    except Exception:
        conn.rollback()
        return jsonify({'success': False, 'error': 'SERVER_ERROR', 'message': '처리 중 오류'}), 500
    finally:
        try:
            conn.autocommit = True
        except Exception:
            pass
        release_db_conn(conn)


@app.route('/api/me/history', methods=['GET'])
@login_required
def api_my_history():
    """내 최근 열람 기록(§21-8). 본인 access_logs 만(남의 활동 아님 — §15-7 위반 아님).

    scope = g.user_id 강제. ★ [외부검증 수정2·Med] 과목 귀속은 access_logs 의 열람 시점 스냅샷
    (al.institution_id·al.subject_code)으로 필터 — 현재 slides.subject_code 가 아니라(사용자/슬라이드
    과목 이동 시 과거 로그 재분류=시간축 오염 차단, §15-7·v3.14 P3 원칙). subject_code NULL 과거 로그는
    제외(과목 귀속 불명). slides 조인은 deployed 필터·제목 표시용. 슬라이드별 최신 1건·최대 15.
    """
    uid = _uid_int()
    if uid is None:
        return jsonify({'success': True, 'history': []})
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT s.id, s.title_ko, s.organ, s.stain, MAX(al.accessed_at) AS last_at
                     FROM access_logs al
                     JOIN slides s ON s.id = al.slide_id
                    WHERE al.user_id = %s
                      AND al.institution_id = %s
                      AND al.subject_code = %s
                      AND al.subject_code IS NOT NULL
                      AND s.deploy_status = 'deployed'
                    GROUP BY s.id, s.title_ko, s.organ, s.stain
                    ORDER BY last_at DESC
                    LIMIT 15""",
                (uid, getattr(g, 'institution_id', None), getattr(g, 'subject_code', None)),
            )
            hist = [{'slide_id': r[0], 'title_ko': r[1] or r[0],
                     'organ': r[2] or '', 'stain': r[3] or '',
                     'accessed_at': r[4].isoformat() if r[4] else None}
                    for r in cur.fetchall()]
        return jsonify({'success': True, 'history': hist})
    finally:
        release_db_conn(conn)


# ── 교수/조교 프론트 표시용 읽기 API (3단계-A) ────────────────────────────────
#   ★ 모두 GET·읽기 전용. 새 접근/권한 판정 로직 없음 — 기존 헬퍼(_course_owner_or_assistant·
#     _visible_slides·POST assistants 의 대상검증식)를 그대로 재사용한다. 슬라이드 접근 판정(§8)·
#     LMS 권한(§21)은 손대지 않는다. 타일·토큰 발급 없음(카탈로그/명단 메타만).

@app.route('/api/courses/<int:cid>/available-slides', methods=['GET'])
@login_required
def api_course_available_slides(cid):
    """슬라이드 배치 모달용 — 편집자(교수·위임 조교)가 접근 가능한 슬라이드 메타 목록.

    권한: _course_owner_or_assistant(편집권·scope) 재사용. 목록: _visible_slides(단일 게이트와
    동일 기준, g 기반) 재사용 — 새 필터 없음. 배치 자체는 기존 POST .../slides 가 _slide_access_allowed
    로 재검증하므로(§8), 이 목록은 표시 후보일 뿐 접근을 부여하지 않는다.
    """
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            ok, _role, err = _course_owner_or_assistant(cur, cid)
            if not ok:
                return err
        data = load_slides()
        vis = _visible_slides(data.get('slides', []))
        # ★ [외부검증 수정1] 후보·배치 두 경로가 같은 과목 집합이 되도록, 이 수업 과목(=g.subject_code)
        #   슬라이드만 후보로 노출한다. 일반 편집자는 _visible_slides 가 이미 과목을 강제하므로 무영향이고,
        #   is_special 편집자(과목·institution 우회)에서만 타 과목 슬라이드를 후보에서 배제한다(배치 가드와 정합).
        subj = getattr(g, 'subject_code', None)
        slides = [{'id': s['id'], 'title_ko': s.get('title_ko', s['id']),
                   'organ': s.get('system', ''), 'stain': s.get('stain', '')}
                  for s in vis if s.get('subject_code') == subj]
        return jsonify({'success': True, 'slides': slides})
    finally:
        release_db_conn(conn)


@app.route('/api/courses/<int:cid>/assistants', methods=['GET'])
@login_required
def api_course_assistants_list(cid):
    """현재 위임 조교 목록(표시명·이메일·user_id) — 편집권자 열람(표시용 GET).

    권한·scope = _course_owner_or_assistant 재사용. 표시명은 institution_rosters.name.
    """
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            ok, _role, err = _course_owner_or_assistant(cur, cid)
            if not ok:
                return err
            cur.execute(
                """SELECT u.id, COALESCE(r.name, ''), u.email
                     FROM course_assistants a
                     JOIN users u ON u.id = a.user_id
                     LEFT JOIN institution_rosters r
                       ON lower(r.email) = lower(u.email)
                      AND r.institution_id = u.institution_id
                      AND r.subject_code = u.subject_code
                    WHERE a.course_id = %s
                    ORDER BY r.name, u.email""",
                (cid,),
            )
            assistants = [{'user_id': r[0], 'name': r[1] or '', 'email': r[2]}
                          for r in cur.fetchall()]
        return jsonify({'success': True, 'assistants': assistants})
    finally:
        release_db_conn(conn)


@app.route('/api/courses/<int:cid>/assistant-candidates', methods=['GET'])
@login_required
def api_course_assistant_candidates(cid):
    """조교 위임 후보 검색 — 교수만(표시용 GET).

    후보 기준 = 기존 POST assistants 의 대상검증식과 동일: 같은 기관(g.institution_id)·같은 과목
    (g.subject_code)·position='조교'. 이미 위임된 사용자는 제외. scope 는 g 에서만(IDOR 차단, §9).
    """
    q = (request.args.get('q') or '').strip()
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            ok, role_in_course, err = _course_owner_or_assistant(cur, cid)
            if not ok:
                return err
            if role_in_course != 'professor':
                return _forbidden_json('조교 검색은 교수만 가능합니다')
            like = '%' + q + '%'
            cur.execute(
                """SELECT u.id, COALESCE(r.name, ''), u.email
                     FROM users u
                     LEFT JOIN institution_rosters r
                       ON lower(r.email) = lower(u.email)
                      AND r.institution_id = u.institution_id
                      AND r.subject_code = u.subject_code
                    WHERE u.institution_id = %s AND u.subject_code = %s AND u.position = '조교'
                      AND u.status = 'active'   -- [외부검증 수정2] locked/pending 비노출(활성 정의 통일 §0)
                      AND NOT EXISTS (SELECT 1 FROM course_assistants a
                                       WHERE a.course_id = %s AND a.user_id = u.id)
                      AND (%s = '' OR COALESCE(r.name, '') ILIKE %s OR u.email ILIKE %s)
                    ORDER BY r.name, u.email
                    LIMIT 20""",
                (g.institution_id, g.subject_code, cid, q, like, like),
            )
            cands = [{'user_id': r[0], 'name': r[1] or '', 'email': r[2]}
                     for r in cur.fetchall()]
        return jsonify({'success': True, 'candidates': cands})
    finally:
        release_db_conn(conn)


# ── 교수/조교 페이지 라우트 (3단계-A) ─────────────────────────────────────────
#   HTML 셸만 렌더 — 데이터는 프론트가 위 기존 API 를 interceptor.js(CSRF 자동) 로 호출한다.
#   권한 게이트는 전부 기존 헬퍼 재사용(새 판정 경로 없음): position(_course_position)·
#   편집권(_course_owner_or_assistant). 비편집자/타 기관·과목은 403 페이지.

def _lms_403_page():
    """LMS 페이지 권한 거부 — JSON 대신 사람용 403 HTML(학생 홈 링크)."""
    html = (
        '<!DOCTYPE html><html lang="ko"><head><meta charset="UTF-8">'
        '<title>SlideAtlas — 접근 권한 없음</title>'
        '<style>body{background:#FAF8F5;color:#0F1F3D;font-family:"Noto Sans KR",sans-serif;'
        'display:flex;align-items:center;justify-content:center;height:100vh;margin:0}'
        '.box{text-align:center}.box h2{font-size:20px;margin-bottom:10px}'
        '.box p{color:#8A91A0;font-size:14px;margin-bottom:18px}'
        '.box a{color:#2C7E99;text-decoration:none;font-weight:600;font-size:14px}</style>'
        '</head><body><div class="box"><h2>접근 권한이 없습니다</h2>'
        '<p>이 수업의 편집 권한이 없거나 다른 기관·과목의 수업입니다.</p>'
        '<a href="/home">← 학습 홈으로</a></div></body></html>'
    )
    return html, 403


def _page_course_role(cid):
    """페이지용 편집권 판정 — (role 또는 None). _course_owner_or_assistant 재사용(새 판정 없음)."""
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            ok, role, _err = _course_owner_or_assistant(cur, cid)
            return role if ok else None
    finally:
        release_db_conn(conn)


@app.route('/teacher/courses')
@page_login_required
def teacher_courses_page():
    """교수/조교 수업 목록 — position∈{교수,조교}만, 그 외 학습 홈으로."""
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            pos = _course_position(cur, g.user_id)
    finally:
        release_db_conn(conn)
    if pos not in ('교수', '조교'):
        return redirect('/home')
    return render_template('teacher_courses.html', is_professor=(pos == '교수'))


@app.route('/teacher/course/<int:cid>')
@page_login_required
def teacher_course_edit_page(cid):
    """수업 편집(주차 구성) — 편집권자(교수·위임 조교)만."""
    role = _page_course_role(cid)
    if role is None:
        return _lms_403_page()
    return render_template('course_edit.html', cid=cid, active_tab='weeks',
                           role_in_course=role, subject_code=g.subject_code)


@app.route('/teacher/course/<int:cid>/assistants')
@page_login_required
def teacher_course_assistants_page(cid):
    """조교 지정 — 교수만(조교 위임은 교수 권한, §21-3)."""
    role = _page_course_role(cid)
    if role is None:
        return _lms_403_page()
    if role != 'professor':
        return _lms_403_page()
    return render_template('assistants.html', cid=cid, active_tab='assistants',
                           role_in_course=role, subject_code=g.subject_code)


@app.route('/teacher/course/<int:cid>/dashboard')
@page_login_required
def teacher_course_dashboard_page(cid):
    """수업 대시보드 — 편집권자(교수·위임 조교)만. 명단/익명집계는 프론트가 기존 roster·stats 호출."""
    role = _page_course_role(cid)
    if role is None:
        return _lms_403_page()
    return render_template('course_dashboard.html', cid=cid, active_tab='dashboard',
                           role_in_course=role, subject_code=g.subject_code)


# ── 학생 LMS 페이지 라우트 (3단계-B) ──────────────────────────────────────────
#   HTML 셸만 렌더 — 데이터·권한 판정은 프론트가 기존 학생 API(GET /api/courses/<cid>,
#   enroll POST/DELETE)를 interceptor.js(CSRF)로 호출해 받는다. 새 권한 경로 없음.
#   수업은 게이트가 아니다(§21-6): 미등록 학생도 상세 열람 가능, 뷰어 진입은 단일 게이트가 최종 판정.

@app.route('/course/<int:cid>')
@page_login_required
def course_detail_page(cid):
    """학생 수업 상세 — 로그인 사용자 누구나 셸 렌더(scope·존재는 GET /api/courses/<cid>가 판정).

    admin-only(콘텐츠 비소비자, 좌석 0·subject 없음)는 학생 수업 화면 대신 포털로(홈과 동일 기준 §6-4).
    """
    if g.role == 'admin' and g.subject_code is None:
        return redirect('/portal')
    return render_template('course.html', cid=cid)


@app.route('/mypage')
@page_login_required
def mypage():
    """마이페이지 — 프로필(읽기전용)·비밀번호 변경·즐겨찾기·최근 열람 기록(§21-8).

    프로필은 서버 렌더 컨텍스트로 전달(읽기전용). 즐겨찾기·기록은 프론트가 GET /api/favorites·
    /api/me/history 를 호출(둘 다 g.user_id scope 강제 — 본인 것만, IDOR 차단).
    """
    display_name, subject_name, institution_name, position = '', '', '', ''
    email = ''
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT u.email, COALESCE(r.name, ''), u.position,
                          COALESCE(i.name_ko, u.institution_id, ''),
                          COALESCE(sc.name_ko, u.subject_code, '')
                     FROM users u
                     LEFT JOIN institution_rosters r
                       ON lower(r.email) = lower(u.email)
                      AND r.institution_id = u.institution_id
                      AND r.subject_code = u.subject_code
                     LEFT JOIN institutions i ON i.id = u.institution_id
                     LEFT JOIN subject_codes sc ON sc.code = u.subject_code
                    WHERE u.id = %s""",
                (g.user_id,),
            )
            row = cur.fetchone()
            if row:
                email = row[0] or ''
                display_name = row[1] or ''
                position = row[2] or ''
                institution_name = row[3] or ''
                subject_name = row[4] or ''
    finally:
        release_db_conn(conn)
    return render_template('mypage.html',
        display_name=display_name, email=email, position=position,
        institution_name=institution_name, subject_name=subject_name,
    )


@app.route('/api/chat', methods=['POST'])
@login_required
def api_chat():
    import urllib.request
    import json as json_mod
    api_key = os.environ.get('ANTHROPIC_API_KEY', '')
    if not api_key:
        return jsonify({'reply': '서버에 API 키가 설정되지 않았습니다.'}), 500
    data = request.get_json()
    user_msg = data.get('message', '')
    # 이용 리포트용 컨텍스트(클라이언트 제공). slide_id로 과목(콘텐츠 축)을 서버에서 재조회한다.
    chat_slide_id = (data.get('slide_id') or '').strip() or None
    chat_tab = (data.get('tab') or '').strip() or None
    if chat_tab not in ('qa', 'quiz'):
        chat_tab = None
    # 탈옥 방어: 클라이언트 system 파라미터 무시, 서버 측 가드레일만 사용 (§12-4 ③)
    system_prompt = (
        '당신은 SlideAtlas의 병리학·조직학 AI 튜터입니다. '
        'SlideAtlas 학습(병리학, 조직학, 슬라이드 판독, 관련 의학 지식)과 무관한 질문에는 '
        '"SlideAtlas 학습 관련 질문만 답변합니다"라고만 응답하세요. '
        '한국어로 답변하세요.'
    )
    if not user_msg:
        return jsonify({'reply': '메시지가 없습니다.'}), 400

    # AI 튜터 호출 1회 로그(best-effort). 과목 축은 슬라이드 콘텐츠 과목으로 기록.
    if chat_slide_id:
        _sinfo = get_slide_institution(chat_slide_id)
        _chat_subject = _sinfo[1] if _sinfo else None
    else:
        _chat_subject = None
    _log_chat(
        getattr(g, 'user_id', None), getattr(g, 'institution_id', None),
        chat_slide_id, chat_tab, _chat_subject,
    )

    if 'JSON' in system_prompt or '형식으로만 응답' in system_prompt:
        payload = json_mod.dumps({
            'model': 'claude-sonnet-4-5',
            'max_tokens': 800,
            'system': system_prompt,
            'messages': [{'role': 'user', 'content': user_msg}]
        }).encode('utf-8')
        req = urllib.request.Request(
            'https://api.anthropic.com/v1/messages',
            data=payload,
            headers={'Content-Type': 'application/json', 'x-api-key': api_key, 'anthropic-version': '2023-06-01'},
            method='POST'
        )
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                result = json_mod.loads(resp.read().decode('utf-8'))
                reply = result['content'][0]['text'] if result.get('content') else '응답 없음'
                return jsonify({'reply': reply})
        except Exception as e:
            return jsonify({'reply': f'API 오류: {str(e)}'}), 500

    payload = json_mod.dumps({
        'model': 'claude-sonnet-4-5',
        'max_tokens': 600,
        'stream': True,
        'system': system_prompt,
        'messages': [{'role': 'user', 'content': user_msg}]
    }).encode('utf-8')
    req = urllib.request.Request(
        'https://api.anthropic.com/v1/messages',
        data=payload,
        headers={
            'Content-Type': 'application/json',
            'x-api-key': api_key,
            'anthropic-version': '2023-06-01'
        },
        method='POST'
    )

    def generate():
        try:
            with urllib.request.urlopen(req, timeout=60) as resp:
                for line in resp:
                    line = line.decode('utf-8').strip()
                    if line.startswith('data: '):
                        chunk = line[6:]
                        if chunk == '[DONE]':
                            break
                        try:
                            obj = json_mod.loads(chunk)
                            if obj.get('type') == 'content_block_delta':
                                text = obj.get('delta', {}).get('text', '')
                                if text:
                                    yield f"data: {json_mod.dumps({'text': text})}\n\n"
                        except Exception:
                            pass
        except Exception as e:
            yield f"data: {json_mod.dumps({'error': str(e)})}\n\n"

    return Response(
        generate(),
        mimetype='text/event-stream',
        headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'}
    )

@app.route('/dzi/<slide_id>.dzi')
@tile_token_required
def dzi_descriptor(slide_id):
    # [Gemini#1] 고빈도 타일 경로 — HMAC tile_token만 검증(DB 조회 없음).
    err = _verify_tile_request(slide_id)
    if err:
        return err
    if slide_id not in SLIDE_CACHE:
        return Response("Loading...", status=503)
    cache = SLIDE_CACHE[slide_id]
    if cache.get('ec2'):
        return Response("Use EC2 tileserver", status=404)
    dz = cache["dz"]
    w, h = dz.level_dimensions[-1]
    xml = f'''<?xml version="1.0" encoding="UTF-8"?>
<Image xmlns="http://schemas.microsoft.com/deepzoom/2008"
  Format="jpeg" Overlap="{OVERLAP}" TileSize="{TILE_SIZE}">
  <Size Width="{w}" Height="{h}"/>
</Image>'''
    resp = Response(xml, mimetype='application/xml')
    resp.headers['Cache-Control'] = 'no-store, no-cache'
    return resp

@app.route('/dzi/<slide_id>_files/<int:level>/<int:col>_<int:row>.jpeg')
@app.route('/dzi/<slide_id>_files/<int:level>/<int:col>_<int:row>.jpg')
@tile_token_required
def dzi_tile(slide_id, level, col, row):
    # [Gemini#1] 최고빈도 타일 경로 — HMAC tile_token만 검증(DB 조회 없음).
    err = _verify_tile_request(slide_id)
    if err:
        return err
    if slide_id not in SLIDE_CACHE:
        img = Image.new('RGB', (TILE_SIZE, TILE_SIZE), (245, 240, 235))
        buf = io.BytesIO()
        img.save(buf, 'JPEG')
        buf.seek(0)
        r = send_file(buf, mimetype='image/jpeg')
        r.headers['Cache-Control'] = 'no-store, no-cache'
        return r
    try:
        dz = SLIDE_CACHE[slide_id]["dz"]
        tile = dz.get_tile(level, (col, row))
        buf = io.BytesIO()
        tile.save(buf, format='JPEG', quality=88)
        buf.seek(0)
        r = send_file(buf, mimetype='image/jpeg')
        r.headers['Cache-Control'] = 'no-store, no-cache'
        return r
    except Exception as e:
        img = Image.new('RGB', (TILE_SIZE, TILE_SIZE), (255, 255, 255))
        buf = io.BytesIO()
        img.save(buf, 'JPEG')
        buf.seek(0)
        r = send_file(buf, mimetype='image/jpeg')
        r.headers['Cache-Control'] = 'no-store, no-cache'
        return r

@app.route('/thumbnail/<slide_id>')
@tile_token_required
def thumbnail(slide_id):
    # [Gemini#1] 고빈도 타일 경로 — HMAC tile_token만 검증(DB 조회 없음).
    err = _verify_tile_request(slide_id)
    if err:
        return err
    if slide_id not in SLIDE_CACHE:
        return Response("Not loaded", status=503)
    cache = SLIDE_CACHE[slide_id]
    if cache.get('ec2'):
        return Response("Use EC2 thumbnail", status=404)
    try:
        slide_obj = cache["slide"]
        thumb = slide_obj.get_thumbnail((280, 200))
        buf = io.BytesIO()
        thumb.save(buf, format='JPEG', quality=80)
        buf.seek(0)
        r = send_file(buf, mimetype='image/jpeg')
        r.headers['Cache-Control'] = 'no-store, no-cache'
        return r
    except Exception as e:
        img = Image.new('RGB', (280, 200), (245, 240, 235))
        buf = io.BytesIO()
        img.save(buf, 'JPEG')
        buf.seek(0)
        return send_file(buf, mimetype='image/jpeg')

@app.route('/api/tile-token')
@login_required
def api_tile_token():
    """[2-2#2] 타일 토큰(TTL 5분) 재발급. 뷰어가 만료 전/후 호출해 끊김 없이 갱신.

    발급 조건은 _slide_access_allowed 게이트 통과(묶음 A) — 접근권 없는 슬라이드는 재발급 거부.
    @login_required로 구독·세션·특별계정 만료도 매 요청 검사된다.
    """
    slide_id = (request.args.get('slide') or '').strip()
    if not slide_id:
        resp = jsonify({'success': False, 'error': 'MISSING_SLIDE'})
        resp.status_code = 400
        resp.headers['Cache-Control'] = 'no-store, no-cache'
        return resp
    allowed, aerr = _slide_access_allowed(slide_id)
    if not allowed:
        return aerr  # 접근권 없는 슬라이드 토큰 재발급 금지 (403/404)
    from flask import g as _g
    token = generate_tile_token(
        str(getattr(_g, 'user_id', '')), getattr(_g, 'institution_id', ''), slide_id)
    resp = jsonify({'success': True, 'token': token})
    resp.headers['Cache-Control'] = 'no-store, no-cache'
    return resp

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    # 이미 로그인된 어드민은 대시보드로
    if session.get('admin_user_id') and _get_admin_user():
        return redirect('/admin')

    error = ''
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        if not email or not password:
            error = '이메일과 비밀번호를 입력하세요.'
        else:
            conn = get_db_conn()
            conn.autocommit = False
            try:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT id, password_hash, role, name, status, locked_at FROM admin_users"
                        " WHERE lower(email) = %s",
                        (email,),
                    )
                    row = cur.fetchone()
                    if row is None:
                        error = '이메일 또는 비밀번호가 올바르지 않습니다.'
                        conn.rollback()
                    else:
                        admin_id, pw_hash, role, name, status, locked_at = row
                        if status != 'active':
                            error = '비활성화된 계정입니다. 관리자에게 문의하세요.'
                            conn.rollback()
                        # [Gemini#1] 무차별 대입 차단: 24h 윈도우 10회 실패 → 잠금, 24h 후 자동 해제.
                        elif _admin_is_locked(cur, admin_id, locked_at):
                            conn.commit()  # 자동해제 미발생 — 트랜잭션 마무리만
                            error = '보안상 계정이 잠겼습니다. 24시간 후 다시 시도하세요.'
                        elif not check_password_hash(pw_hash, password):
                            now_locked = _admin_check_and_increment_failed(cur, admin_id)
                            conn.commit()
                            error = ('보안상 계정이 잠겼습니다. 24시간 후 다시 시도하세요.'
                                     if now_locked else '이메일 또는 비밀번호가 올바르지 않습니다.')
                        else:
                            import secrets as _sec_admin
                            import uuid as _uuid_admin
                            # [Codex#2] 세션 토큰 회전 — 로그인 성공 시 새 토큰을 DB·세션에 저장(매 요청 대조).
                            admin_session_token = str(_uuid_admin.uuid4())
                            cur.execute(
                                "UPDATE admin_users SET last_login = NOW(), session_token = %s, "
                                "failed_attempts = 0, failed_window_start = NULL WHERE id = %s",
                                (admin_session_token, admin_id),
                            )
                            conn.commit()
                            session.clear()
                            session['admin_user_id'] = admin_id
                            session['admin_role'] = role
                            session['admin_name'] = name or email.split('@')[0]
                            session['admin_csrf_token'] = _sec_admin.token_hex(32)
                            session['admin_session_token'] = admin_session_token
                            return redirect('/admin')
            except Exception:
                conn.rollback()
                error = '처리 중 오류가 발생했습니다.'
            finally:
                try:
                    conn.autocommit = True
                except Exception:
                    pass
                release_db_conn(conn)

    return render_template('admin/login.html', error=error)


@app.route('/admin/logout')
def admin_logout():
    # [Codex#2] DB의 session_token도 무효화 → 동일 토큰을 가진 다른 쿠키도 즉시 차단.
    admin_user_id = session.get('admin_user_id')
    if admin_user_id:
        try:
            conn = get_db_conn()
            conn.autocommit = False
            try:
                with conn.cursor() as cur:
                    cur.execute(
                        "UPDATE admin_users SET session_token = NULL WHERE id = %s",
                        (admin_user_id,),
                    )
                conn.commit()
            finally:
                try:
                    conn.autocommit = True
                except Exception:
                    pass
                release_db_conn(conn)
        except Exception:
            pass
    session.pop('admin_user_id', None)
    session.pop('admin_role', None)
    session.pop('admin_name', None)
    session.pop('admin_csrf_token', None)
    session.pop('admin_session_token', None)
    return redirect('/admin/login')


@app.route('/admin')
@admin_required
def admin_dashboard():
    return render_template('admin/dashboard.html', active_page='dash')


def _term_label(term: str) -> str:
    """'2026-fall' → {'year':'2026', 'season':'가을'}"""
    parts = term.split('-')
    if len(parts) == 2:
        season = {'spring': '봄', 'fall': '가을'}.get(parts[1], parts[1])
        return f'{parts[0]} {season}'
    return term


@app.route('/admin/api/dashboard', methods=['GET'])
@admin_required
def api_dashboard():
    from datetime import date, timedelta
    today = date.today()
    d90   = today + timedelta(days=90)

    conn = get_db_conn()
    try:
        with conn.cursor() as cur:

            # ── KPI 1: 활성 구독 기관 수 ──
            cur.execute("""
                SELECT COUNT(DISTINCT institution_id) FROM subscriptions
                WHERE status = 'active' AND subscription_end >= %s
            """, (today,))
            active_inst = cur.fetchone()[0] or 0

            # ── KPI 2: 이번 학기 확정 매출 (현재 접근 가능한 구독만) ──
            cur.execute("""
                SELECT COALESCE(SUM(fee), 0) FROM subscriptions
                WHERE status = 'active'
                  AND access_open_date <= %s
                  AND subscription_end >= %s
                  AND fee IS NOT NULL
            """, (today, today))
            current_revenue = int(cur.fetchone()[0] or 0)

            # ── KPI 3: 활성 사용자 / 좌석 ──
            cur.execute("""
                SELECT COUNT(*) FROM users
                WHERE COALESCE(status, 'active') = 'active'
                  AND COALESCE(is_special, FALSE) = FALSE
            """)
            active_users = cur.fetchone()[0] or 0

            cur.execute("""
                SELECT COALESCE(SUM(max_seats), 0) FROM subscriptions
                WHERE status = 'active' AND subscription_end >= %s
            """, (today,))
            total_seats = int(cur.fetchone()[0] or 0)

            # ── KPI 4: 만료 임박 D-90 ──
            cur.execute("""
                SELECT COUNT(DISTINCT institution_id) FROM subscriptions
                WHERE status = 'active'
                  AND subscription_end >= %s
                  AND subscription_end <= %s
            """, (today, d90))
            expiring_d90 = cur.fetchone()[0] or 0

            # ── 만료 임박 리스트 ──
            cur.execute("""
                SELECT s.institution_id, i.name_ko, s.subject_code,
                       s.subscription_end,
                       (s.subscription_end - %s) AS days_left
                FROM subscriptions s
                JOIN institutions i ON s.institution_id = i.id
                WHERE s.status = 'active' AND s.subscription_end >= %s
                ORDER BY s.subscription_end ASC
                LIMIT 10
            """, (today, today))
            expiry_rows = cur.fetchall()
            expiry_list = []
            for r in expiry_rows:
                dl = r[4]
                if dl <= 90:
                    dday_color = 'danger'
                elif dl <= 180:
                    dday_color = 'warn'
                else:
                    dday_color = 'ok'
                expiry_list.append({
                    'inst_id':   r[0],
                    'inst_name': r[1],
                    'subject':   r[2],
                    'end_date':  r[3].isoformat(),
                    'days_left': dl,
                    'color':     dday_color,
                })

            # ── 학기별 매출 추이 ──
            cur.execute("""
                SELECT start_term,
                       COALESCE(SUM(fee), 0) AS total_fee,
                       MIN(access_open_date) AS term_opens
                FROM subscriptions
                WHERE fee IS NOT NULL
                GROUP BY start_term
                ORDER BY start_term ASC
                LIMIT 8
            """)
            rev_rows = cur.fetchall()
            rev_max = max((r[1] for r in rev_rows), default=1) or 1
            revenue_trend = [{
                'term':   r[0],
                'label':  _term_label(r[0]),
                'fee':    int(r[1]),
                'pct':    round(r[1] / rev_max * 100),
                'future': r[2] > today if r[2] else False,
            } for r in rev_rows]

            # ── 파이프라인 현황 ──
            cur.execute("""
                SELECT
                  COUNT(*) FILTER (WHERE deploy_status = 'deployed')                                          AS deployed,
                  COUNT(*) FILTER (WHERE deploy_status = 'qc_pending'
                                     AND conversion_status IN ('ready','ready_no_mpp'))                       AS qc_pending,
                  COUNT(*) FILTER (WHERE conversion_status IN ('pending','converting','qc_check'))            AS processing,
                  COUNT(*) FILTER (WHERE conversion_status = 'failed' OR deploy_status = 'rejected')         AS failed_rejected
                FROM slides
            """)
            p = cur.fetchone()
            pipeline = {
                'deployed':        p[0] or 0,
                'qc_pending':      p[1] or 0,
                'processing':      p[2] or 0,
                'failed_rejected': p[3] or 0,
            }

            # ── 처리 대기 ──
            open_inquiries = 0
            try:
                cur.execute("SELECT COUNT(*) FROM inquiries WHERE status = 'open'")
                open_inquiries = cur.fetchone()[0] or 0
            except Exception:
                conn.rollback()

            cur.execute("""
                SELECT COUNT(*) FROM slides
                WHERE deploy_status = 'qc_pending'
                  AND conversion_status IN ('ready','ready_no_mpp')
            """)
            qc_pending_slides = cur.fetchone()[0] or 0

            cur.execute("""
                SELECT COUNT(*) FROM slides WHERE conversion_status = 'ready_no_mpp'
            """)
            mpp_missing = cur.fetchone()[0] or 0

        util_pct = round(active_users / total_seats * 100) if total_seats else 0

        return jsonify({
            'ok': True,
            'kpi': {
                'active_inst':     active_inst,
                'current_revenue': current_revenue,
                'active_users':    active_users,
                'total_seats':     total_seats,
                'util_pct':        util_pct,
                'expiring_d90':    expiring_d90,
            },
            'expiry_list':    expiry_list,
            'revenue_trend':  revenue_trend,
            'pipeline':       pipeline,
            'todo': {
                'open_inquiries':   open_inquiries,
                'qc_pending_slides': qc_pending_slides,
                'mpp_missing':      mpp_missing,
                'expiring_renewals': expiring_d90,
            },
        })
    finally:
        release_db_conn(conn)

# ── 슬라이드 관리 (S5-3) ─────────────────────────────────────

@app.route('/admin/slides')
@super_admin_required
def admin_slides_page():
    return render_template('admin/slides.html', active_page='slide')


@app.route('/admin/api/slides/list', methods=['GET'])
@super_admin_required
def api_slides_list():
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, institution_id, subject_code,
                       title_ko, title_en, organ, stain, species,
                       original_format, mpp,
                       conversion_status, deploy_status, reject_reason,
                       knowledge_base,
                       license_source, created_at
                FROM slides
                ORDER BY created_at DESC
            """)
            cols = [d[0] for d in cur.description]
            slides = []
            for row in cur.fetchall():
                s = dict(zip(cols, row))
                s['created_at'] = str(s['created_at'])
                if s['knowledge_base']:
                    try:
                        s['knowledge_base'] = s['knowledge_base'] if isinstance(s['knowledge_base'], dict) else json.loads(s['knowledge_base'])
                    except Exception:
                        s['knowledge_base'] = {}
                slides.append(s)

        # 상태 칩 카운트
        counts = {
            'total': len(slides),
            'processing': sum(1 for s in slides if s['conversion_status'] in ('pending', 'converting', 'qc_check')),
            'qc_pending': sum(1 for s in slides if s['conversion_status'] == 'ready' and s['deploy_status'] == 'qc_pending'),
            'no_mpp': sum(1 for s in slides if s['conversion_status'] == 'ready_no_mpp'),
            'failed': sum(1 for s in slides if s['conversion_status'] == 'failed' or s['deploy_status'] == 'rejected'),
            'deployed': sum(1 for s in slides if s['deploy_status'] == 'deployed'),
        }
        return jsonify({'ok': True, 'slides': slides, 'counts': counts})
    finally:
        release_db_conn(conn)


@app.route('/admin/api/organs', methods=['GET'])
@super_admin_required
def api_organs_list():
    """organ 통제어휘 목록(§18 D28). 개별 추가 모달 드롭다운용 — is_active=TRUE, display_order 정렬."""
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT organ_code, name_ko, name_en, organ_system
                FROM organs
                WHERE is_active = TRUE
                ORDER BY display_order, name_ko
            """)
            cols = [d[0] for d in cur.description]
            organs = [dict(zip(cols, row)) for row in cur.fetchall()]
        return jsonify({'ok': True, 'organs': organs})
    finally:
        release_db_conn(conn)


@app.route('/admin/api/slides/<slide_id>/deploy', methods=['POST'])
@super_admin_required
@admin_csrf_required
def api_slide_deploy(slide_id):
    conn = get_db_conn()
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE slides SET deploy_status='deployed', reject_reason=NULL
                    WHERE id=%s AND conversion_status='ready' AND deploy_status='qc_pending'
                    RETURNING id
                """, (slide_id,))
                if not cur.fetchone():
                    return jsonify({'ok': False, 'error': '배포 가능 상태가 아닙니다'}), 409
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        release_db_conn(conn)


@app.route('/admin/api/slides/<slide_id>/revoke', methods=['POST'])
@super_admin_required
@admin_csrf_required
def api_slide_revoke(slide_id):
    conn = get_db_conn()
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE slides SET deploy_status='qc_pending'
                    WHERE id=%s AND deploy_status='deployed'
                    RETURNING id
                """, (slide_id,))
                if not cur.fetchone():
                    return jsonify({'ok': False, 'error': '배포 중 상태가 아닙니다'}), 409
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        release_db_conn(conn)


@app.route('/admin/api/slides/<slide_id>/reject', methods=['POST'])
@super_admin_required
@admin_csrf_required
def api_slide_reject(slide_id):
    data = request.get_json(silent=True) or {}
    reason = (data.get('reason') or '').strip()
    detail = (data.get('detail') or '').strip()
    reject_reason = f'{reason}: {detail}' if detail else reason
    if not reject_reason:
        return jsonify({'ok': False, 'error': '반려 사유는 필수입니다'}), 400
    conn = get_db_conn()
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE slides SET deploy_status='rejected', reject_reason=%s
                    WHERE id=%s AND deploy_status IN ('qc_pending', 'deployed')
                    RETURNING id
                """, (reject_reason, slide_id))
                if not cur.fetchone():
                    return jsonify({'ok': False, 'error': '반려 가능 상태가 아닙니다'}), 409
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        release_db_conn(conn)


@app.route('/admin/api/slides/<slide_id>/mpp', methods=['POST'])
@super_admin_required
@admin_csrf_required
def api_slide_set_mpp(slide_id):
    data = request.get_json(silent=True) or {}
    try:
        mpp_val = float(data.get('mpp', 0))
    except (TypeError, ValueError):
        return jsonify({'ok': False, 'error': '유효한 MPP 값이 아닙니다'}), 400
    if mpp_val <= 0 or mpp_val > 10:
        return jsonify({'ok': False, 'error': 'MPP 범위를 확인하세요 (0 초과 10 이하)'}), 400
    conn = get_db_conn()
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE slides SET mpp=%s, conversion_status='pending'
                    WHERE id=%s AND conversion_status='ready_no_mpp'
                    RETURNING id
                """, (mpp_val, slide_id))
                if not cur.fetchone():
                    return jsonify({'ok': False, 'error': 'MPP 없음 상태가 아닙니다'}), 409
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        release_db_conn(conn)


@app.route('/admin/api/slides/<slide_id>/kb', methods=['PUT'])
@super_admin_required
@admin_csrf_required
def api_slide_kb(slide_id):
    data = request.get_json(silent=True) or {}
    kb = data.get('knowledge_base')
    deploy_after = data.get('deploy', False)
    if kb is None:
        return jsonify({'ok': False, 'error': 'knowledge_base 누락'}), 400
    conn = get_db_conn()
    try:
        with conn:
            with conn.cursor() as cur:
                if deploy_after:
                    cur.execute("""
                        UPDATE slides SET knowledge_base=%s, deploy_status='deployed', reject_reason=NULL
                        WHERE id=%s AND conversion_status='ready' AND deploy_status='qc_pending'
                        RETURNING id
                    """, (json.dumps(kb, ensure_ascii=False), slide_id))
                else:
                    cur.execute("""
                        UPDATE slides SET knowledge_base=%s WHERE id=%s RETURNING id
                    """, (json.dumps(kb, ensure_ascii=False), slide_id))
                if not cur.fetchone():
                    return jsonify({'ok': False, 'error': '슬라이드를 찾을 수 없거나 배포 불가 상태입니다'}), 409
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        release_db_conn(conn)


@app.route('/admin/api/slides/batch', methods=['POST'])
@super_admin_required
@admin_csrf_required
def api_slides_batch():
    data = request.get_json(silent=True) or {}
    action = data.get('action')  # 'deploy' | 'reject'
    ids = data.get('ids') or []
    if not ids or action not in ('deploy', 'reject'):
        return jsonify({'ok': False, 'error': '잘못된 요청'}), 400
    if len(ids) > 200:
        return jsonify({'ok': False, 'error': '한 번에 최대 200건'}), 400

    conn = get_db_conn()
    try:
        with conn:
            with conn.cursor() as cur:
                if action == 'deploy':
                    cur.execute("""
                        UPDATE slides SET deploy_status='deployed', reject_reason=NULL
                        WHERE id = ANY(%s) AND conversion_status='ready' AND deploy_status='qc_pending'
                    """, (ids,))
                else:
                    reason = (data.get('reason') or '일괄 반려').strip()
                    cur.execute("""
                        UPDATE slides SET deploy_status='rejected', reject_reason=%s
                        WHERE id = ANY(%s) AND deploy_status IN ('qc_pending', 'deployed')
                    """, (reason, ids))
                updated = cur.rowcount
        return jsonify({'ok': True, 'updated': updated})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        release_db_conn(conn)


@app.route('/admin/api/slides/add', methods=['POST'])
@super_admin_required
@admin_csrf_required
def api_slide_add():
    """개별 슬라이드 추가 (메타데이터). 파일 업로드는 파이프라인 별도 처리."""
    data = request.get_json(silent=True) or {}
    subject_code = (data.get('subject_code') or '').upper()
    title_ko = (data.get('title_ko') or '').strip()
    if not subject_code or not title_ko:
        return jsonify({'ok': False, 'error': '과목·제목(한국어)은 필수입니다'}), 400

    # organ 통제어휘(§18 D28): organ_code 필수(Med#1). 누락/빈 값이면 신규 INSERT 거부.
    #   ⚠ 기존 NULL organ_code 행(D24 잔재)은 본 변경과 무관 — 신규 추가 경로에만 강제한다.
    organ_code = (data.get('organ_code') or '').strip().lower()
    if not organ_code:
        return jsonify({'ok': False, 'error': 'organ_code(장기)는 필수입니다'}), 400

    conn = get_db_conn()
    try:
        with conn:
            with conn.cursor() as cur:
                # organ_code 는 organs 마스터에 활성 행으로 존재해야 한다(임의값 거부).
                # 표시 연속성을 위해 organ 컬럼엔 선택 organ_code 의 name_ko 를 함께 기록
                # (home 계통 필터 s['system']·admin 목록 escH(s.organ) 등 기존 표시 경로 무변경).
                cur.execute(
                    "SELECT name_ko FROM organs WHERE organ_code=%s AND is_active=TRUE",
                    (organ_code,))
                row = cur.fetchone()
                if not row:
                    return jsonify({'ok': False, 'error': f'유효하지 않은 organ_code: {organ_code}'}), 400
                organ_name = row[0]

                # SA-{SUBJECT} 접두사로 다음 채번
                cur.execute("""
                    SELECT id FROM slides WHERE id LIKE %s ORDER BY id DESC LIMIT 1
                """, (f'SA-{subject_code}-%',))
                last = cur.fetchone()
                if last:
                    try:
                        last_num = int(last[0].split('-')[-1])
                    except ValueError:
                        last_num = 0
                    next_num = last_num + 1
                else:
                    next_num = 1
                new_id = f'SA-{subject_code}-{next_num:03d}'

                cur.execute("""
                    INSERT INTO slides (
                        id, institution_id, subject_code,
                        title_ko, title_en, description,
                        stain, organ, organ_code, species, license_source,
                        original_format, conversion_status, deploy_status
                    ) VALUES (%s,'SA',%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'pending','qc_pending')
                """, (
                    new_id, subject_code,
                    title_ko,
                    (data.get('title_en') or '').strip() or None,
                    (data.get('description') or '').strip() or None,
                    (data.get('stain') or '').strip() or None,
                    organ_name,         # organ(표시용) = 선택 organ_code 의 name_ko
                    organ_code,         # organ_code(정규 앵커) = 통제어휘 PK
                    (data.get('species') or 'human').strip(),
                    (data.get('license_source') or '').strip() or None,
                    (data.get('original_format') or '').upper() or None,
                ))
        return jsonify({'ok': True, 'id': new_id})
    except psycopg2.IntegrityError:
        return jsonify({'ok': False, 'error': f'ID {new_id} 충돌. 재시도하세요'}), 409
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        release_db_conn(conn)


@app.route('/admin/api/slide', methods=['POST'])
@super_admin_required
@admin_csrf_required
def admin_save_slide():
    # ⚠ DEPRECATED (organ 정규화 §18 D28): 이 레거시 경로는 프론트엔드에서 호출되지 않는다
    #   (개별 추가는 /admin/api/slides/add → api_slide_add 가 organ_code 통제어휘로 기록).
    #   본 핸들러는 payload['system']→organ 자유텍스트 단독 기록이라 organ_code 정규화를 우회하므로
    #   신규 사용 금지. 라우트는 어드민 인증/CSRF/세션잠금 테스트(tests/test_auth.py)가 참조하므로
    #   존치하되, 슬라이드 메타 저장 용도로는 재사용하지 말 것. v1.5 정리 대상.
    #
    # [Med#2 레거시 하드블록 — 통제어휘 우회 차단] 인증·CSRF·세션잠금 게이트(데코레이터)는 그대로
    #   통과시켜 어드민 게이트 테스트(tests/test_auth.py 401/403)에 영향이 없게 하되, organ 자유텍스트
    #   저장(organ_code 정규화 우회)에 도달하기 전에 410 Gone 으로 막는다. 슬라이드 메타 추가는
    #   /admin/api/slides/add(api_slide_add, organ_code 통제어휘 필수)만 사용한다(§18 D28).
    return jsonify({
        'ok': False,
        'error': '이 경로는 더 이상 사용되지 않습니다. 슬라이드 추가는 통제어휘(organ_code) 경로를 사용하세요.',
    }), 410

@app.route('/admin/api/slide/<slide_id>', methods=['DELETE'])
@super_admin_required
@admin_csrf_required
def admin_delete_slide(slide_id):
    # [ROLLBACK] JSON 방식 -- RDS 전환 전 코드
    # try:
    #     data = load_slides()
    #     data['slides'] = [s for s in data.get('slides', []) if s['id'] != slide_id]
    #     save_slides(data)
    #     return jsonify({'ok': True})
    # except Exception as e:
    #     return jsonify({'ok': False, 'error': str(e)})
    try:
        conn = get_db_conn()
        try:
            with conn:
                with conn.cursor() as cur:
                    cur.execute("DELETE FROM slides WHERE id = %s", (slide_id,))
        finally:
            release_db_conn(conn)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

# ── 기관 관리 (S5-2) ──────────────────────────────────────

import calendar as _calendar
from datetime import date as _date


def _sem_dates(term: str):
    """'2026-fall' → (access_open_date, subscription_end) as date objects.
    윤년은 calendar.monthrange로 자동 처리."""
    year_s, season = term.split('-')
    year = int(year_s)
    if season == 'spring':
        return _date(year, 2, 1), _date(year, 8, 31)
    else:  # fall
        last_feb = _calendar.monthrange(year + 1, 2)[1]
        return _date(year, 8, 1), _date(year + 1, 2, last_feb)


def _sub_status(open_date, end_date):
    # [item4] 구독 상태(대시보드 표시)도 게이트와 동일 KST 기준 — 자정~오전9시 하루 어긋남 제거(§18 D10).
    #   P2 포털 구독카드·슈퍼관리자 기관목록이 공유하는 단일 상태 함수.
    today = _today_kst()
    if end_date < today:
        return 'expired', '만료'
    if open_date <= today:
        return 'active', '구독중'
    delta = (open_date - today).days
    if delta <= 60:
        return 'upcoming', '오픈예정'
    return 'pending', '대기'


def _next_term(term: str) -> str:
    year_s, season = term.split('-')
    year = int(year_s)
    return f'{year}-fall' if season == 'spring' else f'{year + 1}-spring'


def _term_end_after(start_term: str, term_count: int) -> str:
    """start_term 에서 term_count 학기 후의 마지막 학기 식별자."""
    year_s, season = start_term.split('-')
    year = int(year_s)
    idx = year * 2 + (0 if season == 'spring' else 1)
    idx += term_count - 1
    return f'{idx // 2}-{"spring" if idx % 2 == 0 else "fall"}'


PLAN_SEATS = {'department': 50, 'standard': 150, 'campus': 300, 'institution': 500}

# 구독 입력 검증 상수/한계 (Codex#5). 잘못된 타입·범위는 500이 아니라 400으로 거른다.
VALID_PLANS = set(PLAN_SEATS) | {'custom'}
MAX_TERM_COUNT = 20        # 학기 수 상한 (10년)
MAX_SEATS_LIMIT = 100000   # 좌석 수 상한
MAX_FEE_LIMIT = 10 ** 12   # 구독료 상한(원)
import re as _re_validate
START_TERM_RE = _re_validate.compile(r'^\d{4}-(spring|fall)$')   # 학기 식별자 형식(Codex#3)


def _parse_int_field(raw, name, *, default=None, minimum=0, maximum=None, allow_none=False):
    """int 변환 + 범위 검증. 성공 시 (value, None), 실패 시 (None, error_message).

    빈 값/None은 default로 대체(allow_none이면 None 허용). 콤마 포함 금액 문자열도 허용.
    """
    if raw is None or (isinstance(raw, str) and raw.strip() == ''):
        if allow_none:
            return default, None
        raw = default
    if raw is None and allow_none:
        return None, None
    try:
        if isinstance(raw, bool):           # bool은 int 서브타입 — 명시적 거부
            raise ValueError()
        value = int(str(raw).replace(',', '').strip())
    except (ValueError, TypeError):
        return None, f'{name} 값이 올바르지 않습니다(정수 필요)'
    if value < minimum:
        return None, f'{name} 값은 {minimum} 이상이어야 합니다'
    if maximum is not None and value > maximum:
        return None, f'{name} 값이 허용 범위를 초과했습니다'
    return value, None


def _validate_array_of_dicts(value, name):
    """value가 list이고 모든 원소가 dict인지 검증. 정상이면 None, 아니면 에러 메시지(Codex#4)."""
    if not isinstance(value, list):
        return f'{name}은(는) 배열이어야 합니다'
    if not all(isinstance(x, dict) for x in value):
        return f'{name}의 각 항목은 객체여야 합니다'
    return None


def _subject_codes_set():
    """subject_codes 테이블의 코드 집합 (allowlist, 하드코딩 금지 §6-1)."""
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT code FROM subject_codes")
            return {r[0] for r in cur.fetchall()}
    finally:
        release_db_conn(conn)


def _validate_subscription_payload(raw, valid_subjects):
    """구독 1건 입력 검증·정규화. 성공 시 (fields_dict, None), 실패 시 (None, error_message).

    fields: subject_code, plan, max_seats, term_count, fee, start_term, payment_method.
    plan·subject_code allowlist, start_term 형식, term_count·max_seats·fee 타입·범위 검증(Codex#3·#5).
    """
    if not isinstance(raw, dict):
        return None, '구독 항목 형식이 올바르지 않습니다'   # [Codex#4] 원소 타입 검증
    start_term = (raw.get('start_term') or '').strip()
    subject_code = (raw.get('subject_code') or '').strip().upper()
    plan = (raw.get('plan') or '').strip().lower()
    if not START_TERM_RE.match(start_term):
        # [Codex#3] 'YYYY-spring' | 'YYYY-fall'만 허용 — _sem_dates 계산이 깨지지 않게 사전 차단.
        return None, "start_term 형식이 올바르지 않습니다('YYYY-spring' 또는 'YYYY-fall')"
    if subject_code not in valid_subjects:
        return None, f'알 수 없는 과목코드입니다: {subject_code or "(빈값)"}'
    if plan not in VALID_PLANS:
        return None, f'알 수 없는 플랜입니다: {plan or "(빈값)"}'

    term_count, err = _parse_int_field(raw.get('term_count'), 'term_count',
                                       default=1, minimum=1, maximum=MAX_TERM_COUNT)
    if err:
        return None, err
    # max_seats 미지정 시 플랜 기본값. custom 플랜은 명시 좌석 필요.
    seats_raw = raw.get('max_seats')
    if seats_raw is None or (isinstance(seats_raw, str) and seats_raw.strip() == ''):
        seats_default = PLAN_SEATS.get(plan)
        if seats_default is None:
            return None, 'custom 플랜은 max_seats를 명시해야 합니다'
        max_seats = seats_default
    else:
        max_seats, err = _parse_int_field(seats_raw, 'max_seats', minimum=1, maximum=MAX_SEATS_LIMIT)
        if err:
            return None, err
    fee, err = _parse_int_field(raw.get('fee'), 'fee', default=None,
                                minimum=0, maximum=MAX_FEE_LIMIT, allow_none=True)
    if err:
        return None, err
    payment_method = (raw.get('payment_method') or '학기 선불')
    return {
        'subject_code': subject_code, 'plan': plan, 'max_seats': max_seats,
        'term_count': term_count, 'fee': fee, 'start_term': start_term,
        'payment_method': payment_method,
    }, None


@app.route('/admin/institutions')
@super_admin_required
def admin_institutions_page():
    return render_template('admin/institutions.html', active_page='inst')


@app.route('/admin/api/institutions', methods=['GET'])
@super_admin_required
def api_institutions_list():
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT i.id, i.name_ko, i.name_en, i.university, i.college, i.domain,
                       i.admin_contacts
                FROM institutions i
                ORDER BY i.name_ko
            """)
            cols = [d[0] for d in cur.description]
            insts = [dict(zip(cols, r)) for r in cur.fetchall()]

            for inst in insts:
                # used_seats: institution_rosters가 아직 없을 수 있으므로 0으로 폴백
                try:
                    cur.execute("""
                        SELECT subject_code, COUNT(*) AS used_seats
                        FROM institution_rosters
                        WHERE institution_id = %s
                        GROUP BY subject_code
                    """, (inst['id'],))
                    used_map = {r[0]: r[1] for r in cur.fetchall()}
                except Exception:
                    conn.rollback()
                    used_map = {}

                cur.execute("""
                    SELECT id, subject_code, plan, max_seats,
                           start_term, term_count,
                           access_open_date, subscription_end,
                           fee, payment_method, status
                    FROM subscriptions
                    WHERE institution_id = %s
                    ORDER BY subscription_end DESC
                """, (inst['id'],))
                scols = [d[0] for d in cur.description]
                subs = []
                for row in cur.fetchall():
                    sub = dict(zip(scols, row))
                    open_d = sub['access_open_date']
                    end_d = sub['subscription_end']
                    today = _today_kst()    # [item4] _sub_status(KST)와 D-day 기준일 일치(어긋남 방지)
                    status_key, status_label = _sub_status(open_d, end_d)
                    if status_key == 'active':
                        dday = f'D-{(end_d - today).days} 만료'
                    elif status_key in ('upcoming', 'pending'):
                        dday = f'D-{(open_d - today).days} 오픈'
                    else:
                        dday = f'D+{(today - end_d).days} 만료'
                    sub['status_key'] = status_key
                    sub['status_label'] = status_label
                    sub['dday'] = dday
                    sub['used_seats'] = used_map.get(sub['subject_code'], 0)
                    sub['access_open_date'] = str(open_d)
                    sub['subscription_end'] = str(end_d)
                    # 이력 (갱신 모달용)
                    cur.execute("""
                        SELECT event, plan, max_seats, start_term, term_count, fee, note, created_at
                        FROM subscription_history
                        WHERE subscription_id = %s
                        ORDER BY created_at
                    """, (sub['id'],))
                    hcols = [d[0] for d in cur.description]
                    sub['history'] = [
                        {**dict(zip(hcols, h)), 'created_at': str(h[-1])}
                        for h in cur.fetchall()
                    ]
                    subs.append(sub)
                inst['subscriptions'] = subs
                inst['admin_contacts'] = inst.get('admin_contacts') or []

        return jsonify({'ok': True, 'institutions': insts})
    finally:
        release_db_conn(conn)


def _upsert_admin_roster(cur, inst_id, contacts):
    """admin_contacts 각 행을 institution_rosters에 관리자(role='admin')로 등록/갱신.

    roster는 (institution_id, subject_code, email) 독립 행이며 관리자 행은 센티넬 과목코드
    '__ADMIN__'을 쓴다 → 같은 이메일이 과목 명단('HST' 등)과 충돌 없이 공존(§9).
    role='admin'(시스템 권한, 포털 접근)과 position(교수/조교 등, 표시용)은 별개 개념이다.
    이미 존재하는 이메일은 ON CONFLICT DO NOTHING(중복 무시) 후 name/position만 갱신.
    반환: 등록 대상 (email, name) 목록(등록 수 집계용 — 초대 메일은 발송하지 않는다, §9).
    """
    invited = []
    for c in contacts:
        email = (c.get('email') or '').strip().lower()
        if not email:
            continue
        name = (c.get('name') or '').strip() or None
        position = (c.get('position') or '').strip() or None
        cur.execute(
            """INSERT INTO institution_rosters
                   (institution_id, subject_code, email, name, role, position)
               VALUES (%s, %s, %s, %s, 'admin', %s)
               ON CONFLICT DO NOTHING""",
            (inst_id, ADMIN_ROSTER_SUBJECT, email, name, position),
        )
        # 기존 행이 있었으면 표시 정보(name/position)만 최신화(권한 role='admin' 유지).
        cur.execute(
            """UPDATE institution_rosters
                  SET name = %s, position = %s, role = 'admin'
                WHERE institution_id = %s AND subject_code = %s AND lower(email) = %s""",
            (name, position, inst_id, ADMIN_ROSTER_SUBJECT, email),
        )
        invited.append((email, name or ''))
    return invited


@app.route('/admin/api/institutions', methods=['POST'])
@super_admin_required
@admin_csrf_required
def api_institution_create():
    data = request.get_json(silent=True) or {}
    inst_id = (data.get('id') or '').strip().upper()
    university = (data.get('university') or '').strip()
    college = (data.get('college') or '').strip()
    name_en = (data.get('name_en') or '').strip() or None
    domain = (data.get('domain') or '').strip() or None
    contacts = data.get('admin_contacts') if data.get('admin_contacts') is not None else []
    subs_data = data.get('subscriptions') if data.get('subscriptions') is not None else []

    if not inst_id or not university or not college:
        return jsonify({'ok': False, 'error': '기관코드·학교명·단과대는 필수입니다'}), 400
    if not inst_id.replace('-', '').replace('_', '').isalnum():
        return jsonify({'ok': False, 'error': '기관코드는 영문·숫자·하이픈만 허용됩니다'}), 400
    # [Codex#4] 배열 타입 검증: admin_contacts·subscriptions가 list이고 각 원소가 dict인지 확인.
    cerr = _validate_array_of_dicts(contacts, 'admin_contacts')
    if cerr:
        return jsonify({'ok': False, 'error': cerr}), 400
    serr = _validate_array_of_dicts(subs_data, 'subscriptions')
    if serr:
        return jsonify({'ok': False, 'error': serr}), 400
    if len(contacts) > 5:
        return jsonify({'ok': False, 'error': '관리자는 최대 5명입니다'}), 400

    # [Codex#5] 트랜잭션 진입 전 구독 입력 전수 검증 → 잘못된 타입·범위는 400(500 아님).
    #   UI가 보낸 완전 빈 행(과목·학기·플랜 모두 공란)은 placeholder로 보고 건너뛴다.
    valid_subjects = _subject_codes_set()
    parsed_subs = []
    for sub in subs_data:
        if not (sub.get('subject_code') or '').strip() and \
           not (sub.get('start_term') or '').strip() and \
           not (sub.get('plan') or '').strip():
            continue
        fields, err = _validate_subscription_payload(sub, valid_subjects)
        if err:
            return jsonify({'ok': False, 'error': err}), 400
        parsed_subs.append(fields)

    name_ko = f'{university} {college}'
    conn = get_db_conn()
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute("SELECT id FROM institutions WHERE id = %s", (inst_id,))
                if cur.fetchone():
                    return jsonify({'ok': False, 'error': f'기관코드 {inst_id}는 이미 사용 중입니다'}), 409

                cur.execute("""
                    INSERT INTO institutions (id, name_ko, name_en, university, college, domain, admin_contacts)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (inst_id, name_ko, name_en, university, college, domain,
                      json.dumps(contacts, ensure_ascii=False)))

                for f in parsed_subs:
                    subject_code = f['subject_code']
                    plan = f['plan']
                    max_seats = f['max_seats']
                    term_count = f['term_count']
                    fee = f['fee']
                    start_term = f['start_term']
                    payment_method = f['payment_method']
                    end_term = _term_end_after(start_term, term_count)
                    open_date, _ = _sem_dates(start_term)
                    _, end_date = _sem_dates(end_term)
                    cur.execute("""
                        INSERT INTO subscriptions
                            (institution_id, subject_code, plan, max_seats,
                             start_term, term_count, access_open_date, subscription_end,
                             fee, payment_method)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        RETURNING id
                    """, (inst_id, subject_code, plan, max_seats,
                          start_term, term_count, open_date, end_date,
                          fee, payment_method))
                    sub_id = cur.fetchone()[0]
                    cur.execute("""
                        INSERT INTO subscription_history
                            (subscription_id, event, plan, max_seats, start_term, term_count, fee, created_by)
                        VALUES (%s, 'initial', %s, %s, %s, %s, %s, %s)
                    """, (sub_id, plan, max_seats, start_term, term_count, fee,
                          g.admin_user_id))

                    cur.execute("""
                        INSERT INTO institution_subject_access (institution_id, subject_code)
                        VALUES (%s, %s)
                        ON CONFLICT DO NOTHING
                    """, (inst_id, subject_code))

                # 관리자 명단 등록 (§9): admin_contacts → institution_rosters(role='admin').
                #   같은 트랜잭션으로 기관·구독과 함께 커밋. CEO 확정: 별도 포털 초대 메일은
                #   자동 발송하지 않는다 — 관리자도 학생과 동일하게 본인이 가입(register→verify)하며
                #   roster __ADMIN__ 행이 가입 대조의 근거다(§9).
                registered = _upsert_admin_roster(cur, inst_id, contacts)

        return jsonify({'ok': True, 'id': inst_id,
                        'admins_registered': len(registered)})
    except psycopg2.IntegrityError as e:
        return jsonify({'ok': False, 'error': str(e)}), 409
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        release_db_conn(conn)


@app.route('/admin/api/institutions/<inst_id>', methods=['GET'])
@super_admin_required
def api_institution_detail(inst_id):
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, name_ko, name_en, university, college, domain, admin_contacts
                FROM institutions WHERE id = %s
            """, (inst_id,))
            row = cur.fetchone()
            if not row:
                return jsonify({'ok': False, 'error': '기관을 찾을 수 없습니다'}), 404
            cols = [d[0] for d in cur.description]
            inst = dict(zip(cols, row))
            inst['admin_contacts'] = inst.get('admin_contacts') or []

            cur.execute("""
                SELECT s.id, s.subject_code, s.plan, s.max_seats,
                       s.start_term, s.term_count,
                       s.access_open_date, s.subscription_end,
                       s.fee, s.payment_method, s.status
                FROM subscriptions s
                WHERE s.institution_id = %s
                ORDER BY s.created_at
            """, (inst_id,))
            scols = [d[0] for d in cur.description]
            subs = []
            for sr in cur.fetchall():
                sub = dict(zip(scols, sr))
                sub['access_open_date'] = str(sub['access_open_date'])
                sub['subscription_end'] = str(sub['subscription_end'])
                cur.execute("""
                    SELECT event, plan, max_seats, start_term, term_count, fee, note, created_at
                    FROM subscription_history
                    WHERE subscription_id = %s
                    ORDER BY created_at
                """, (sub['id'],))
                hcols = [d[0] for d in cur.description]
                sub['history'] = [dict(zip(hcols, h)) for h in cur.fetchall()]
                for h in sub['history']:
                    h['created_at'] = str(h['created_at'])
                subs.append(sub)
            inst['subscriptions'] = subs

        return jsonify({'ok': True, 'institution': inst})
    finally:
        release_db_conn(conn)


@app.route('/admin/api/institutions/<inst_id>', methods=['PUT'])
@super_admin_required
@admin_csrf_required
def api_institution_update(inst_id):
    data = request.get_json(silent=True) or {}
    university = (data.get('university') or '').strip()
    college = (data.get('college') or '').strip()
    if not university or not college:
        return jsonify({'ok': False, 'error': '학교명·단과대는 필수입니다'}), 400
    name_ko = f'{university} {college}'
    name_en = (data.get('name_en') or '').strip() or None
    domain = (data.get('domain') or '').strip() or None
    contacts = data.get('admin_contacts') if data.get('admin_contacts') is not None else []
    # [Codex#4] 배열 타입 검증
    cerr = _validate_array_of_dicts(contacts, 'admin_contacts')
    if cerr:
        return jsonify({'ok': False, 'error': cerr}), 400
    if len(contacts) > 5:
        return jsonify({'ok': False, 'error': '관리자는 최대 5명입니다'}), 400

    # 현재 폼이 보낸 관리자 이메일 집합(소문자 정규화).
    contact_emails = {
        (c.get('email') or '').strip().lower()
        for c in contacts if (c.get('email') or '').strip()
    }

    conn = get_db_conn()
    try:
        to_add = set()
        to_remove = set()
        with conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE institutions
                    SET name_ko=%s, name_en=%s, university=%s, college=%s,
                        domain=%s, admin_contacts=%s
                    WHERE id=%s
                """, (name_ko, name_en, university, college, domain,
                      json.dumps(contacts, ensure_ascii=False), inst_id))

                # 관리자 명단 변경분 반영 (§9). 기존 admin roster와 폼 입력의 차집합 계산.
                cur.execute(
                    """SELECT lower(email) FROM institution_rosters
                       WHERE institution_id = %s AND role = 'admin'""",
                    (inst_id,),
                )
                existing = {r[0] for r in cur.fetchall()}
                to_add = contact_emails - existing
                to_remove = existing - contact_emails

                # 추가/유지: roster upsert(role='admin' 유지, name/position 갱신).
                _upsert_admin_roster(cur, inst_id, contacts)

                # 제거 = "포털 관리 권한만 회수"다(관리/열람 분리 원칙). admin roster 행(__ADMIN__)만
                #   DELETE한다. users 계정·다른 과목 roster 행은 절대 건드리지 않는다 → 겸직(학생/교수)이면
                #   슬라이드 열람 유지, 순수 관리자였으면 단지 포털 접근만 사라진다(계정 정지 아님).
                for em in to_remove:
                    cur.execute(
                        """DELETE FROM institution_rosters
                           WHERE institution_id = %s AND subject_code = %s AND lower(email) = %s""",
                        (inst_id, ADMIN_ROSTER_SUBJECT, em),
                    )

        # CEO 확정: 포털 초대 메일 자동 발송 없음 — 관리자는 본인이 가입(register→verify)한다(§9).
        return jsonify({'ok': True, 'admins_added': len(to_add),
                        'admins_removed': len(to_remove)})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        release_db_conn(conn)


@app.route('/admin/api/institutions/<inst_id>/subscriptions', methods=['POST'])
@super_admin_required
@admin_csrf_required
def api_subscription_add(inst_id):
    data = request.get_json(silent=True) or {}
    # [Codex#5] 타입·범위·allowlist 검증 → 실패 시 400(500 아님).
    fields, err = _validate_subscription_payload(data, _subject_codes_set())
    if err:
        return jsonify({'ok': False, 'error': err}), 400
    subject_code = fields['subject_code']
    plan = fields['plan']
    max_seats = fields['max_seats']
    term_count = fields['term_count']
    fee = fields['fee']
    start_term = fields['start_term']
    payment_method = fields['payment_method']

    end_term = _term_end_after(start_term, term_count)
    open_date, _ = _sem_dates(start_term)
    _, end_date = _sem_dates(end_term)

    conn = get_db_conn()
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute("SELECT id FROM institutions WHERE id=%s", (inst_id,))
                if not cur.fetchone():
                    return jsonify({'ok': False, 'error': '기관을 찾을 수 없습니다'}), 404
                cur.execute("""
                    INSERT INTO subscriptions
                        (institution_id, subject_code, plan, max_seats,
                         start_term, term_count, access_open_date, subscription_end,
                         fee, payment_method)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING id
                """, (inst_id, subject_code, plan, max_seats,
                      start_term, term_count, open_date, end_date,
                      fee, payment_method))
                sub_id = cur.fetchone()[0]
                cur.execute("""
                    INSERT INTO subscription_history
                        (subscription_id, event, plan, max_seats, start_term, term_count, fee, created_by)
                    VALUES (%s, 'initial', %s, %s, %s, %s, %s, %s)
                """, (sub_id, plan, max_seats, start_term, term_count, fee, g.admin_user_id))
                cur.execute("""
                    INSERT INTO institution_subject_access (institution_id, subject_code)
                    VALUES (%s, %s) ON CONFLICT DO NOTHING
                """, (inst_id, subject_code))
        return jsonify({'ok': True, 'subscription_id': sub_id})
    except psycopg2.IntegrityError:
        return jsonify({'ok': False, 'error': '이미 동일 기관·과목·시작학기 구독이 존재합니다'}), 409
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        release_db_conn(conn)


@app.route('/admin/api/institutions/<inst_id>/subscriptions/<int:sub_id>/renew', methods=['POST'])
@super_admin_required
@admin_csrf_required
def api_subscription_renew(inst_id, sub_id):
    data = request.get_json(silent=True) or {}
    # [Codex#5] 타입·범위·plan allowlist 검증 → 실패 시 400. subject_code는 기존 구독에서 승계.
    plan = (data.get('plan') or '').strip().lower()
    if plan and plan not in VALID_PLANS:
        return jsonify({'ok': False, 'error': f'알 수 없는 플랜입니다: {plan}'}), 400
    extra_seats, err = _parse_int_field(data.get('extra_seats'), 'extra_seats',
                                        default=0, minimum=0, maximum=MAX_SEATS_LIMIT)
    if err:
        return jsonify({'ok': False, 'error': err}), 400
    term_count, err = _parse_int_field(data.get('term_count'), 'term_count',
                                       default=1, minimum=1, maximum=MAX_TERM_COUNT)
    if err:
        return jsonify({'ok': False, 'error': err}), 400
    fee, err = _parse_int_field(data.get('fee'), 'fee', default=None,
                                minimum=0, maximum=MAX_FEE_LIMIT, allow_none=True)
    if err:
        return jsonify({'ok': False, 'error': err}), 400
    payment_method = data.get('payment_method') or '학기 선불'
    note = data.get('note') or None

    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT institution_id, subject_code, plan, max_seats, start_term, term_count,
                       subscription_end
                FROM subscriptions WHERE id=%s AND institution_id=%s
            """, (sub_id, inst_id))
            row = cur.fetchone()
        if not row:
            return jsonify({'ok': False, 'error': '구독을 찾을 수 없습니다'}), 404
        _, subject_code, old_plan, old_seats, old_start, old_count, old_end = row
        new_plan = plan or old_plan
        new_seats = (PLAN_SEATS.get(new_plan, old_seats) + extra_seats) if plan else (old_seats + extra_seats)
        old_end_term = _term_end_after(old_start, old_count)
        new_start = _next_term(old_end_term)
        end_term = _term_end_after(new_start, term_count)
        open_date, _ = _sem_dates(new_start)
        _, end_date = _sem_dates(end_term)

        with conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO subscriptions
                        (institution_id, subject_code, plan, max_seats,
                         start_term, term_count, access_open_date, subscription_end,
                         fee, payment_method)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING id
                """, (inst_id, subject_code, new_plan, new_seats,
                      new_start, term_count, open_date, end_date,
                      fee, payment_method))
                new_sub_id = cur.fetchone()[0]
                cur.execute("""
                    INSERT INTO subscription_history
                        (subscription_id, event, plan, max_seats, start_term, term_count, fee, note, created_by)
                    VALUES (%s, 'renewal', %s, %s, %s, %s, %s, %s, %s)
                """, (new_sub_id, new_plan, new_seats, new_start, term_count, fee, note, g.admin_user_id))
        return jsonify({'ok': True, 'subscription_id': new_sub_id,
                        'start_term': new_start, 'end_term': end_term,
                        'access_open_date': str(open_date),
                        'subscription_end': str(end_date)})
    except psycopg2.IntegrityError:
        return jsonify({'ok': False, 'error': '해당 학기 구독이 이미 존재합니다'}), 409
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        release_db_conn(conn)


# ── 접근 제어 (S5-4) ─────────────────────────────────────────

# v1.0: is_active 컬럼 없이 코드 파생.
# subject_codes 테이블에 is_active BOOLEAN 컬럼 추가는 두 번째 모듈 출시 시 ALTER로 진행.
_ACTIVE_SUBJECTS = frozenset({'HST'})   # v1.0: 조직학만 활성


@app.route('/admin/access')
@super_admin_required
def admin_access_page():
    return render_template('admin/access.html', active_page='access')


@app.route('/admin/api/access/modules', methods=['GET'])
@super_admin_required
def api_access_modules():
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT code, name_ko, name_en FROM subject_codes ORDER BY code")
            subjects = [{'code': r[0], 'name_ko': r[1], 'name_en': r[2]} for r in cur.fetchall()]

            cur.execute("""
                SELECT subject_code, COUNT(*) AS cnt
                FROM slides
                WHERE deploy_status = 'deployed'
                GROUP BY subject_code
            """)
            slide_counts = {r[0]: r[1] for r in cur.fetchall()}

        modules = []
        for s in subjects:
            code = s['code']
            is_active = code in _ACTIVE_SUBJECTS
            modules.append({
                'code': code,
                'name_ko': s['name_ko'],
                'name_en': s['name_en'],
                'is_active': is_active,
                'slide_count': slide_counts.get(code, 0),
                'grant_mode': '전 구독 자동 부여 (기본 모듈)' if is_active else '기관별 토글 (출시 후)',
                'addon': '—' if is_active else '과목별 독립 좌석',
            })
        return jsonify({'ok': True, 'modules': modules})
    finally:
        release_db_conn(conn)


@app.route('/admin/api/access/matrix', methods=['GET'])
@super_admin_required
def api_access_matrix():
    """기관×과목 접근 매트릭스. 읽기 전용 (v1.0 미리보기)."""
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT id, name_ko FROM institutions ORDER BY name_ko")
            institutions = [{'id': r[0], 'name_ko': r[1]} for r in cur.fetchall()]

            # 매트릭스 열: 주요 과목만 (HST·PATH·PARA)
            matrix_subjects = ['HST', 'PATH', 'PARA']

            cur.execute("""
                SELECT institution_id, subject_code, granted
                FROM institution_subject_access
                WHERE subject_code = ANY(%s)
            """, (matrix_subjects,))
            access_map = {}
            for inst_id, subj, granted in cur.fetchall():
                access_map.setdefault(inst_id, {})[subj] = granted

        matrix = []
        for inst in institutions:
            row = {'id': inst['id'], 'name_ko': inst['name_ko'], 'access': {}}
            for subj in matrix_subjects:
                if subj in _ACTIVE_SUBJECTS:
                    # HST: 항상 true (잠김)
                    row['access'][subj] = {'granted': True, 'locked': True}
                else:
                    granted = access_map.get(inst['id'], {}).get(subj, False)
                    row['access'][subj] = {'granted': granted, 'locked': True}  # v1.0: 모두 잠김(미리보기)
            matrix.append(row)

        return jsonify({'ok': True, 'subjects': matrix_subjects, 'matrix': matrix})
    finally:
        release_db_conn(conn)


# ── 이용 리포트 (S5-5) ───────────────────────────────────

@app.route('/admin/reports')
@super_admin_required
def admin_reports_page():
    return render_template('admin/reports.html', active_page='reports')


def _reports_date_range(period: str, inst_id: str, subject_code: str, conn):
    """기간 파라미터 → (start_date, end_date) 또는 (None, None)."""
    from datetime import date, timedelta
    if period == '30d':
        today = date.today()
        return today - timedelta(days=30), today
    if period == 'all':
        return None, None
    # 기본: 현재 활성 구독 기간
    with conn.cursor() as cur:
        cur.execute("""
            SELECT access_open_date, subscription_end
            FROM subscriptions
            WHERE institution_id = %s AND subject_code = %s
              AND status = 'active'
            ORDER BY subscription_end DESC LIMIT 1
        """, (inst_id, subject_code))
        row = cur.fetchone()
    if row:
        return row[0], row[1]
    return None, None


@app.route('/admin/api/reports/institutions', methods=['GET'])
@super_admin_required
def api_reports_institutions():
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT id, name_ko FROM institutions ORDER BY name_ko")
            rows = [{'id': r[0], 'name_ko': r[1]} for r in cur.fetchall()]
        return jsonify({'ok': True, 'institutions': rows})
    finally:
        release_db_conn(conn)


@app.route('/admin/api/reports/kpi', methods=['GET'])
@super_admin_required
def api_reports_kpi():
    inst_id      = request.args.get('inst_id', '').strip()
    subject_code = request.args.get('subject_code', 'HST').strip()
    period       = request.args.get('period', 'term')

    if not inst_id:
        return jsonify({'ok': False, 'error': '기관을 선택하세요'}), 400

    conn = get_db_conn()
    try:
        start_dt, end_dt = _reports_date_range(period, inst_id, subject_code, conn)

        # 날짜 필터 SQL 조건
        date_where = ""
        date_params: list = []
        if start_dt and end_dt:
            date_where = "AND al.accessed_at BETWEEN %s AND %s + INTERVAL '1 day'"
            date_params = [start_dt, end_dt]

        with conn.cursor() as cur:
            # 활성 사용자 수 + 최대 좌석
            # [§18 D9] 과목별 산출 — active_users는 (기관 × 과목) 좌석 카운터다(§15-7).
            #   subject_code 필터를 추가해 좌석 소진율 = 활성(과목)/max_seats(과목)이 되게 한다.
            #   기관 롤업은 과목별 산출값의 합(과목 카드별 KPI)으로 자연 구성된다(과목축 미혼합).
            cur.execute("""
                SELECT COUNT(u.id)
                FROM users u
                WHERE u.institution_id = %s
                  AND u.subject_code = %s
                  AND COALESCE(u.status, 'active') = 'active'
                  AND COALESCE(u.is_special, FALSE) = FALSE
            """, (inst_id, subject_code))
            active_users = cur.fetchone()[0] or 0

            cur.execute("""
                SELECT COALESCE(SUM(max_seats), 0)
                FROM subscriptions
                WHERE institution_id = %s AND subject_code = %s AND status = 'active'
            """, (inst_id, subject_code))
            max_seats = cur.fetchone()[0] or 0

            # 총 열람 수
            cur.execute(f"""
                SELECT COUNT(al.id)
                FROM access_logs al
                JOIN users u ON al.user_id = u.id
                WHERE u.institution_id = %s {date_where}
            """, [inst_id] + date_params)
            total_views = cur.fetchone()[0] or 0

            # AI 튜터 질문 수 (chat_logs는 created_at 컬럼, access_logs와 별개)
            ai_questions = 0
            chat_date_where = ""
            if start_dt and end_dt:
                chat_date_where = "AND cl.created_at BETWEEN %s AND %s + INTERVAL '1 day'"
            try:
                cur.execute(f"""
                    SELECT COUNT(cl.id)
                    FROM chat_logs cl
                    WHERE cl.institution_id = %s {chat_date_where}
                """, [inst_id] + date_params)
                ai_questions = cur.fetchone()[0] or 0
            except Exception:
                conn.rollback()

            # 마지막 활동
            cur.execute(f"""
                SELECT MAX(al.accessed_at)
                FROM access_logs al
                JOIN users u ON al.user_id = u.id
                WHERE u.institution_id = %s {date_where}
            """, [inst_id] + date_params)
            last_activity = cur.fetchone()[0]

        util_pct = round(active_users / max_seats * 100) if max_seats > 0 else 0
        per_user = round(ai_questions / active_users, 1) if active_users > 0 else 0

        return jsonify({
            'ok': True,
            'active_users': active_users,
            'max_seats':    max_seats,
            'util_pct':     util_pct,
            'total_views':  total_views,
            'ai_questions': ai_questions,
            'per_user':     per_user,
            'last_activity': last_activity.strftime('%Y-%m-%d %H:%M') if last_activity else None,
        })
    finally:
        release_db_conn(conn)


@app.route('/admin/api/reports/weekly', methods=['GET'])
@super_admin_required
def api_reports_weekly():
    inst_id      = request.args.get('inst_id', '').strip()
    subject_code = request.args.get('subject_code', 'HST').strip()
    period       = request.args.get('period', 'term')

    if not inst_id:
        return jsonify({'ok': False, 'error': '기관을 선택하세요'}), 400

    conn = get_db_conn()
    try:
        start_dt, end_dt = _reports_date_range(period, inst_id, subject_code, conn)

        date_where = ""
        date_params: list = []
        if start_dt and end_dt:
            date_where = "AND al.accessed_at BETWEEN %s AND %s + INTERVAL '1 day'"
            date_params = [start_dt, end_dt]

        with conn.cursor() as cur:
            cur.execute(f"""
                SELECT
                    TO_CHAR(DATE_TRUNC('week', al.accessed_at), 'MM/DD') AS week_label,
                    COUNT(DISTINCT al.user_id) AS logins
                FROM access_logs al
                JOIN users u ON al.user_id = u.id
                WHERE u.institution_id = %s {date_where}
                GROUP BY DATE_TRUNC('week', al.accessed_at)
                ORDER BY DATE_TRUNC('week', al.accessed_at)
                LIMIT 12
            """, [inst_id] + date_params)
            rows = cur.fetchall()

        weeks = [{'label': r[0], 'count': r[1]} for r in rows]
        max_val = max((w['count'] for w in weeks), default=1)
        for w in weeks:
            w['pct'] = round(w['count'] / max_val * 100) if max_val else 0

        return jsonify({'ok': True, 'weeks': weeks})
    finally:
        release_db_conn(conn)


@app.route('/admin/api/reports/top-slides', methods=['GET'])
@super_admin_required
def api_reports_top_slides():
    inst_id      = request.args.get('inst_id', '').strip()
    subject_code = request.args.get('subject_code', 'HST').strip()
    period       = request.args.get('period', 'term')

    if not inst_id:
        return jsonify({'ok': False, 'error': '기관을 선택하세요'}), 400

    conn = get_db_conn()
    try:
        start_dt, end_dt = _reports_date_range(period, inst_id, subject_code, conn)

        date_where = ""
        date_params: list = []
        if start_dt and end_dt:
            date_where = "AND al.accessed_at BETWEEN %s AND %s + INTERVAL '1 day'"
            date_params = [start_dt, end_dt]

        with conn.cursor() as cur:
            cur.execute(f"""
                SELECT s.title_ko, s.stain, COUNT(al.id) AS views
                FROM access_logs al
                JOIN users u ON al.user_id = u.id
                JOIN slides s ON al.slide_id = s.id
                WHERE u.institution_id = %s
                  AND s.subject_code = %s {date_where}
                GROUP BY s.id, s.title_ko, s.stain
                ORDER BY views DESC
                LIMIT 5
            """, [inst_id, subject_code] + date_params)
            rows = cur.fetchall()

        slides = [{'title': r[0], 'stain': r[1], 'views': r[2]} for r in rows]
        max_val = slides[0]['views'] if slides else 1
        for sl in slides:
            sl['pct'] = round(sl['views'] / max_val * 100) if max_val else 0

        return jsonify({'ok': True, 'slides': slides})
    finally:
        release_db_conn(conn)


def _xlsx_safe(v):
    """[2-2#4] 스프레드시트 수식 주입 방어(§18 D9): =,+,-,@(및 탭/CR로 시작) 문자열 셀은
    앞에 작은따옴표(')를 붙여 텍스트로 무력화. 숫자/None 등은 그대로 둔다.
    """
    if isinstance(v, str) and v and v[0] in ('=', '+', '-', '@', '\t', '\r'):
        return "'" + v
    return v


@app.route('/admin/api/reports/export/excel', methods=['GET'])
@super_admin_required
def api_reports_export_excel():
    inst_id      = request.args.get('inst_id', '').strip()
    subject_code = request.args.get('subject_code', 'HST').strip()
    period       = request.args.get('period', 'term')

    if not inst_id:
        return jsonify({'ok': False, 'error': '기관을 선택하세요'}), 400

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({'ok': False, 'error': 'openpyxl 미설치'}), 500

    conn = get_db_conn()
    try:
        start_dt, end_dt = _reports_date_range(period, inst_id, subject_code, conn)

        date_where = ""
        date_params: list = []
        if start_dt and end_dt:
            date_where = "AND al.accessed_at BETWEEN %s AND %s + INTERVAL '1 day'"
            date_params = [start_dt, end_dt]

        with conn.cursor() as cur:
            # 기관명
            cur.execute("SELECT name_ko FROM institutions WHERE id = %s", (inst_id,))
            row = cur.fetchone()
            inst_name = row[0] if row else inst_id

            # KPI
            cur.execute("""
                SELECT COUNT(u.id) FROM users u
                WHERE u.institution_id = %s
                  AND COALESCE(u.status, 'active') = 'active'
                  AND COALESCE(u.is_special, FALSE) = FALSE
            """, (inst_id,))
            active_users = cur.fetchone()[0] or 0

            cur.execute("""
                SELECT COALESCE(SUM(max_seats), 0) FROM subscriptions
                WHERE institution_id = %s AND subject_code = %s AND status = 'active'
            """, (inst_id, subject_code))
            max_seats = cur.fetchone()[0] or 0

            cur.execute(f"""
                SELECT COUNT(al.id) FROM access_logs al
                JOIN users u ON al.user_id = u.id
                WHERE u.institution_id = %s {date_where}
            """, [inst_id] + date_params)
            total_views = cur.fetchone()[0] or 0

            # 주간 추세
            cur.execute(f"""
                SELECT TO_CHAR(DATE_TRUNC('week', al.accessed_at), 'YYYY-MM-DD') AS wk,
                       COUNT(DISTINCT al.user_id)
                FROM access_logs al
                JOIN users u ON al.user_id = u.id
                WHERE u.institution_id = %s {date_where}
                GROUP BY DATE_TRUNC('week', al.accessed_at)
                ORDER BY 1 LIMIT 12
            """, [inst_id] + date_params)
            weekly_rows = cur.fetchall()

            # 많이 본 슬라이드
            cur.execute(f"""
                SELECT s.id, s.title_ko, s.stain, COUNT(al.id) AS views
                FROM access_logs al
                JOIN users u ON al.user_id = u.id
                JOIN slides s ON al.slide_id = s.id
                WHERE u.institution_id = %s
                  AND s.subject_code = %s {date_where}
                GROUP BY s.id, s.title_ko, s.stain
                ORDER BY views DESC LIMIT 10
            """, [inst_id, subject_code] + date_params)
            top_slides = cur.fetchall()
    finally:
        release_db_conn(conn)

    # Excel 생성
    wb = openpyxl.Workbook()
    HDR = Font(bold=True, color='FFFFFF')
    HDR_FILL = PatternFill('solid', fgColor='0F1F3D')
    center = Alignment(horizontal='center')

    # 시트 1: 요약
    ws = wb.active
    ws.title = '요약'
    ws.append([_xlsx_safe(f'{inst_name} 이용 리포트')])
    ws['A1'].font = Font(bold=True, size=13)
    period_label = {'term': '이번 학기', '30d': '최근 30일', 'all': '전체 기간'}.get(period, period)
    ws.append([_xlsx_safe(f'기간: {period_label}  /  과목: {subject_code}')])
    ws.append([])
    ws.append(['지표', '수치'])
    for cell in ws[4]:
        cell.font = HDR; cell.fill = HDR_FILL; cell.alignment = center
    ws.append(['활성 사용자', active_users])
    ws.append(['최대 좌석', max_seats])
    ws.append(['좌석 소진율 (%)', round(active_users / max_seats * 100) if max_seats else 0])
    ws.append(['총 슬라이드 열람', total_views])
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14

    # 시트 2: 주간 추세
    ws2 = wb.create_sheet('주간 추세')
    ws2.append(['주차 시작일', '순방문자(주)'])
    for cell in ws2[1]:
        cell.font = HDR; cell.fill = HDR_FILL
    for wk, cnt in weekly_rows:
        ws2.append([_xlsx_safe(wk), cnt])
    ws2.column_dimensions['A'].width = 16
    ws2.column_dimensions['B'].width = 14

    # 시트 3: 많이 본 슬라이드
    ws3 = wb.create_sheet('많이 본 슬라이드')
    ws3.append(['슬라이드 ID', '제목', '염색', '열람 수'])
    for cell in ws3[1]:
        cell.font = HDR; cell.fill = HDR_FILL
    for slide_id, title, stain, views in top_slides:
        ws3.append([_xlsx_safe(slide_id), _xlsx_safe(title), _xlsx_safe(stain), views])
    for col in ['A', 'B', 'C', 'D']:
        ws3.column_dimensions[col].width = 18

    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from flask import send_file as _sf
    fname = f'SlideAtlas_report_{inst_id}_{period}.xlsx'
    return _sf(buf, as_attachment=True, download_name=fname,
               mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── 특별 계정 (S5-6) ─────────────────────────────────────

@app.route('/admin/special')
@super_admin_required
def admin_special_page():
    return render_template('admin/special.html', active_page='special')


@app.route('/admin/api/special/accounts', methods=['GET'])
@super_admin_required
def api_special_accounts_list():
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    u.id, u.email, u.institution_id, i.name_ko AS inst_name,
                    u.special_purpose, u.special_expires_at, u.special_review_at,
                    u.special_created_by, a.name AS created_by_name,
                    COALESCE(u.status, 'active') AS status,
                    u.created_at
                FROM users u
                LEFT JOIN institutions i ON u.institution_id = i.id
                LEFT JOIN admin_users  a ON u.special_created_by = a.id
                WHERE COALESCE(u.is_special, FALSE) = TRUE
                ORDER BY u.created_at DESC
            """)
            rows = cur.fetchall()

        from datetime import date, timedelta
        today = date.today()
        accounts = []
        for r in rows:
            expires = r[5]
            review  = r[6]
            dday    = None
            dday_type = None  # 'warn'|'danger'|'ok'
            if expires:
                delta = (expires - today).days
                dday = delta
                dday_type = 'danger' if delta < 0 else ('warn' if delta <= 30 else 'ok')

            accounts.append({
                'id':              r[0],
                'email':           r[1],
                'institution_id':  r[2],
                'inst_name':       r[3],
                'purpose':         r[4],
                'expires_at':      expires.isoformat() if expires else None,
                'review_at':       review.isoformat()  if review  else None,
                'created_by_name': r[8],
                'status':          r[9],
                'dday':            dday,
                'dday_type':       dday_type,
                'no_expiry':       expires is None,
            })

        return jsonify({'ok': True, 'accounts': accounts})
    finally:
        release_db_conn(conn)


@app.route('/admin/api/special/accounts', methods=['POST'])
@super_admin_required
@admin_csrf_required
def api_special_accounts_create():
    data = request.get_json(silent=True) or {}
    email       = (data.get('email') or '').strip().lower()
    purpose     = (data.get('purpose') or '').strip()
    expires_at  = data.get('expires_at') or None
    review_at   = data.get('review_at')  or None
    inst_id     = (data.get('institution_id') or '').strip() or None

    if not email or not purpose:
        return jsonify({'ok': False, 'error': '이메일과 용도는 필수입니다'}), 400

    # 이메일 기본 검증
    if '@' not in email or len(email) > 200:
        return jsonify({'ok': False, 'error': '이메일 형식이 올바르지 않습니다'}), 400

    conn = get_db_conn()
    try:
        with conn:
            with conn.cursor() as cur:
                # 이미 존재하면 is_special 업데이트
                cur.execute("SELECT id FROM users WHERE email = %s", (email,))
                existing = cur.fetchone()
                if existing:
                    # [Codex 2R#1 §0] 특별계정은 좌석을 점유하지 않는다(CEO 결정) → 승격 시 subject_code=NULL
                    #   로 정리해 active_seat_count(P2 좌석)·P3 active_users 양쪽에서 동일하게 빠지게 한다
                    #   ('subject_code 있는 active 사용자'만 셈 → 같은 집합). position 도 함께 비운다.
                    cur.execute("""
                        UPDATE users SET
                            is_special         = TRUE,
                            special_purpose    = %s,
                            special_expires_at = %s,
                            special_review_at  = %s,
                            special_created_by = %s,
                            institution_id     = COALESCE(%s, institution_id),
                            status             = 'active',
                            subject_code       = NULL,
                            position           = NULL
                        WHERE id = %s
                    """, (purpose, expires_at, review_at,
                          g.admin_user_id, inst_id, existing[0]))
                    user_id = existing[0]
                else:
                    from werkzeug.security import generate_password_hash as _gph
                    import secrets as _sec
                    temp_pw = _gph(_sec.token_hex(16))
                    cur.execute("""
                        INSERT INTO users
                            (email, password_hash, institution_id, is_special,
                             special_purpose, special_expires_at, special_review_at,
                             special_created_by, status)
                        VALUES (%s, %s, %s, TRUE, %s, %s, %s, %s, 'active')
                        RETURNING id
                    """, (email, temp_pw, inst_id, purpose,
                          expires_at, review_at, g.admin_user_id))
                    user_id = cur.fetchone()[0]

        return jsonify({'ok': True, 'user_id': user_id})
    finally:
        release_db_conn(conn)


@app.route('/admin/api/special/accounts/<int:user_id>/deactivate', methods=['POST'])
@super_admin_required
@admin_csrf_required
def api_special_accounts_deactivate(user_id):
    conn = get_db_conn()
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE users SET status = 'suspended'
                    WHERE id = %s AND COALESCE(is_special, FALSE) = TRUE
                    RETURNING id
                """, (user_id,))
                if not cur.fetchone():
                    return jsonify({'ok': False, 'error': '계정을 찾을 수 없습니다'}), 404
        return jsonify({'ok': True})
    finally:
        release_db_conn(conn)


@app.route('/admin/api/special/accounts/<int:user_id>/activate', methods=['POST'])
@super_admin_required
@admin_csrf_required
def api_special_accounts_activate(user_id):
    conn = get_db_conn()
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE users SET status = 'active'
                    WHERE id = %s AND COALESCE(is_special, FALSE) = TRUE
                    RETURNING id
                """, (user_id,))
                if not cur.fetchone():
                    return jsonify({'ok': False, 'error': '계정을 찾을 수 없습니다'}), 404
        return jsonify({'ok': True})
    finally:
        release_db_conn(conn)


# ── 공지 관리 (S5-7) ─────────────────────────────────────

@app.route('/admin/notices')
@admin_required
def admin_notices_page():
    return render_template('admin/notices.html', active_page='notice')


@app.route('/admin/api/notices/list', methods=['GET'])
@admin_required
def api_notices_list():
    view = request.args.get('view', 'active')   # 'active' | 'archive'
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            if view == 'archive':
                cur.execute("""
                    SELECT id, title, body, is_published, display_order,
                           created_by, updated_by, archived_at, updated_at,
                           ac.name AS created_name, ac.role AS created_role,
                           au.name AS updated_name, au.role AS updated_role
                    FROM announcements a
                    LEFT JOIN admin_users ac ON a.created_by = ac.id
                    LEFT JOIN admin_users au ON a.updated_by = au.id
                    WHERE is_archived = TRUE
                    ORDER BY archived_at DESC
                """)
            else:
                cur.execute("""
                    SELECT id, title, body, is_published, display_order,
                           created_by, updated_by, archived_at, updated_at,
                           ac.name AS created_name, ac.role AS created_role,
                           au.name AS updated_name, au.role AS updated_role
                    FROM announcements a
                    LEFT JOIN admin_users ac ON a.created_by = ac.id
                    LEFT JOIN admin_users au ON a.updated_by = au.id
                    WHERE is_archived = FALSE
                    ORDER BY is_published DESC, display_order ASC, updated_at DESC
                """)
            rows = cur.fetchall()

            # 게시 중 개수
            cur.execute("""
                SELECT COUNT(*) FROM announcements
                WHERE is_published = TRUE AND is_archived = FALSE
            """)
            published_count = cur.fetchone()[0] or 0

            cur.execute("""
                SELECT COUNT(*) FROM announcements WHERE is_archived = TRUE
            """)
            archive_count = cur.fetchone()[0] or 0

        def _row(r):
            editor = r[11] or r[9]  # updated_name 우선, 없으면 created_name
            editor_role = r[12] or r[10]
            return {
                'id':             r[0],
                'title':          r[1],
                'body':           r[2],
                'is_published':   r[3],
                'display_order':  r[4],
                'archived_at':    r[7].strftime('%m-%d') if r[7] else None,
                'updated_at':     r[8].strftime('%m-%d') if r[8] else None,
                'editor':         editor or '',
                'editor_role':    'super' if editor_role == 'super_admin' else 'staff',
            }

        return jsonify({
            'ok': True,
            'items': [_row(r) for r in rows],
            'published_count': published_count,
            'archive_count':   archive_count,
        })
    finally:
        release_db_conn(conn)


@app.route('/admin/api/notices', methods=['POST'])
@admin_required
@admin_csrf_required
def api_notices_create():
    data = request.get_json(silent=True) or {}
    title   = (data.get('title') or '').strip()
    body    = (data.get('body')  or '').strip()
    publish = bool(data.get('is_published', False))

    if not title:
        return jsonify({'ok': False, 'error': '제목을 입력하세요'}), 400

    conn = get_db_conn()
    try:
        with conn:
            with conn.cursor() as cur:
                # 게시 순서: 기존 최대 display_order + 1
                if publish:
                    cur.execute("""
                        SELECT COALESCE(MAX(display_order), 0)
                        FROM announcements WHERE is_published = TRUE AND is_archived = FALSE
                    """)
                    next_order = (cur.fetchone()[0] or 0) + 1
                else:
                    next_order = 0

                cur.execute("""
                    INSERT INTO announcements
                        (title, body, is_published, display_order, created_by, updated_by)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    RETURNING id
                """, (title, body, publish, next_order,
                      g.admin_user_id, g.admin_user_id))
                new_id = cur.fetchone()[0]

        return jsonify({'ok': True, 'id': new_id})
    finally:
        release_db_conn(conn)


@app.route('/admin/api/notices/<int:notice_id>', methods=['PUT'])
@admin_required
@admin_csrf_required
def api_notices_update(notice_id):
    data = request.get_json(silent=True) or {}
    title         = (data.get('title') or '').strip()
    body          = data.get('body', '')
    is_published  = data.get('is_published')   # None = don't change
    display_order = data.get('display_order')  # None = don't change

    if not title:
        return jsonify({'ok': False, 'error': '제목을 입력하세요'}), 400

    conn = get_db_conn()
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute("SELECT id FROM announcements WHERE id = %s AND is_archived = FALSE", (notice_id,))
                if not cur.fetchone():
                    return jsonify({'ok': False, 'error': '공지를 찾을 수 없습니다'}), 404

                set_clauses = ["title = %s", "body = %s", "updated_by = %s", "updated_at = NOW()"]
                params: list = [title, body or '', g.admin_user_id]

                if is_published is not None:
                    set_clauses.append("is_published = %s")
                    params.append(bool(is_published))
                if display_order is not None:
                    set_clauses.append("display_order = %s")
                    params.append(int(display_order))

                params.append(notice_id)
                cur.execute(
                    f"UPDATE announcements SET {', '.join(set_clauses)} WHERE id = %s",
                    params,
                )

        return jsonify({'ok': True})
    finally:
        release_db_conn(conn)


@app.route('/admin/api/notices/<int:notice_id>/toggle', methods=['POST'])
@admin_required
@admin_csrf_required
def api_notices_toggle(notice_id):
    conn = get_db_conn()
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT is_published, display_order
                    FROM announcements WHERE id = %s AND is_archived = FALSE
                """, (notice_id,))
                row = cur.fetchone()
                if not row:
                    return jsonify({'ok': False, 'error': '공지를 찾을 수 없습니다'}), 404

                currently_published = row[0]
                new_published = not currently_published

                if new_published:
                    cur.execute("""
                        SELECT COALESCE(MAX(display_order), 0)
                        FROM announcements WHERE is_published = TRUE AND is_archived = FALSE AND id != %s
                    """, (notice_id,))
                    new_order = (cur.fetchone()[0] or 0) + 1
                else:
                    new_order = 0

                cur.execute("""
                    UPDATE announcements
                    SET is_published = %s, display_order = %s,
                        updated_by = %s, updated_at = NOW()
                    WHERE id = %s
                """, (new_published, new_order, g.admin_user_id, notice_id))

        return jsonify({'ok': True, 'is_published': new_published})
    finally:
        release_db_conn(conn)


@app.route('/admin/api/notices/<int:notice_id>/archive', methods=['POST'])
@admin_required
@admin_csrf_required
def api_notices_archive(notice_id):
    conn = get_db_conn()
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE announcements
                    SET is_archived = TRUE, is_published = FALSE, display_order = 0,
                        archived_at = NOW(), updated_by = %s, updated_at = NOW()
                    WHERE id = %s AND is_archived = FALSE
                    RETURNING id
                """, (g.admin_user_id, notice_id))
                if not cur.fetchone():
                    return jsonify({'ok': False, 'error': '공지를 찾을 수 없습니다'}), 404

        return jsonify({'ok': True})
    finally:
        release_db_conn(conn)


@app.route('/admin/api/notices/<int:notice_id>/restore', methods=['POST'])
@admin_required
@admin_csrf_required
def api_notices_restore(notice_id):
    conn = get_db_conn()
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE announcements
                    SET is_archived = FALSE, is_published = FALSE, display_order = 0,
                        archived_at = NULL, updated_by = %s, updated_at = NOW()
                    WHERE id = %s AND is_archived = TRUE
                    RETURNING id
                """, (g.admin_user_id, notice_id))
                if not cur.fetchone():
                    return jsonify({'ok': False, 'error': '공지를 찾을 수 없습니다'}), 404

        return jsonify({'ok': True})
    finally:
        release_db_conn(conn)


@app.route('/admin/api/notices/<int:notice_id>', methods=['DELETE'])
@admin_required
@admin_csrf_required
def api_notices_hard_delete(notice_id):
    conn = get_db_conn()
    try:
        with conn:
            with conn.cursor() as cur:
                # 보관함에 있는 항목만 완전 삭제 가능
                cur.execute("""
                    DELETE FROM announcements
                    WHERE id = %s AND is_archived = TRUE
                    RETURNING id
                """, (notice_id,))
                if not cur.fetchone():
                    return jsonify({'ok': False, 'error': '보관함에 있는 공지만 완전 삭제할 수 있습니다'}), 404

        return jsonify({'ok': True})
    finally:
        release_db_conn(conn)


# ── 1:1 문의 (S5-8) ──────────────────────────────────────

def _send_inquiry_reply_email(to_email: str, inquiry_title: str, reply_body: str) -> bool:
    """답변 이메일 발송 (현재: Gmail SMTP 임시 구현).

    # TODO: SES 도메인 인증(slide-atlas.net) 완료 후 boto3 SES로 교체.
    #   교체 시 이 함수만 수정하면 됨 — 호출부는 변경 없음.
    #   SES 발신 주소: noreply@slide-atlas.net (인증 후 확정).
    #   자격증명은 환경변수 AWS_ACCESS_KEY_ID / AWS_SECRET_ACCESS_KEY / AWS_REGION.
    #   (하드코딩 금지 — .env 또는 Render 환경변수에만 설정)
    """
    try:
        import smtplib
        import re as _re
        import html as _html
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText as _MIMEText
        # GMAIL_USER / GMAIL_APP_PW 는 .env 또는 Render 환경변수에서만 읽음 (하드코딩 금지)
        sender = os.environ.get('GMAIL_USER', '')
        if not sender:
            return False
        # [2-2#3] 헤더 주입 방지: Subject/To에서 개행·제어문자 제거(거부). To에 CR/LF 있으면 발송 거부.
        if _re.search(r'[\r\n]', to_email or ''):
            print('[S5-8] 답변 이메일: 수신 주소에 개행 포함 — 발송 거부')
            return False
        safe_subject_title = _re.sub(r'[\r\n\t]+', ' ', inquiry_title or '').strip()[:150]
        # [2-2#3] HTML 주입 방지: 사용자 입력(제목·본문) escaping.
        esc_title = _html.escape(inquiry_title or '')
        esc_body  = _html.escape(reply_body or '')
        msg = MIMEMultipart('alternative')
        msg['Subject'] = f'[SlideAtlas] "{safe_subject_title}" 문의 답변'
        msg['From']    = sender
        msg['To']      = to_email
        html = f"""<p>안녕하세요, SlideAtlas입니다.</p>
<p><b>'{esc_title}'</b> 문의에 대한 답변입니다.</p>
<hr style="border:none;border-top:1px solid #eee;margin:16px 0">
<p style="white-space:pre-wrap">{esc_body}</p>
<p style="color:#888;font-size:12px;margin-top:24px">SlideAtlas | atlaslab.co.kr</p>"""
        msg.attach(_MIMEText(html, 'html'))
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as s:
            s.login(sender, os.environ['GMAIL_APP_PW'])
            s.send_message(msg)
        return True
    except Exception as e:
        print(f'[S5-8] 답변 이메일 발송 실패: {e}')
        return False


@app.route('/admin/inquiries')
@admin_required
def admin_inquiries_page():
    return render_template('admin/inquiries.html', active_page='inquiry')


@app.route('/admin/api/inquiries/list', methods=['GET'])
@admin_required
def api_inquiries_list():
    status_filter = request.args.get('status', 'all')   # 'all'|'open'|'answered'
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            where = ""
            params: list = []
            if status_filter in ('open', 'answered'):
                where = "WHERE i.status = %s"
                params = [status_filter]

            cur.execute(f"""
                SELECT i.id, i.title, i.status,
                       i.user_email, i.user_name,
                       i.institution_id, inst.name_ko AS inst_name,
                       i.created_at
                FROM inquiries i
                LEFT JOIN institutions inst ON i.institution_id = inst.id
                {where}
                ORDER BY
                    CASE WHEN i.status = 'open' THEN 0 ELSE 1 END,
                    i.created_at DESC
            """, params)
            rows = cur.fetchall()

            # 미답변 수
            cur.execute("SELECT COUNT(*) FROM inquiries WHERE status = 'open'")
            open_count = cur.fetchone()[0] or 0

        items = [{
            'id':         r[0],
            'title':      r[1] or '(제목 없음)',
            'status':     r[2],
            'email':      r[3] or '',
            'name':       r[4] or '',
            'inst_id':    r[5] or '',
            'inst_name':  r[6] or '',
            'created_at': r[7].strftime('%m-%d') if r[7] else '',
        } for r in rows]

        return jsonify({'ok': True, 'items': items, 'open_count': open_count})
    finally:
        release_db_conn(conn)


@app.route('/admin/api/inquiries/<int:inq_id>', methods=['GET'])
@admin_required
def api_inquiries_detail(inq_id):
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT i.id, i.title, i.body, i.status,
                       i.user_email, i.user_name,
                       i.institution_id, inst.name_ko AS inst_name,
                       i.created_at
                FROM inquiries i
                LEFT JOIN institutions inst ON i.institution_id = inst.id
                WHERE i.id = %s
            """, (inq_id,))
            row = cur.fetchone()
            if not row:
                return jsonify({'ok': False, 'error': '문의를 찾을 수 없습니다'}), 404

            cur.execute("""
                SELECT r.id, r.body, r.sent_via_ses, r.created_at,
                       a.name, a.role
                FROM inquiry_replies r
                LEFT JOIN admin_users a ON r.created_by = a.id
                WHERE r.inquiry_id = %s
                ORDER BY r.created_at ASC
            """, (inq_id,))
            reply_rows = cur.fetchall()

        inq = {
            'id':        row[0],
            'title':     row[1] or '(제목 없음)',
            'body':      row[2] or '',
            'status':    row[3],
            'email':     row[4] or '',
            'name':      row[5] or '',
            'inst_id':   row[6] or '',
            'inst_name': row[7] or '',
            'created_at': row[8].strftime('%Y-%m-%d %H:%M') if row[8] else '',
        }
        replies = [{
            'id':           r[0],
            'body':         r[1],
            'sent_via_ses': r[2],
            'created_at':   r[3].strftime('%Y-%m-%d %H:%M') if r[3] else '',
            'author':       r[4] or '',
            'author_role':  'super' if r[5] == 'super_admin' else 'staff',
        } for r in reply_rows]

        return jsonify({'ok': True, 'inquiry': inq, 'replies': replies})
    finally:
        release_db_conn(conn)


@app.route('/admin/api/inquiries/<int:inq_id>/reply', methods=['POST'])
@admin_required
@admin_csrf_required
def api_inquiries_reply(inq_id):
    data = request.get_json(silent=True) or {}
    body = (data.get('body') or '').strip()
    if not body:
        return jsonify({'ok': False, 'error': '답변 내용을 입력하세요'}), 400

    conn = get_db_conn()
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT title, user_email FROM inquiries WHERE id = %s
                """, (inq_id,))
                row = cur.fetchone()
                if not row:
                    return jsonify({'ok': False, 'error': '문의를 찾을 수 없습니다'}), 404

                title, to_email = row

                # 이메일 발송. 답변 레코드는 감사 목적상 항상 기록(sent_via_ses에 결과 보존).
                sent = _send_inquiry_reply_email(to_email or '', title or '', body)

                cur.execute("""
                    INSERT INTO inquiry_replies
                        (inquiry_id, body, created_by, sent_via_ses)
                    VALUES (%s, %s, %s, %s)
                    RETURNING id
                """, (inq_id, body, g.admin_user_id, sent))
                reply_id = cur.fetchone()[0]

                # [2-2#3] 메일 발송 성공 시에만 answered로 전환. 실패 시 open 유지(조용한 실패 방지)
                #   → 운영자가 재발송하도록 경고 반환.
                if sent:
                    cur.execute("""
                        UPDATE inquiries SET status = 'answered' WHERE id = %s
                    """, (inq_id,))

        if not sent:
            print(f'[S5-8] 문의 #{inq_id} 답변 메일 발송 실패 — 상태 open 유지(answered 미전환)')
            return jsonify({
                'ok': True, 'reply_id': reply_id, 'sent_via_ses': False,
                'warning': '답변은 기록됐으나 메일 발송에 실패했습니다. 상태를 미답변으로 유지합니다.',
            })
        return jsonify({'ok': True, 'reply_id': reply_id, 'sent_via_ses': True})
    finally:
        release_db_conn(conn)


# ── 고객센터 ──────────────────────────────────────────────

@app.route('/faq')
def faq():
    is_logged = bool(request.cookies.get('access_token'))
    return render_template('faq.html', is_logged=is_logged)


@app.route('/manual')
def manual():
    is_logged = bool(request.cookies.get('access_token'))
    return render_template('manual.html', is_logged=is_logged)


@app.route('/support', methods=['GET', 'POST'])
def support():
    is_logged = bool(request.cookies.get('access_token'))

    if request.method == 'GET':
        return render_template('support.html', is_logged=is_logged,
                               success=False, error=None, form={})

    # POST: 문의 접수
    name    = request.form.get('name', '').strip()
    email   = request.form.get('email', '').strip()
    phone   = request.form.get('phone', '').strip()
    subject = request.form.get('subject', '').strip()
    content = request.form.get('content', '').strip()
    privacy = request.form.get('privacy_agreed', '')

    form_data = {'name': name, 'email': email, 'phone': phone,
                 'subject': subject, 'content': content,
                 'privacy_agreed': bool(privacy)}

    # 서버측 필수값 검증
    if not email or not subject or not content or not privacy:
        return render_template('support.html', is_logged=is_logged,
                               success=False,
                               error='필수 항목을 모두 입력해 주세요.',
                               form=form_data)

    # [H2] 로그인 사용자면 user_id·institution_id 자동 캡처(§15-10), 비로그인이면 NULL.
    # access_token 쿠키를 best-effort 디코드(인증 게이트 아님 — 익명 접수도 허용).
    user_id = None
    inst_id = None
    if is_logged:
        try:
            from auth.decorators import decode_token, COOKIE_NAME
            _payload = decode_token(request.cookies.get(COOKIE_NAME, ''))
            user_id = _payload.get('sub')
            inst_id = _payload.get('institution_id')
        except Exception:
            user_id = None
            inst_id = None

    # [H2] inquiries 실제 스키마 컬럼에 매핑: subject→title, content→body,
    #      email→user_email, name→user_name, status='open'.
    #      phone·privacy_agreed는 스키마에 컬럼이 없어 저장하지 않는다(아래 CEO 승인 항목 참조).
    #      ★실패 은폐(except: pass) 제거 — INSERT 실패 시 사용자에게 실패를 알리고 서버 로그 기록.
    inserted = False
    try:
        conn = get_db_conn()
        try:
            with conn:
                with conn.cursor() as cur:
                    cur.execute("""
                        INSERT INTO inquiries
                            (user_id, institution_id, title, body,
                             user_email, user_name, status)
                        VALUES (%s, %s, %s, %s, %s, %s, 'open')
                    """, (user_id, inst_id, subject, content, email, name or None))
            inserted = True
        finally:
            release_db_conn(conn)
    except Exception:
        app.logger.exception("1:1 문의 INSERT 실패 (email=%s)", email)
        inserted = False

    if not inserted:
        return render_template('support.html', is_logged=is_logged,
                               success=False,
                               error='문의 접수에 실패했습니다. 잠시 후 다시 시도해 주세요.',
                               form=form_data)

    return render_template('support.html', is_logged=is_logged,
                           success=True, error=None, form={})


if __name__ == '__main__':
    print(f"\n✅ SlideAtlas 서버 시작!")
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, threaded=True)
